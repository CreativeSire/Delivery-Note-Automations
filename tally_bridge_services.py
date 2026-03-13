from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
import html
import re
from shutil import copy2
from io import BytesIO
import mimetypes
from pathlib import Path
from urllib import error as urlerror
from urllib import request as urlrequest
from uuid import uuid4

from flask import current_app
from sqlalchemy import func, select
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

from models import SalesOrderRun, TallyBridgeProfile, TallyBridgeRun, TallyDiagnosticsArtifact, TallyDiagnosticsRun, db, utcnow
from services import ServiceError, _string_value
from workflow_services import WorkflowError, create_sku_automator_run, export_sales_order_run_to_workbook

try:
    import xlrd  # type: ignore
except Exception:  # pragma: no cover
    xlrd = None

CONNECTION_MODE_OPTIONS = [
    ("manual_fallback", "Manual fallback"),
    ("file_drop", "File drop"),
    ("hybrid", "Hybrid"),
    ("xml_http", "XML / HTTP"),
]

YES_NO_UNKNOWN_OPTIONS = [
    ("unknown", "Unknown"),
    ("yes", "Yes"),
    ("no", "No"),
]

CASE_STATUS_OPTIONS = [
    ("missing", "Missing"),
    ("uploaded", "Uploaded"),
    ("reviewed", "Reviewed"),
    ("linked", "Linked"),
    ("unlinked", "Unlinked"),
]

RUN_STATUS_OPTIONS = [
    ("draft", "Draft"),
    ("evidence_uploaded", "Evidence uploaded"),
    ("manual_case_ready", "Manual case ready"),
    ("upload_case_ready", "Upload case ready"),
    ("compared", "Compared"),
    ("bridge_mode_decided", "Bridge mode decided"),
]

BRIDGE_RUN_STATUS_OPTIONS = [
    ("ready_to_send", "Ready to send"),
    ("queued_outbound", "Queued outbound"),
    ("staged_for_tally", "Staged for Tally"),
    ("sent_to_tally", "Sent to Tally"),
    ("confirmed_in_tally", "Confirmed in Tally"),
    ("register_received", "Register received"),
    ("linked_to_sku_automator", "Linked to SKU Automator"),
    ("needs_attention", "Needs attention"),
    ("failed", "Failed"),
]

ARTIFACT_GROUP_OPTIONS = [
    ("manual_linked", "Manual linked case"),
    ("uploaded_unlinked", "Uploaded unlinked case"),
    ("environment", "Environment / settings"),
    ("other", "Other"),
]

ARTIFACT_TYPE_OPTIONS = [
    ("sales_order", "Sales Order voucher"),
    ("delivery_note", "Delivery Note voucher"),
    ("sales_invoice", "Sales Invoice voucher"),
    ("register_export", "Register export"),
    ("xml_dump", "XML dump"),
    ("settings_export", "Settings / config"),
    ("screenshot", "Screenshot"),
    ("other", "Other"),
]

REFERENCE_PREFIX_PATTERN = re.compile(r"\b(?:BP|VT|NV)-(\d{6,})\b", re.IGNORECASE)
REFERENCE_RAW_PATTERN = re.compile(r"\b\d{6,}\b")


@dataclass
class TallyBridgeSummary:
    profiles: list[TallyBridgeProfile]
    active_profile: TallyBridgeProfile | None
    latest_run: TallyDiagnosticsRun | None
    run_count: int
    open_run_count: int
    artifact_count: int
    recent_runs: list[TallyDiagnosticsRun]
    outbound_run_count: int
    outbound_open_count: int
    clear_outbound_count: int
    warning_outbound_count: int
    blocked_outbound_count: int
    outbound_guard_filter: str
    ready_pipeline_count: int
    sent_pipeline_count: int
    confirmed_pipeline_count: int
    register_pipeline_count: int
    linked_pipeline_count: int
    blocked_outbound_runs: list["TallyBridgeQueueItem"]
    recent_outbound_runs: list["TallyBridgeQueueItem"]


@dataclass
class TallyDiagnosticsArtifactGroup:
    code: str
    label: str
    artifacts: list[TallyDiagnosticsArtifact]


@dataclass
class TallyDiagnosticsDetail:
    run: TallyDiagnosticsRun
    artifact_groups: list[TallyDiagnosticsArtifactGroup]
    artifact_count: int
    recommended_mode_label: str
    link_integrity: "TallyLinkIntegritySummary"


@dataclass
class TallyLinkIntegrityCase:
    code: str
    label: str
    artifact_count: int
    sales_order_count: int
    delivery_note_count: int
    sales_invoice_count: int
    analyzable_count: int
    shared_all_three: list[str]
    shared_so_to_dn: list[str]
    shared_dn_to_si: list[str]
    shared_so_to_si: list[str]
    verdict: str
    status: str


@dataclass
class TallyLinkIntegritySummary:
    manual_case: TallyLinkIntegrityCase
    upload_case: TallyLinkIntegrityCase
    comparison_verdict: str


@dataclass
class TallyBridgeRunDetail:
    run: TallyBridgeRun
    recommended_mode_label: str
    payload_exists: bool
    staged_exists: bool
    register_exists: bool
    endpoint_response_exists: bool
    link_guard: "TallyBridgeLinkGuard"
    auto_fetch_allowed: bool
    direct_send_allowed: bool


@dataclass
class TallyBridgeLinkGuard:
    status: str
    title: str
    message: str
    diagnostics_run_id: str | None
    diagnostics_title: str | None


@dataclass
class TallyBridgeQueueItem:
    run: TallyBridgeRun
    link_guard: TallyBridgeLinkGuard
    pipeline_stage: str
    pipeline_stage_label: str


def build_tally_bridge_summary(*, guard_filter: str = "all") -> TallyBridgeSummary:
    profiles = list(db.session.scalars(select(TallyBridgeProfile).order_by(TallyBridgeProfile.is_active.desc(), TallyBridgeProfile.name.asc())))
    active_profile = next((profile for profile in profiles if profile.is_active), profiles[0] if profiles else None)
    latest_run = db.session.scalar(select(TallyDiagnosticsRun).order_by(TallyDiagnosticsRun.created_at.desc()).limit(1))
    run_count = db.session.scalar(select(func.count(TallyDiagnosticsRun.id))) or 0
    open_run_count = db.session.scalar(
        select(func.count(TallyDiagnosticsRun.id)).where(TallyDiagnosticsRun.status.not_in(["bridge_mode_decided"]))
    ) or 0
    artifact_count = db.session.scalar(select(func.count(TallyDiagnosticsArtifact.id))) or 0
    recent_runs = list(db.session.scalars(select(TallyDiagnosticsRun).order_by(TallyDiagnosticsRun.created_at.desc()).limit(6)))
    outbound_run_count = db.session.scalar(select(func.count(TallyBridgeRun.id))) or 0
    outbound_open_count = db.session.scalar(
        select(func.count(TallyBridgeRun.id)).where(TallyBridgeRun.status.not_in(["confirmed_in_tally", "failed"]))
    ) or 0
    outbound_runs = list(db.session.scalars(select(TallyBridgeRun).order_by(TallyBridgeRun.created_at.desc()).limit(24)))
    queue_items: list[TallyBridgeQueueItem] = []
    for run in outbound_runs:
        pipeline_stage = _derive_tally_bridge_pipeline_stage(run)
        queue_items.append(
            TallyBridgeQueueItem(
                run=run,
                link_guard=resolve_tally_bridge_link_guard(run.profile_id),
                pipeline_stage=pipeline_stage,
                pipeline_stage_label=_tally_bridge_pipeline_stage_label(pipeline_stage),
            )
        )
    clear_outbound_runs = [item for item in queue_items if item.link_guard.status == "clear"]
    warning_outbound_runs = [item for item in queue_items if item.link_guard.status == "warning"]
    blocked_outbound_runs = [item for item in queue_items if item.link_guard.status == "blocked"]
    selected_filter = guard_filter if guard_filter in {"all", "clear", "warning", "blocked"} else "all"
    if selected_filter == "clear":
        recent_outbound_runs = clear_outbound_runs
    elif selected_filter == "warning":
        recent_outbound_runs = warning_outbound_runs
    elif selected_filter == "blocked":
        recent_outbound_runs = blocked_outbound_runs
    else:
        recent_outbound_runs = queue_items
    ready_pipeline_count = sum(1 for item in queue_items if item.pipeline_stage == "ready")
    sent_pipeline_count = sum(1 for item in queue_items if item.pipeline_stage == "sent")
    confirmed_pipeline_count = sum(1 for item in queue_items if item.pipeline_stage == "confirmed")
    register_pipeline_count = sum(1 for item in queue_items if item.pipeline_stage == "register_received")
    linked_pipeline_count = sum(1 for item in queue_items if item.pipeline_stage == "linked")
    return TallyBridgeSummary(
        profiles=profiles,
        active_profile=active_profile,
        latest_run=latest_run,
        run_count=run_count,
        open_run_count=open_run_count,
        artifact_count=artifact_count,
        recent_runs=recent_runs,
        outbound_run_count=outbound_run_count,
        outbound_open_count=outbound_open_count,
        clear_outbound_count=len(clear_outbound_runs),
        warning_outbound_count=len(warning_outbound_runs),
        blocked_outbound_count=len(blocked_outbound_runs),
        outbound_guard_filter=selected_filter,
        ready_pipeline_count=ready_pipeline_count,
        sent_pipeline_count=sent_pipeline_count,
        confirmed_pipeline_count=confirmed_pipeline_count,
        register_pipeline_count=register_pipeline_count,
        linked_pipeline_count=linked_pipeline_count,
        blocked_outbound_runs=blocked_outbound_runs,
        recent_outbound_runs=recent_outbound_runs,
    )


def build_tally_diagnostics_detail(run_id: str) -> TallyDiagnosticsDetail | None:
    run = db.session.get(TallyDiagnosticsRun, run_id)
    if run is None:
        return None

    artifacts = list(
        db.session.scalars(
            select(TallyDiagnosticsArtifact)
            .where(TallyDiagnosticsArtifact.run_id == run_id)
            .order_by(TallyDiagnosticsArtifact.created_at.desc(), TallyDiagnosticsArtifact.id.desc())
        )
    )
    groups: list[TallyDiagnosticsArtifactGroup] = []
    for code, label in ARTIFACT_GROUP_OPTIONS:
        groups.append(
            TallyDiagnosticsArtifactGroup(
                code=code,
                label=label,
                artifacts=[artifact for artifact in artifacts if artifact.artifact_group == code],
            )
        )
    grouped_ids = {artifact.id for group in groups for artifact in group.artifacts}
    other_artifacts = [artifact for artifact in artifacts if artifact.id not in grouped_ids]
    if other_artifacts:
        groups.append(TallyDiagnosticsArtifactGroup(code="unclassified", label="Unclassified", artifacts=other_artifacts))

    recommended_mode = derive_tally_bridge_mode(run)
    link_integrity = build_tally_link_integrity_summary(artifacts)
    return TallyDiagnosticsDetail(
        run=run,
        artifact_groups=groups,
        artifact_count=len(artifacts),
        recommended_mode_label=_mode_label(recommended_mode),
        link_integrity=link_integrity,
    )


def get_tally_bridge_profile(profile_id: int) -> TallyBridgeProfile | None:
    return db.session.get(TallyBridgeProfile, profile_id)


def get_tally_diagnostics_run(run_id: str) -> TallyDiagnosticsRun | None:
    return db.session.get(TallyDiagnosticsRun, run_id)


def get_tally_diagnostics_artifact(artifact_id: int) -> TallyDiagnosticsArtifact | None:
    return db.session.get(TallyDiagnosticsArtifact, artifact_id)


def get_tally_bridge_run(run_id: str) -> TallyBridgeRun | None:
    return db.session.get(TallyBridgeRun, run_id)


def build_tally_bridge_run_detail(run_id: str) -> TallyBridgeRunDetail | None:
    run = db.session.get(TallyBridgeRun, run_id)
    if run is None:
        return None
    payload_path = Path(run.payload_storage_path)
    staged_path = Path(run.staged_storage_path) if run.staged_storage_path else None
    endpoint_response_path = Path(run.endpoint_response_storage_path) if run.endpoint_response_storage_path else None
    link_guard = resolve_tally_bridge_link_guard(run.profile_id)
    direct_send_allowed = bool(
        run.profile
        and run.profile.connection_mode in {"xml_http", "hybrid"}
        and _is_http_endpoint(run.profile.endpoint_url)
        and link_guard.status == "clear"
        and (run.profile.capabilities_json or {}).get("probe_status") == "success"
    )
    return TallyBridgeRunDetail(
        run=run,
        recommended_mode_label=_mode_label(run.bridge_mode),
        payload_exists=payload_path.exists(),
        staged_exists=bool(staged_path and staged_path.exists()),
        register_exists=bool(run.register_storage_path and Path(run.register_storage_path).exists()),
        endpoint_response_exists=bool(endpoint_response_path and endpoint_response_path.exists()),
        link_guard=link_guard,
        auto_fetch_allowed=bool(run.profile and run.profile.connection_mode in {"file_drop", "hybrid"} and link_guard.status == "clear"),
        direct_send_allowed=direct_send_allowed,
    )


def save_tally_bridge_profile(values: dict[str, str], *, profile_id: int | None = None) -> TallyBridgeProfile:
    profile = db.session.get(TallyBridgeProfile, profile_id) if profile_id else None
    if profile is None:
        profile = TallyBridgeProfile()
        db.session.add(profile)

    name = _string_value(values.get("name"))
    if not name:
        raise ServiceError("Please give this Tally profile a name first.")

    connection_mode = _choice_value(values.get("connection_mode"), CONNECTION_MODE_OPTIONS, "manual_fallback")
    is_active = _string_value(values.get("is_active")).lower() != "no"

    profile.name = name
    profile.connection_mode = connection_mode
    profile.company_name = _nullable_text(values.get("company_name"))
    profile.tally_version = _nullable_text(values.get("tally_version"))
    profile.endpoint_url = _nullable_text(values.get("endpoint_url"))
    profile.machine_name = _nullable_text(values.get("machine_name"))
    profile.notes = _nullable_text(values.get("notes"))
    profile.is_active = is_active

    existing_capabilities = dict(profile.capabilities_json or {})
    profile.capabilities_json = {
        **existing_capabilities,
        "xml_http": _choice_value(values.get("profile_xml_http"), YES_NO_UNKNOWN_OPTIONS, "unknown"),
        "outbound_import": _choice_value(values.get("profile_outbound_import"), YES_NO_UNKNOWN_OPTIONS, "unknown"),
        "register_fetch": _choice_value(values.get("profile_register_fetch"), YES_NO_UNKNOWN_OPTIONS, "unknown"),
    }
    if any(value != "unknown" for value in profile.capabilities_json.values()):
        profile.last_checked_at = utcnow()

    if is_active:
        db.session.query(TallyBridgeProfile).filter(TallyBridgeProfile.id != profile.id).update(
            {"is_active": False}, synchronize_session=False
        )

    db.session.commit()
    return profile


def probe_tally_bridge_profile(profile_id: int) -> TallyBridgeProfile:
    profile = db.session.get(TallyBridgeProfile, profile_id)
    if profile is None:
        raise ServiceError("This Tally Bridge profile could not be found.")

    endpoint = _nullable_text(profile.endpoint_url)
    if not endpoint:
        raise ServiceError("Add the Tally XML / HTTP endpoint first.")
    if not endpoint.lower().startswith(("http://", "https://")):
        raise ServiceError("The Tally XML / HTTP probe needs an http:// or https:// endpoint.")

    payload = _build_tally_probe_envelope(profile.company_name)
    request = urlrequest.Request(
        endpoint,
        data=payload.encode("utf-8"),
        method="POST",
        headers={
            "Content-Type": "text/xml; charset=utf-8",
            "Accept": "text/xml, application/xml",
            "User-Agent": "DALA-Tally-Bridge/1.0",
        },
    )
    checked_at = utcnow()
    capabilities = dict(profile.capabilities_json or {})

    try:
        with urlrequest.urlopen(request, timeout=12) as response:
            response_bytes = response.read(4096)
            http_status = response.getcode() or 200
        response_text = response_bytes.decode("utf-8", errors="replace")
        xml_ok = _looks_like_tally_xml_response(response_text)
        capabilities["xml_http"] = "yes" if xml_ok else capabilities.get("xml_http", "unknown")
        capabilities["probe_status"] = "success" if xml_ok else "unexpected_response"
        capabilities["probe_http_status"] = http_status
        capabilities["probe_message"] = (
            "Tally XML / HTTP endpoint responded with a Tally-style XML envelope."
            if xml_ok
            else "Endpoint responded, but the payload did not look like a Tally XML response."
        )
        capabilities["probe_excerpt"] = _compact_probe_excerpt(response_text)
    except urlerror.HTTPError as exc:
        response_text = exc.read(4096).decode("utf-8", errors="replace")
        capabilities["xml_http"] = "no"
        capabilities["probe_status"] = "http_error"
        capabilities["probe_http_status"] = exc.code
        capabilities["probe_message"] = f"HTTP {exc.code} returned from the configured Tally endpoint."
        capabilities["probe_excerpt"] = _compact_probe_excerpt(response_text)
    except urlerror.URLError as exc:
        capabilities["xml_http"] = "no"
        capabilities["probe_status"] = "connection_error"
        capabilities["probe_http_status"] = None
        capabilities["probe_message"] = f"Could not reach the configured Tally endpoint: {exc.reason}"
        capabilities["probe_excerpt"] = ""

    capabilities["probe_checked_at"] = checked_at.isoformat()
    profile.capabilities_json = capabilities
    profile.last_checked_at = checked_at
    db.session.commit()
    return profile


def create_tally_diagnostics_run(values: dict[str, str]) -> TallyDiagnosticsRun:
    title = _string_value(values.get("title"))
    if not title:
        raise ServiceError("Please name this Tally diagnostics run first.")

    profile = _resolve_profile(values.get("profile_id"))
    run = TallyDiagnosticsRun(
        id=uuid4().hex,
        profile_id=profile.id if profile is not None else None,
        title=title,
        status="draft",
        notes=_nullable_text(values.get("notes")),
    )
    run.recommended_mode = derive_tally_bridge_mode(run)
    db.session.add(run)
    db.session.commit()
    return run


def update_tally_diagnostics_run(run_id: str, values: dict[str, str]) -> TallyDiagnosticsRun:
    run = db.session.get(TallyDiagnosticsRun, run_id)
    if run is None:
        raise ServiceError("This Tally diagnostics run could not be found.")

    run.status = _choice_value(values.get("status"), RUN_STATUS_OPTIONS, run.status)
    run.xml_http_supported = _choice_value(values.get("xml_http_supported"), YES_NO_UNKNOWN_OPTIONS, run.xml_http_supported)
    run.outbound_import_supported = _choice_value(
        values.get("outbound_import_supported"), YES_NO_UNKNOWN_OPTIONS, run.outbound_import_supported
    )
    run.register_fetch_supported = _choice_value(
        values.get("register_fetch_supported"), YES_NO_UNKNOWN_OPTIONS, run.register_fetch_supported
    )
    run.dn_link_supported = _choice_value(values.get("dn_link_supported"), YES_NO_UNKNOWN_OPTIONS, run.dn_link_supported)
    run.manual_case_status = _choice_value(values.get("manual_case_status"), CASE_STATUS_OPTIONS, run.manual_case_status)
    run.uploaded_case_status = _choice_value(
        values.get("uploaded_case_status"), CASE_STATUS_OPTIONS, run.uploaded_case_status
    )
    run.findings_summary = _nullable_text(values.get("findings_summary"))
    run.notes = _nullable_text(values.get("notes"))
    run.recommended_mode = derive_tally_bridge_mode(run)
    if run.status == "bridge_mode_decided":
        run.completed_at = utcnow()

    db.session.commit()
    return run


def add_tally_diagnostics_artifact(
    run_id: str,
    *,
    file_storage: object,
    artifact_group: str,
    artifact_type: str,
    description: str | None = None,
) -> TallyDiagnosticsArtifact:
    run = db.session.get(TallyDiagnosticsRun, run_id)
    if run is None:
        raise ServiceError("This Tally diagnostics run could not be found.")

    filename = _string_value(getattr(file_storage, "filename", ""))
    if not filename:
        raise ServiceError("Please choose a Tally artifact file first.")

    payload = file_storage.read()
    if hasattr(file_storage, "stream"):
        file_storage.stream.seek(0)
    if not payload:
        raise ServiceError("The selected artifact file is empty.")

    group_code = _choice_value(artifact_group, ARTIFACT_GROUP_OPTIONS, "other")
    type_code = _choice_value(artifact_type, ARTIFACT_TYPE_OPTIONS, "other")
    safe_name = secure_filename(filename) or f"artifact-{uuid4().hex}"
    stored_name = f"{uuid4().hex}-{safe_name}"
    target_dir = Path(current_app.instance_path) / "tally_bridge" / run.id
    target_dir.mkdir(parents=True, exist_ok=True)
    target_path = target_dir / stored_name
    target_path.write_bytes(payload)

    artifact = TallyDiagnosticsArtifact(
        run_id=run.id,
        artifact_group=group_code,
        artifact_type=type_code,
        filename=filename,
        content_type=_string_value(getattr(file_storage, "mimetype", "")) or None,
        storage_path=str(target_path),
        description=_nullable_text(description),
        file_size=len(payload),
    )
    db.session.add(artifact)

    if group_code == "manual_linked" and run.manual_case_status == "missing":
        run.manual_case_status = "uploaded"
    if group_code == "uploaded_unlinked" and run.uploaded_case_status == "missing":
        run.uploaded_case_status = "uploaded"
    if run.status == "draft":
        run.status = "evidence_uploaded"

    db.session.commit()
    return artifact


def create_tally_bridge_run_from_sales_order(
    sales_order_run_id: str,
    *,
    profile_id: int | None = None,
    notes: str | None = None,
) -> TallyBridgeRun:
    sales_order_run = db.session.get(SalesOrderRun, sales_order_run_id)
    if sales_order_run is None:
        raise ServiceError("This Sales Order run could not be found.")
    if sales_order_run.rows_needing_review > 0:
        raise ServiceError("Resolve all Sales Order review items before sending this run into Tally Bridge.")
    if sales_order_run.rows_ready <= 0:
        raise ServiceError("This Sales Order run has no ready rows to send into Tally Bridge.")

    profile = db.session.get(TallyBridgeProfile, profile_id) if profile_id else _resolve_profile(None)
    bridge_mode = profile.connection_mode if profile is not None else "manual_fallback"
    link_guard = resolve_tally_bridge_link_guard(profile.id if profile is not None else None)

    try:
        payload_filename, payload = export_sales_order_run_to_workbook(sales_order_run_id)
    except WorkflowError as exc:
        raise ServiceError(str(exc)) from exc

    run_id = uuid4().hex
    target_dir = _bridge_storage_dir("outbound", run_id)
    safe_name = secure_filename(payload_filename) or f"sales-order-{run_id}.xlsx"
    payload_path = target_dir / safe_name
    payload_path.write_bytes(payload)

    run = TallyBridgeRun(
        id=run_id,
        profile_id=profile.id if profile is not None else None,
        sales_order_run_id=sales_order_run.id,
        status="needs_attention" if link_guard.status == "blocked" else "ready_to_send",
        bridge_mode=bridge_mode,
        payload_filename=payload_filename,
        payload_storage_path=str(payload_path),
        payload_content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        rows_ready=sales_order_run.rows_ready,
        notes=_nullable_text(notes),
        error_message=link_guard.message if link_guard.status == "blocked" else None,
    )
    db.session.add(run)
    db.session.commit()
    return run


def send_tally_bridge_run_to_endpoint(run_id: str) -> TallyBridgeRun:
    run = db.session.get(TallyBridgeRun, run_id)
    if run is None:
        raise ServiceError("This Tally Bridge run could not be found.")
    assert_tally_bridge_link_guard(run, action_label="send this payload into the Tally endpoint")
    if run.profile is None:
        raise ServiceError("Link this run to a Tally profile before sending it directly.")

    endpoint = _nullable_text(run.profile.endpoint_url)
    if not _is_http_endpoint(endpoint):
        raise ServiceError("The active Tally profile must have an http:// or https:// endpoint for direct send.")

    capabilities = dict(run.profile.capabilities_json or {})
    if capabilities.get("probe_status") != "success":
        raise ServiceError("Run a successful XML / HTTP probe on this Tally profile before direct send.")

    payload_path = Path(run.payload_storage_path)
    if not payload_path.exists():
        raise ServiceError("The stored Sales Order payload for this bridge run is missing.")

    payload_bytes = payload_path.read_bytes()
    request = urlrequest.Request(
        endpoint,
        data=payload_bytes,
        method="POST",
        headers={
            "Content-Type": run.payload_content_type or "application/octet-stream",
            "Accept": "application/xml, text/xml, application/json, text/plain",
            "User-Agent": "DALA-Tally-Bridge/1.0",
            "X-DALA-Bridge-Run-ID": run.id,
            "X-DALA-Source-Run-ID": run.sales_order_run_id,
            "X-DALA-Filename": run.payload_filename,
        },
    )

    sent_at = utcnow()
    try:
        with urlrequest.urlopen(request, timeout=20) as response:
            response_bytes = response.read(4096)
            http_status = response.getcode() or 200
            response_content_type = response.headers.get_content_type() if response.headers else "application/octet-stream"
    except urlerror.HTTPError as exc:
        response_bytes = exc.read(4096)
        response_content_type = exc.headers.get_content_type() if exc.headers else "text/plain"
        _store_tally_endpoint_response(run, response_bytes, response_content_type)
        run.endpoint_http_status = exc.code
        run.status = "needs_attention"
        run.error_message = f"Endpoint returned HTTP {exc.code} during direct send."
        db.session.commit()
        raise ServiceError(run.error_message) from exc
    except urlerror.URLError as exc:
        run.status = "needs_attention"
        run.error_message = f"Could not reach the Tally endpoint during direct send: {exc.reason}"
        db.session.commit()
        raise ServiceError(run.error_message) from exc

    _store_tally_endpoint_response(run, response_bytes, response_content_type)
    run.endpoint_http_status = http_status
    run.sent_at = sent_at
    if _response_confirms_direct_send(response_bytes, response_content_type):
        run.status = "confirmed_in_tally"
        run.confirmed_at = utcnow()
        run.error_message = None
    else:
        run.status = "sent_to_tally"
        run.confirmed_at = None
        run.error_message = None
    db.session.commit()
    return run


def update_tally_bridge_run_status(run_id: str, values: dict[str, str]) -> TallyBridgeRun:
    run = db.session.get(TallyBridgeRun, run_id)
    if run is None:
        raise ServiceError("This Tally Bridge run could not be found.")

    next_status = _choice_value(values.get("status"), BRIDGE_RUN_STATUS_OPTIONS, run.status)
    if next_status in {"sent_to_tally", "confirmed_in_tally"}:
        assert_tally_bridge_link_guard(run, action_label=f"mark this run as {next_status.replace('_', ' ')}")

    run.status = next_status
    run.notes = _nullable_text(values.get("notes"))
    error_message = _nullable_text(values.get("error_message"))
    run.error_message = error_message if run.status in {"needs_attention", "failed"} else None
    if run.status in {"sent_to_tally", "confirmed_in_tally"} and run.sent_at is None:
        run.sent_at = utcnow()
    if run.status == "confirmed_in_tally":
        run.confirmed_at = utcnow()
    elif run.status not in {"confirmed_in_tally"}:
        run.confirmed_at = None if run.status == "failed" else run.confirmed_at

    db.session.commit()
    return run


def stage_tally_bridge_run_to_profile_target(run_id: str) -> TallyBridgeRun:
    run = db.session.get(TallyBridgeRun, run_id)
    if run is None:
        raise ServiceError("This Tally Bridge run could not be found.")
    assert_tally_bridge_link_guard(run, action_label="stage this payload for Tally")
    if run.profile is None:
        raise ServiceError("Link this run to a Tally profile before staging it.")
    destination = _nullable_text(run.profile.endpoint_url)
    if not destination:
        raise ServiceError("This Tally profile does not have a local file-drop target yet.")

    destination_dir = Path(destination)
    if not destination_dir.exists() or not destination_dir.is_dir():
        raise ServiceError("The profile endpoint must point to an existing local folder before staging can work.")

    payload_path = Path(run.payload_storage_path)
    if not payload_path.exists():
        raise ServiceError("The stored Sales Order payload for this bridge run is missing.")

    staged_name = secure_filename(run.payload_filename) or payload_path.name
    staged_path = destination_dir / staged_name
    copy2(payload_path, staged_path)

    run.staged_storage_path = str(staged_path)
    run.status = "staged_for_tally"
    run.sent_at = utcnow()
    run.error_message = None
    db.session.commit()
    return run


def import_tally_register_for_bridge_run(run_id: str, *, file_storage: object) -> TallyBridgeRun:
    run = db.session.get(TallyBridgeRun, run_id)
    if run is None:
        raise ServiceError("This Tally Bridge run could not be found.")
    assert_tally_bridge_link_guard(run, action_label="continue into SKU Automator")

    filename = _string_value(getattr(file_storage, "filename", ""))
    if not filename:
        raise ServiceError("Please choose the returned Tally register file first.")

    payload = file_storage.read()
    if hasattr(file_storage, "stream"):
        file_storage.stream.seek(0)
    if not payload:
        raise ServiceError("The returned Tally register file is empty.")

    return _attach_register_payload_to_bridge_run(
        run,
        payload=payload,
        filename=filename,
        content_type=_string_value(getattr(file_storage, "mimetype", "")) or None,
    )


def pull_tally_register_from_profile_target(run_id: str) -> TallyBridgeRun:
    run = db.session.get(TallyBridgeRun, run_id)
    if run is None:
        raise ServiceError("This Tally Bridge run could not be found.")
    guard = resolve_tally_bridge_link_guard(run.profile_id)
    if guard.status != "clear":
        raise ServiceError("Auto-fetch is only available after the Tally chain is verified as clear in diagnostics.")
    if run.profile is None or run.profile.connection_mode not in {"file_drop", "hybrid"}:
        raise ServiceError("Auto-fetch needs a file-drop or hybrid Tally profile.")

    destination = _nullable_text(run.profile.endpoint_url)
    if not destination:
        raise ServiceError("This Tally profile does not have a watched local folder configured yet.")

    watched_dir = Path(destination)
    if not watched_dir.exists() or not watched_dir.is_dir():
        raise ServiceError("The watched local folder for this Tally profile does not exist.")

    candidate = _select_returned_register_candidate(run, watched_dir)
    if candidate is None:
        raise ServiceError("No returned Tally register file was found in the watched folder yet.")

    content_type = mimetypes.guess_type(candidate.name)[0] or "application/octet-stream"
    return _attach_register_payload_to_bridge_run(
        run,
        payload=candidate.read_bytes(),
        filename=candidate.name,
        content_type=content_type,
    )


def build_tally_link_integrity_summary(artifacts: list[TallyDiagnosticsArtifact]) -> TallyLinkIntegritySummary:
    by_group = {
        "manual_linked": [artifact for artifact in artifacts if artifact.artifact_group == "manual_linked"],
        "uploaded_unlinked": [artifact for artifact in artifacts if artifact.artifact_group == "uploaded_unlinked"],
    }
    manual_case = _analyze_link_integrity_case("manual_linked", "Manual linked case", by_group["manual_linked"])
    upload_case = _analyze_link_integrity_case("uploaded_unlinked", "Uploaded comparison case", by_group["uploaded_unlinked"])

    if manual_case.status == "linked" and upload_case.status == "broken":
        comparison_verdict = "The manual case appears linked, but the uploaded case breaks the reference chain after import."
    elif manual_case.status == "linked" and upload_case.status == "linked":
        comparison_verdict = "Both the manual and uploaded cases appear to preserve the same order reference chain."
    elif manual_case.status == "missing":
        comparison_verdict = "Upload one full manual SO -> DN -> SI sample first so the bridge can compare it against the imported path."
    elif upload_case.status == "missing":
        comparison_verdict = "Upload the imported comparison case so the bridge can prove whether Tally keeps the same chain."
    else:
        comparison_verdict = "The bridge has partial evidence, but the chain verdict is still incomplete. Add or improve the voucher artifacts."

    return TallyLinkIntegritySummary(
        manual_case=manual_case,
        upload_case=upload_case,
        comparison_verdict=comparison_verdict,
    )


def resolve_tally_bridge_link_guard(profile_id: int | None) -> TallyBridgeLinkGuard:
    diagnostics_run = _latest_relevant_diagnostics_run(profile_id)
    if diagnostics_run is None:
        return TallyBridgeLinkGuard(
            status="warning",
            title="No chain diagnostics yet",
            message="No Tally diagnostics run has been recorded for this profile yet. The bridge can still prepare the payload, but chain integrity is not proven.",
            diagnostics_run_id=None,
            diagnostics_title=None,
        )

    if diagnostics_run.status != "bridge_mode_decided":
        return TallyBridgeLinkGuard(
            status="warning",
            title="Diagnostics still open",
            message="The latest Tally diagnostics run has not been finalized yet. Review the evidence before relying on this bridge path.",
            diagnostics_run_id=diagnostics_run.id,
            diagnostics_title=diagnostics_run.title,
        )

    detail = build_tally_diagnostics_detail(diagnostics_run.id)
    if detail is None:
        return TallyBridgeLinkGuard(
            status="warning",
            title="Diagnostics unavailable",
            message="The bridge could not load the latest Tally diagnostics detail, so chain integrity is not verified.",
            diagnostics_run_id=diagnostics_run.id,
            diagnostics_title=diagnostics_run.title,
        )

    if diagnostics_run.dn_link_supported == "no" or detail.link_integrity.upload_case.status == "broken":
        return TallyBridgeLinkGuard(
            status="blocked",
            title="Imported chain is broken",
            message=detail.link_integrity.comparison_verdict,
            diagnostics_run_id=diagnostics_run.id,
            diagnostics_title=diagnostics_run.title,
        )

    if detail.link_integrity.upload_case.status == "linked" and diagnostics_run.dn_link_supported == "yes":
        return TallyBridgeLinkGuard(
            status="clear",
            title="Imported chain verified",
            message=detail.link_integrity.comparison_verdict,
            diagnostics_run_id=diagnostics_run.id,
            diagnostics_title=diagnostics_run.title,
        )

    return TallyBridgeLinkGuard(
        status="warning",
        title="Chain proof is incomplete",
        message=detail.link_integrity.comparison_verdict,
        diagnostics_run_id=diagnostics_run.id,
        diagnostics_title=diagnostics_run.title,
    )


def assert_tally_bridge_link_guard(run: TallyBridgeRun, *, action_label: str) -> None:
    guard = resolve_tally_bridge_link_guard(run.profile_id)
    if guard.status != "blocked":
        return
    raise ServiceError(
        f"Stop before you {action_label}. {guard.message}"
        + (f" Review diagnostics run '{guard.diagnostics_title}' first." if guard.diagnostics_title else "")
    )


def _attach_register_payload_to_bridge_run(
    run: TallyBridgeRun,
    *,
    payload: bytes,
    filename: str,
    content_type: str | None,
) -> TallyBridgeRun:
    target_dir = _bridge_storage_dir("registers", run.id)
    safe_name = secure_filename(filename) or f"register-{run.id}.xlsx"
    register_path = target_dir / f"{uuid4().hex}-{safe_name}"
    register_path.write_bytes(payload)

    inbound_upload = FileStorage(
        stream=BytesIO(payload),
        filename=filename,
        content_type=content_type,
    )
    try:
        sku_run = create_sku_automator_run(inbound_upload)
    except WorkflowError as exc:
        raise ServiceError(str(exc)) from exc

    run.register_filename = filename
    run.register_storage_path = str(register_path)
    run.register_content_type = inbound_upload.content_type
    run.register_received_at = utcnow()
    run.sku_automator_run_id = sku_run.id
    run.status = "linked_to_sku_automator"
    run.error_message = None
    db.session.commit()
    return run


def _store_tally_endpoint_response(run: TallyBridgeRun, response_bytes: bytes, content_type: str | None) -> None:
    target_dir = _bridge_storage_dir("outbound", run.id)
    suffix = _response_file_suffix(content_type)
    response_path = target_dir / f"endpoint-response{suffix}"
    response_path.write_bytes(response_bytes)
    run.endpoint_response_storage_path = str(response_path)
    run.endpoint_response_content_type = content_type or "application/octet-stream"


def _response_file_suffix(content_type: str | None) -> str:
    normalized = (content_type or "").lower()
    if "xml" in normalized:
        return ".xml"
    if "json" in normalized:
        return ".json"
    if "html" in normalized:
        return ".html"
    return ".txt"


def _response_confirms_direct_send(response_bytes: bytes, content_type: str | None) -> bool:
    response_text = response_bytes.decode("utf-8", errors="replace").strip()
    normalized = response_text.lower()
    if not normalized:
        return False
    if "json" in (content_type or "").lower():
        return any(
            token in normalized
            for token in ('"status":"success"', '"status": "success"', '"status":"confirmed"', '"accepted":true')
        )
    if "<lineerror" in normalized:
        return False
    return any(
        token in normalized
        for token in (
            "<created>",
            "<altered>",
            "<lastvchid>",
            "<tallymessage",
            "<status>success</status>",
            "<status>confirmed</status>",
        )
    )


def _is_http_endpoint(endpoint: str | None) -> bool:
    normalized = _string_value(endpoint).lower()
    return normalized.startswith("http://") or normalized.startswith("https://")


def _derive_tally_bridge_pipeline_stage(run: TallyBridgeRun) -> str:
    if run.sku_automator_run_id or run.status == "linked_to_sku_automator":
        return "linked"
    if run.register_received_at or run.status == "register_received":
        return "register_received"
    if run.confirmed_at or run.status == "confirmed_in_tally":
        return "confirmed"
    if run.sent_at or run.status == "sent_to_tally":
        return "sent"
    return "ready"


def _tally_bridge_pipeline_stage_label(stage: str) -> str:
    return {
        "ready": "Ready to send",
        "sent": "Sent to Tally",
        "confirmed": "Confirmed in Tally",
        "register_received": "Register received",
        "linked": "SKU Automator linked",
    }.get(stage, stage.replace("_", " ").title())


def derive_tally_bridge_mode(run: TallyDiagnosticsRun) -> str:
    if (
        run.xml_http_supported == "yes"
        and run.outbound_import_supported == "yes"
        and run.register_fetch_supported == "yes"
        and run.dn_link_supported == "yes"
    ):
        return "xml_http"
    if run.outbound_import_supported == "yes" and run.register_fetch_supported == "yes":
        return "hybrid"
    if run.outbound_import_supported == "yes":
        return "file_drop"
    return "manual_fallback"


def _latest_relevant_diagnostics_run(profile_id: int | None) -> TallyDiagnosticsRun | None:
    query = select(TallyDiagnosticsRun)
    if profile_id is not None:
        query = query.where(TallyDiagnosticsRun.profile_id == profile_id)
    query = query.order_by(TallyDiagnosticsRun.completed_at.desc(), TallyDiagnosticsRun.created_at.desc())
    run = db.session.scalar(query.limit(1))
    if run is not None or profile_id is None:
        return run
    return db.session.scalar(
        select(TallyDiagnosticsRun).order_by(TallyDiagnosticsRun.completed_at.desc(), TallyDiagnosticsRun.created_at.desc()).limit(1)
    )


def _select_returned_register_candidate(run: TallyBridgeRun, watched_dir: Path) -> Path | None:
    allowed_suffixes = {".xlsx", ".xls", ".csv", ".tsv", ".txt"}
    excluded_paths = {
        Path(run.payload_storage_path).resolve(),
    }
    if run.staged_storage_path:
        excluded_paths.add(Path(run.staged_storage_path).resolve())
    if run.register_storage_path:
        excluded_paths.add(Path(run.register_storage_path).resolve())

    threshold = run.sent_at or run.created_at
    if threshold is not None:
        if threshold.tzinfo is None:
            threshold = threshold.replace(tzinfo=UTC)
        else:
            threshold = threshold.astimezone(UTC)
    candidates: list[Path] = []
    for path in watched_dir.iterdir():
        if not path.is_file() or path.suffix.lower() not in allowed_suffixes:
            continue
        try:
            resolved = path.resolve()
        except OSError:
            continue
        if resolved in excluded_paths:
            continue
        modified_at = datetime.fromtimestamp(path.stat().st_mtime, tz=UTC)
        if threshold is not None and modified_at < threshold:
            continue
        candidates.append(path)

    if not candidates:
        return None

    def candidate_rank(path: Path) -> tuple[int, float]:
        name = path.name.lower()
        score = 0
        if "register" in name:
            score += 3
        if "salesorder" in name or "sales-order" in name:
            score += 2
        if "dala" in name:
            score += 1
        return (score, path.stat().st_mtime)

    candidates.sort(key=candidate_rank, reverse=True)
    return candidates[0]


def _resolve_profile(raw_profile_id: str | None) -> TallyBridgeProfile | None:
    text = _string_value(raw_profile_id)
    if text.isdigit():
        profile = db.session.get(TallyBridgeProfile, int(text))
        if profile is not None:
            return profile
    active_profile = db.session.scalar(select(TallyBridgeProfile).where(TallyBridgeProfile.is_active.is_(True)).limit(1))
    if active_profile is not None:
        return active_profile
    return db.session.scalar(select(TallyBridgeProfile).order_by(TallyBridgeProfile.id.asc()).limit(1))


def _build_tally_probe_envelope(company_name: str | None) -> str:
    company_xml = ""
    if company_name:
        company_xml = f"<SVCURRENTCOMPANY>{html.escape(company_name)}</SVCURRENTCOMPANY>"
    return (
        "<ENVELOPE>"
        "<HEADER>"
        "<VERSION>1</VERSION>"
        "<TALLYREQUEST>Export</TALLYREQUEST>"
        "<TYPE>Data</TYPE>"
        "<ID>List of Companies</ID>"
        "</HEADER>"
        "<BODY>"
        "<DESC>"
        "<STATICVARIABLES>"
        "<SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>"
        f"{company_xml}"
        "</STATICVARIABLES>"
        "</DESC>"
        "</BODY>"
        "</ENVELOPE>"
    )


def _looks_like_tally_xml_response(payload: str) -> bool:
    normalized = payload.strip().lower()
    if not normalized:
        return False
    if "<envelope" not in normalized or "<body>" not in normalized:
        return False
    return any(marker in normalized for marker in ("<tallymessage", "<company", "<collection", "<lineerror"))


def _compact_probe_excerpt(payload: str) -> str:
    compact = re.sub(r"\s+", " ", payload).strip()
    return compact[:400]


def _nullable_text(value: object) -> str | None:
    text = _string_value(value)
    return text or None


def _choice_value(value: object, options: list[tuple[str, str]], default: str) -> str:
    selected = _string_value(value)
    allowed = {code for code, _ in options}
    return selected if selected in allowed else default


def _mode_label(code: str) -> str:
    options = dict(CONNECTION_MODE_OPTIONS)
    return options.get(code, code.replace("_", " ").title())


def _bridge_storage_dir(*parts: str) -> Path:
    target_dir = Path(current_app.instance_path) / "tally_bridge"
    for part in parts:
        target_dir = target_dir / part
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir


def _analyze_link_integrity_case(code: str, label: str, artifacts: list[TallyDiagnosticsArtifact]) -> TallyLinkIntegrityCase:
    role_refs: dict[str, set[str]] = {
        "sales_order": set(),
        "delivery_note": set(),
        "sales_invoice": set(),
    }
    analyzable_count = 0
    role_counts = {role: 0 for role in role_refs}

    for artifact in artifacts:
        if artifact.artifact_type not in role_refs:
            continue
        role_counts[artifact.artifact_type] += 1
        refs = _extract_references_from_artifact(artifact)
        if refs:
            analyzable_count += 1
            role_refs[artifact.artifact_type].update(refs)

    so_refs = role_refs["sales_order"]
    dn_refs = role_refs["delivery_note"]
    si_refs = role_refs["sales_invoice"]
    shared_so_to_dn = sorted(so_refs & dn_refs)
    shared_dn_to_si = sorted(dn_refs & si_refs)
    shared_so_to_si = sorted(so_refs & si_refs)
    shared_all_three = sorted(so_refs & dn_refs & si_refs)

    missing_roles = [
        role.replace("_", " ")
        for role, count in role_counts.items()
        if count == 0
    ]
    if not artifacts:
        status = "missing"
        verdict = "No voucher evidence has been uploaded for this case yet."
    elif missing_roles:
        status = "incomplete"
        verdict = f"Need {', '.join(missing_roles)} artifact(s) to compare the full SO -> DN -> SI chain."
    elif analyzable_count == 0:
        status = "incomplete"
        verdict = "Voucher artifacts are present, but the bridge could not read any comparable references from them yet."
    elif shared_all_three:
        status = "linked"
        verdict = "Common references appear in the Sales Order, Delivery Note, and Sales Invoice artifacts, so the chain looks preserved."
    elif not shared_so_to_dn:
        status = "broken"
        verdict = "The Delivery Note does not share a comparable reference with the Sales Order in the uploaded evidence."
    elif not shared_dn_to_si:
        status = "broken"
        verdict = "The Sales Invoice does not share a comparable reference with the Delivery Note in the uploaded evidence."
    else:
        status = "partial"
        verdict = "Some references overlap, but the bridge cannot prove one full SO -> DN -> SI chain from the current evidence."

    return TallyLinkIntegrityCase(
        code=code,
        label=label,
        artifact_count=len(artifacts),
        sales_order_count=role_counts["sales_order"],
        delivery_note_count=role_counts["delivery_note"],
        sales_invoice_count=role_counts["sales_invoice"],
        analyzable_count=analyzable_count,
        shared_all_three=shared_all_three[:8],
        shared_so_to_dn=shared_so_to_dn[:8],
        shared_dn_to_si=shared_dn_to_si[:8],
        shared_so_to_si=shared_so_to_si[:8],
        verdict=verdict,
        status=status,
    )


def _extract_references_from_artifact(artifact: TallyDiagnosticsArtifact) -> set[str]:
    path = Path(artifact.storage_path)
    if not path.exists():
        return set()
    text = _extract_text_payload(path, artifact.filename)
    if not text:
        return set()

    references = {match.group(1) for match in REFERENCE_PREFIX_PATTERN.finditer(text)}
    references.update(match.group(0) for match in REFERENCE_RAW_PATTERN.finditer(text))
    return {value for value in references if value and "/" not in value}


def _extract_text_payload(path: Path, filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".csv", ".tsv", ".xml"}:
        return _decode_text_payload(path.read_bytes())
    if suffix in {".xlsx", ".xlsm"}:
        return _read_openpyxl_text(path)
    if suffix == ".xls" and xlrd is not None:
        return _read_xlrd_text(path)
    return ""


def _decode_text_payload(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return ""


def _read_openpyxl_text(path: Path) -> str:
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return ""

    chunks: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            line = " ".join(_string_value(cell) for cell in row if _string_value(cell))
            if line:
                chunks.append(line)
    return "\n".join(chunks)


def _read_xlrd_text(path: Path) -> str:
    try:
        workbook = xlrd.open_workbook(path.as_posix())
    except Exception:
        return ""

    chunks: list[str] = []
    for sheet_index in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(sheet_index)
        for row_index in range(sheet.nrows):
            values = [_string_value(value) for value in sheet.row_values(row_index) if _string_value(value)]
            if values:
                chunks.append(" ".join(values))
    return "\n".join(chunks)
