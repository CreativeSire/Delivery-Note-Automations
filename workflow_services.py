from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, select

from models import Product, ProductAlias, SalesOrderLine, SalesOrderRun, SkuAutomatorLine, SkuAutomatorRun, db
from services import (
    ProductMatch,
    apply_invoice_classification_to_record,
    build_prefixed_reference,
    load_brand_partner_rules,
    normalize_sku,
    resolve_product_match,
    split_prefixed_reference,
    suggest_products,
)

try:
    import xlrd  # type: ignore
except Exception:  # pragma: no cover
    xlrd = None

SALES_ORDER_SOURCE_SHEET = "Order Item List"
SALES_ORDER_OUTPUT_SHEET = "Sales Order"
SALES_ORDER_HEADERS = [
    "Invoice Date",
    "Order Number",
    "Voucher Type Name",
    "Party name",
    "stock Item Name",
    "Quantity",
    "Rate",
    "Amount",
    "SalesLedger Name",
    "VAT",
    "VAT%",
    "VAT Amount",
]
SALES_ORDER_REQUIRED_HEADERS = {
    "date of order": "invoice_date",
    "order number": "order_number",
    "retailer name": "retailer_name",
    "item name": "source_sku",
    "uom": "source_uom",
    "quantity": "source_quantity",
    "price": "source_rate",
    "total": "source_amount",
}
TALLY_REGISTER_SOURCE_SHEET = "Sales Order- Dala Register"
SKU_REGISTER_HEADERS = ["Date", "Stores", "SKU", "Qty", "Price", "Value", "Order Reference No.", "Voucher No."]
SKU_MATRIX_HEADERS = ["Stores"]
VOUCHER_TYPE_NAME = "Sales Order- Dala"
SALES_LEDGER_NAME = "Inventory Pool"
VAT_LABEL = "VAT"
VAT_RATE = Decimal("7.5")
CASE_UNITS = {"case", "cases", "ctn", "carton", "cartons"}
SALES_ORDER_CATEGORY_COLORS = {
    "BP": {"row_fill": "FFF5E3", "key_fill": "FFC28A25", "key_font": "FFFDF7"},
    "VT": {"row_fill": "E4F7F9", "key_fill": "FF0A7787", "key_font": "FFFFFFFF"},
    "NV": {"row_fill": "F7EEF4", "key_fill": "FF8A5C78", "key_font": "FFFFFFFF"},
}


class WorkflowError(Exception):
    pass


@dataclass
class ModuleSummary:
    run_count: int
    latest_run: Any | None
    recent_runs: list[Any]


@dataclass
class ProductReviewGroup:
    source_sku: str
    occurrences: int
    order_numbers: list[str]
    stores: list[str]
    suggestions: list[Product]


@dataclass
class SalesOrderRunSummary:
    run: SalesOrderRun
    unresolved_groups: list[ProductReviewGroup]
    ready_lines: int
    preview_lines: list[SalesOrderLine]


@dataclass
class SkuMatrixRow:
    store_name: str
    order_references: list[str]
    values_by_sku: dict[str, Decimal]
    total_quantity: Decimal


@dataclass
class SkuAutomatorRunSummary:
    run: SkuAutomatorRun
    unresolved_groups: list[ProductReviewGroup]
    ready_lines: int
    preview_lines: list[SkuAutomatorLine]
    matrix_headers: list[str]
    matrix_rows: list[SkuMatrixRow]


def build_sales_order_summary() -> ModuleSummary:
    return ModuleSummary(
        run_count=db.session.scalar(select(func.count(SalesOrderRun.id))) or 0,
        latest_run=db.session.scalar(select(SalesOrderRun).order_by(SalesOrderRun.created_at.desc()).limit(1)),
        recent_runs=list(db.session.scalars(select(SalesOrderRun).order_by(SalesOrderRun.created_at.desc()).limit(8))),
    )


def build_sku_automator_summary() -> ModuleSummary:
    return ModuleSummary(
        run_count=db.session.scalar(select(func.count(SkuAutomatorRun.id))) or 0,
        latest_run=db.session.scalar(select(SkuAutomatorRun).order_by(SkuAutomatorRun.created_at.desc()).limit(1)),
        recent_runs=list(db.session.scalars(select(SkuAutomatorRun).order_by(SkuAutomatorRun.created_at.desc()).limit(8))),
    )


def create_sales_order_run(file_storage: Any) -> SalesOrderRun:
    workbook = _load_openpyxl_workbook(file_storage, data_only=True)
    sheet = _resolve_sales_order_sheet(workbook)
    header_map = _map_headers(sheet, SALES_ORDER_REQUIRED_HEADERS)

    run = SalesOrderRun(
        id=uuid4().hex,
        original_filename=file_storage.filename or "sales-order-source.xlsx",
        source_sheet_name=sheet.title,
        status="needs_review",
        row_count=0,
        rows_ready=0,
        rows_needing_review=0,
    )
    db.session.add(run)
    db.session.flush()

    products = list(db.session.scalars(select(Product)))
    aliases = list(db.session.scalars(select(ProductAlias)))
    bp_rules = load_brand_partner_rules()

    for row_index in range(2, sheet.max_row + 1):
        source_sku = _string_value(sheet.cell(row_index, header_map["source_sku"]).value)
        order_number = _string_value(sheet.cell(row_index, header_map["order_number"]).value)
        retailer_name = _clean_party_name(_string_value(sheet.cell(row_index, header_map["retailer_name"]).value))
        if not source_sku or not order_number or not retailer_name:
            continue

        invoice_date = _iso_date_value(sheet.cell(row_index, header_map["invoice_date"]).value)
        source_uom = _string_value(sheet.cell(row_index, header_map["source_uom"]).value)
        source_quantity = _decimal_value(sheet.cell(row_index, header_map["source_quantity"]).value)
        source_rate = _decimal_value(sheet.cell(row_index, header_map["source_rate"]).value)
        source_amount = _decimal_value(sheet.cell(row_index, header_map["source_amount"]).value)

        if source_quantity is None or source_quantity <= Decimal("0"):
            continue
        if source_rate is None and source_amount is not None and source_quantity != Decimal("0"):
            source_rate = (source_amount / source_quantity).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
        if source_rate is None or source_rate <= Decimal("0"):
            continue
        if source_amount is None:
            source_amount = (source_quantity * source_rate).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

        run.row_count += 1
        line = SalesOrderLine(
            run_id=run.id,
            invoice_date=invoice_date,
            order_number=order_number,
            retailer_name=retailer_name,
            source_sku=source_sku,
            normalized_source_sku=normalize_sku(source_sku),
            source_uom=source_uom or None,
            source_quantity=source_quantity,
            source_rate=source_rate,
            source_amount=source_amount,
            raw_reference_no=order_number,
        )

        match = resolve_product_match(source_sku, products, aliases)
        if match is None:
            line.status = "needs_review"
            run.rows_needing_review += 1
        else:
            try:
                _apply_product_to_sales_order_line(line, match, bp_rules=bp_rules)
                run.rows_ready += 1
            except WorkflowError:
                line.status = "needs_review"
                run.rows_needing_review += 1

        db.session.add(line)

    run.status = "ready" if run.rows_needing_review == 0 else "needs_review"
    db.session.commit()
    return run


def create_sku_automator_run(file_storage: Any) -> SkuAutomatorRun:
    filename = file_storage.filename or "sales-order-register.xlsx"
    payload = file_storage.read()
    if hasattr(file_storage, "stream"):
        file_storage.stream.seek(0)
    sheet_name, rows = _load_tabular_workbook(payload)

    run = SkuAutomatorRun(
        id=uuid4().hex,
        original_filename=filename,
        source_sheet_name=sheet_name,
        status="needs_review",
        voucher_count=0,
        order_reference_count=0,
        line_count=0,
        rows_ready=0,
        rows_needing_review=0,
        store_count=0,
    )
    db.session.add(run)
    db.session.flush()

    products = list(db.session.scalars(select(Product)))
    aliases = list(db.session.scalars(select(ProductAlias)))
    bp_rules = load_brand_partner_rules()

    current_header: dict[str, str] | None = None
    voucher_numbers: set[str] = set()
    order_refs: set[str] = set()
    stores: set[str] = set()

    for row in rows:
        if _is_tally_header_row(row):
            current_header = {
                "order_date": _iso_date_value(row[0]),
                "store_name": _string_value(row[1]),
                "voucher_no": _string_value(row[2]),
                "order_reference_no": _string_value(row[3]),
            }
            if current_header["voucher_no"]:
                voucher_numbers.add(current_header["voucher_no"])
            if current_header["order_reference_no"]:
                order_refs.add(current_header["order_reference_no"])
            if current_header["store_name"]:
                stores.add(_normalize_store_key(current_header["store_name"]))
            continue

        if current_header is None or not _is_tally_item_row(row):
            continue

        source_sku = _string_value(row[1])
        source_value = _decimal_value(row[7])
        if not source_sku or source_value is None or source_value <= Decimal("0"):
            continue

        run.line_count += 1
        parsed_category, parsed_reference = split_prefixed_reference(current_header["order_reference_no"])
        line = SkuAutomatorLine(
            run_id=run.id,
            order_date=current_header["order_date"],
            store_name=current_header["store_name"],
            voucher_no=current_header["voucher_no"] or None,
            order_reference_no=current_header["order_reference_no"] or None,
            source_sku=source_sku,
            normalized_source_sku=normalize_sku(source_sku),
            source_value=source_value,
            raw_reference_no=parsed_reference or (current_header["order_reference_no"] or None),
            invoice_category=parsed_category,
            prefixed_reference_no=(current_header["order_reference_no"] or None) if parsed_category else None,
            classification_source="prefixed_reference" if parsed_category else None,
        )

        match = resolve_product_match(source_sku, products, aliases)
        if match is None:
            line.status = "needs_review"
            run.rows_needing_review += 1
        else:
            try:
                _apply_product_to_sku_automator_line(line, match, bp_rules=bp_rules)
                run.rows_ready += 1
            except WorkflowError:
                line.status = "needs_review"
                run.rows_needing_review += 1

        db.session.add(line)

    run.voucher_count = len(voucher_numbers)
    run.order_reference_count = len(order_refs)
    run.store_count = len(stores)
    run.status = "ready" if run.rows_needing_review == 0 else "needs_review"
    db.session.commit()
    return run


def build_sales_order_run_summary(run_id: str) -> SalesOrderRunSummary | None:
    run = db.session.get(SalesOrderRun, run_id)
    if run is None:
        return None

    unresolved_lines = list(
        db.session.scalars(
            select(SalesOrderLine).where(SalesOrderLine.run_id == run_id, SalesOrderLine.status == "needs_review")
        )
    )
    unresolved_groups = _group_review_lines(unresolved_lines, line_type="sales_order")
    preview_lines = list(
        db.session.scalars(
            select(SalesOrderLine)
            .where(SalesOrderLine.run_id == run_id, SalesOrderLine.status == "ready")
            .order_by(SalesOrderLine.id.asc())
            .limit(18)
        )
    )
    return SalesOrderRunSummary(
        run=run,
        unresolved_groups=unresolved_groups,
        ready_lines=run.rows_ready,
        preview_lines=preview_lines,
    )


def build_sku_automator_run_summary(run_id: str) -> SkuAutomatorRunSummary | None:
    run = db.session.get(SkuAutomatorRun, run_id)
    if run is None:
        return None

    unresolved_lines = list(
        db.session.scalars(
            select(SkuAutomatorLine).where(SkuAutomatorLine.run_id == run_id, SkuAutomatorLine.status == "needs_review")
        )
    )
    unresolved_groups = _group_review_lines(unresolved_lines, line_type="sku_automator")
    preview_lines = list(
        db.session.scalars(
            select(SkuAutomatorLine)
            .where(SkuAutomatorLine.run_id == run_id, SkuAutomatorLine.status == "ready")
            .order_by(SkuAutomatorLine.id.asc())
            .limit(18)
        )
    )
    matrix_headers, matrix_rows = build_sku_automator_matrix(run_id)
    return SkuAutomatorRunSummary(
        run=run,
        unresolved_groups=unresolved_groups,
        ready_lines=run.rows_ready,
        preview_lines=preview_lines,
        matrix_headers=matrix_headers,
        matrix_rows=matrix_rows[:16],
    )


def apply_sales_order_review_decisions(run_id: str, mapping: dict[str, int]) -> SalesOrderRun:
    run = db.session.get(SalesOrderRun, run_id)
    if run is None:
        raise WorkflowError("This sales-order run could not be found.")

    bp_rules = load_brand_partner_rules()
    for source_sku, product_id in mapping.items():
        product = db.session.get(Product, product_id)
        if product is None:
            raise WorkflowError(f"Selected product for '{source_sku}' could not be found.")
        _validate_sales_order_product(product)
        _persist_alias(source_sku, product)

        lines = list(
            db.session.scalars(
                select(SalesOrderLine).where(SalesOrderLine.run_id == run_id, SalesOrderLine.source_sku == source_sku)
            )
        )
        for line in lines:
            _apply_product_to_sales_order_line(line, ProductMatch(product, "approved-alias"), bp_rules=bp_rules)

    _refresh_sales_order_run(run)
    db.session.commit()
    return run


def apply_sku_automator_review_decisions(run_id: str, mapping: dict[str, int]) -> SkuAutomatorRun:
    run = db.session.get(SkuAutomatorRun, run_id)
    if run is None:
        raise WorkflowError("This SKU Automator run could not be found.")

    bp_rules = load_brand_partner_rules()
    for source_sku, product_id in mapping.items():
        product = db.session.get(Product, product_id)
        if product is None:
            raise WorkflowError(f"Selected product for '{source_sku}' could not be found.")
        _validate_sku_automator_product(product)
        _persist_alias(source_sku, product)

        lines = list(
            db.session.scalars(
                select(SkuAutomatorLine).where(SkuAutomatorLine.run_id == run_id, SkuAutomatorLine.source_sku == source_sku)
            )
        )
        for line in lines:
            _apply_product_to_sku_automator_line(line, ProductMatch(product, "approved-alias"), bp_rules=bp_rules)

    _refresh_sku_automator_run(run)
    db.session.commit()
    return run


def export_sales_order_run_to_workbook(run_id: str) -> tuple[str, bytes]:
    run = db.session.get(SalesOrderRun, run_id)
    if run is None:
        raise WorkflowError("This sales-order run could not be found.")
    if run.rows_needing_review > 0:
        raise WorkflowError("Resolve all review items before downloading the Sales Order file.")

    lines = list(
        db.session.scalars(
            select(SalesOrderLine)
            .where(SalesOrderLine.run_id == run_id, SalesOrderLine.status == "ready")
            .order_by(SalesOrderLine.id.asc())
        )
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SALES_ORDER_OUTPUT_SHEET
    _write_styled_header(sheet, SALES_ORDER_HEADERS)
    sheet.column_dimensions["A"].width = max(sheet.column_dimensions["A"].width or 0, 16)
    for line in lines:
        sheet.append(
            [
                _excel_date_value(line.invoice_date),
                line.prefixed_reference_no or line.raw_reference_no or line.order_number,
                VOUCHER_TYPE_NAME,
                line.retailer_name,
                line.resolved_sku_name,
                line.resolved_quantity_text,
                float(line.resolved_rate or 0),
                float(line.resolved_amount or 0),
                SALES_LEDGER_NAME,
                VAT_LABEL if line.resolved_vatable else "",
                float(VAT_RATE) if line.resolved_vatable else "",
                float(_vat_amount(line.resolved_amount or Decimal("0"))) if line.resolved_vatable else "",
            ]
        )
        _apply_sales_order_category_style(sheet, sheet.max_row, line.invoice_category)
        date_cell = sheet.cell(sheet.max_row, 1)
        if isinstance(date_cell.value, datetime):
            date_cell.number_format = "yyyy-mm-dd"
    payload = _save_workbook(workbook)
    run.status = "exported"
    run.exported_at = datetime.now(UTC)
    db.session.commit()
    filename = f"DALA Sales Order - {_clean_filename_part(Path(run.original_filename).stem, 'Orders')}.xlsx"
    return filename, payload


def build_sku_automator_matrix(run_id: str) -> tuple[list[str], list[SkuMatrixRow]]:
    lines = list(
        db.session.scalars(
            select(SkuAutomatorLine)
            .where(SkuAutomatorLine.run_id == run_id, SkuAutomatorLine.status == "ready")
            .order_by(SkuAutomatorLine.store_name.asc(), SkuAutomatorLine.id.asc())
        )
    )
    if not lines:
        return [], []

    sku_headers = sorted({line.resolved_sku_name for line in lines if line.resolved_sku_name})
    stores: dict[str, SkuMatrixRow] = {}
    for line in lines:
        store_key = _normalize_store_key(line.store_name)
        if store_key not in stores:
            stores[store_key] = SkuMatrixRow(
                store_name=line.store_name,
                order_references=[],
                values_by_sku={sku: Decimal("0") for sku in sku_headers},
                total_quantity=Decimal("0"),
            )
        row = stores[store_key]
        display_reference = line.prefixed_reference_no or line.order_reference_no
        if display_reference and display_reference not in row.order_references:
            row.order_references.append(display_reference)
        quantity = line.resolved_quantity or Decimal("0")
        if line.resolved_sku_name:
            row.values_by_sku[line.resolved_sku_name] = (
                row.values_by_sku.get(line.resolved_sku_name, Decimal("0")) + quantity
            )
        row.total_quantity += quantity

    matrix_rows = sorted(stores.values(), key=lambda item: item.store_name.upper())
    return sku_headers, matrix_rows


def export_sku_automator_run_to_workbook(run_id: str) -> tuple[str, bytes]:
    run = db.session.get(SkuAutomatorRun, run_id)
    if run is None:
        raise WorkflowError("This SKU Automator run could not be found.")
    if run.rows_needing_review > 0:
        raise WorkflowError("Resolve all review items before downloading the SKU Automator output.")

    lines = list(
        db.session.scalars(
            select(SkuAutomatorLine)
            .where(SkuAutomatorLine.run_id == run_id, SkuAutomatorLine.status == "ready")
            .order_by(SkuAutomatorLine.order_date.asc(), SkuAutomatorLine.store_name.asc(), SkuAutomatorLine.id.asc())
        )
    )
    matrix_headers, matrix_rows = build_sku_automator_matrix(run_id)
    workbook = Workbook()
    register = workbook.active
    register.title = TALLY_REGISTER_SOURCE_SHEET
    _write_styled_header(register, SKU_REGISTER_HEADERS)
    for line in lines:
        register.append(
            [
                _excel_date_value(line.order_date),
                line.store_name,
                line.resolved_sku_name,
                float(line.resolved_quantity or 0),
                float(line.resolved_rate or 0),
                float(line.source_value),
                line.prefixed_reference_no or line.order_reference_no or "",
                line.voucher_no or "",
            ]
        )

    matrix_sheet = workbook.create_sheet("Store SKU Matrix")
    _write_styled_header(matrix_sheet, SKU_MATRIX_HEADERS + matrix_headers)
    for row in matrix_rows:
        matrix_sheet.append([row.store_name] + [float(row.values_by_sku.get(sku, Decimal("0"))) for sku in matrix_headers])

    payload = _save_workbook(workbook)
    run.status = "exported"
    run.exported_at = datetime.now(UTC)
    db.session.commit()
    filename = f"SKU Automator - {_clean_filename_part(Path(run.original_filename).stem, 'Register')}.xlsx"
    return filename, payload


def _group_review_lines(lines: list[Any], line_type: str) -> list[ProductReviewGroup]:
    grouped: dict[str, list[Any]] = defaultdict(list)
    for line in lines:
        grouped[line.source_sku].append(line)

    products = list(db.session.scalars(select(Product).where(Product.is_active.is_(True))))
    groups: list[ProductReviewGroup] = []
    for source_sku, items in grouped.items():
        order_numbers: set[str] = set()
        stores: set[str] = set()
        for item in items:
            if line_type == "sales_order":
                order_numbers.add(item.order_number)
                stores.add(item.retailer_name)
            else:
                order_numbers.update(value for value in [item.order_reference_no, item.voucher_no] if value)
                stores.add(item.store_name)

        groups.append(
            ProductReviewGroup(
                source_sku=source_sku,
                occurrences=len(items),
                order_numbers=sorted(order_numbers),
                stores=sorted(stores),
                suggestions=suggest_products(source_sku, products),
            )
        )
    groups.sort(key=lambda item: item.source_sku)
    return groups


def _apply_product_to_sales_order_line(
    line: SalesOrderLine,
    match: ProductMatch,
    *,
    bp_rules: list[Any] | None = None,
) -> None:
    product = match.product
    _validate_sales_order_product(product)
    source_quantity = line.source_quantity or Decimal("0")
    quantity_text, quantity_value = _resolve_sales_order_quantity(line.source_uom, source_quantity, product)
    resolved_rate = _resolve_sales_order_rate(line.source_uom, line.source_rate or Decimal("0"), product)
    resolved_amount = (quantity_value * resolved_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    line.product_id = product.id
    line.status = "ready"
    line.matched_by = match.match_method
    line.resolved_sku_name = product.sku_name
    line.resolved_uom = product.uom
    line.resolved_quantity = quantity_value
    line.resolved_quantity_text = quantity_text
    line.resolved_rate = resolved_rate
    line.resolved_amount = resolved_amount
    line.resolved_vatable = bool(product.vatable)
    classification = apply_invoice_classification_to_record(
        line,
        product=product,
        store_name=line.retailer_name,
        sku_name=product.sku_name,
        raw_reference_no=line.raw_reference_no or line.order_number,
        bp_rules=bp_rules,
    )
    line.prefixed_reference_no = classification.prefixed_reference_no
    line.raw_reference_no = classification.raw_reference_no or line.raw_reference_no


def _apply_product_to_sku_automator_line(
    line: SkuAutomatorLine,
    match: ProductMatch,
    *,
    bp_rules: list[Any] | None = None,
) -> None:
    product = match.product
    _validate_sku_automator_product(product)
    quantity = (line.source_value / product.price).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    line.product_id = product.id
    line.status = "ready"
    line.matched_by = match.match_method
    line.resolved_sku_name = product.sku_name
    line.resolved_quantity = quantity
    line.resolved_rate = product.price
    classification = apply_invoice_classification_to_record(
        line,
        product=product,
        store_name=line.store_name,
        sku_name=product.sku_name,
        raw_reference_no=line.raw_reference_no or line.order_reference_no,
        bp_rules=bp_rules,
    )
    line.prefixed_reference_no = classification.prefixed_reference_no
    line.raw_reference_no = classification.raw_reference_no or line.raw_reference_no


def _refresh_sales_order_run(run: SalesOrderRun) -> None:
    run.rows_ready = db.session.scalar(
        select(func.count(SalesOrderLine.id)).where(SalesOrderLine.run_id == run.id, SalesOrderLine.status == "ready")
    ) or 0
    run.rows_needing_review = db.session.scalar(
        select(func.count(SalesOrderLine.id)).where(
            SalesOrderLine.run_id == run.id, SalesOrderLine.status == "needs_review"
        )
    ) or 0
    run.row_count = db.session.scalar(select(func.count(SalesOrderLine.id)).where(SalesOrderLine.run_id == run.id)) or 0
    run.status = "ready" if run.rows_needing_review == 0 else "needs_review"


def _refresh_sku_automator_run(run: SkuAutomatorRun) -> None:
    run.rows_ready = db.session.scalar(
        select(func.count(SkuAutomatorLine.id)).where(SkuAutomatorLine.run_id == run.id, SkuAutomatorLine.status == "ready")
    ) or 0
    run.rows_needing_review = db.session.scalar(
        select(func.count(SkuAutomatorLine.id)).where(
            SkuAutomatorLine.run_id == run.id, SkuAutomatorLine.status == "needs_review"
        )
    ) or 0
    run.line_count = db.session.scalar(select(func.count(SkuAutomatorLine.id)).where(SkuAutomatorLine.run_id == run.id)) or 0
    ready_lines = list(
        db.session.scalars(
            select(SkuAutomatorLine).where(SkuAutomatorLine.run_id == run.id, SkuAutomatorLine.status == "ready")
        )
    )
    run.voucher_count = len({line.voucher_no for line in ready_lines if line.voucher_no})
    run.order_reference_count = len({line.order_reference_no for line in ready_lines if line.order_reference_no})
    run.store_count = len({_normalize_store_key(line.store_name) for line in ready_lines if line.store_name})
    run.status = "ready" if run.rows_needing_review == 0 else "needs_review"


def _resolve_sales_order_sheet(workbook: Any) -> Any:
    if SALES_ORDER_SOURCE_SHEET in workbook.sheetnames:
        sheet = workbook[SALES_ORDER_SOURCE_SHEET]
        _map_headers(sheet, SALES_ORDER_REQUIRED_HEADERS)
        return sheet
    for name in workbook.sheetnames:
        sheet = workbook[name]
        try:
            _map_headers(sheet, SALES_ORDER_REQUIRED_HEADERS)
            return sheet
        except WorkflowError:
            continue
    raise WorkflowError("The workbook must include a Pep-up order sheet with retailer, item, quantity, and price columns.")


def _map_headers(sheet: Any, required_headers: dict[str, str]) -> dict[str, int]:
    normalized_headers: dict[str, int] = {}
    for column_index in range(1, sheet.max_column + 1):
        header = _normalize_header(sheet.cell(1, column_index).value)
        if header:
            normalized_headers[header] = column_index

    mapped: dict[str, int] = {}
    for header_name, mapped_name in required_headers.items():
        column_index = normalized_headers.get(header_name)
        if column_index is None:
            raise WorkflowError(f"Required column '{header_name}' was not found in sheet '{sheet.title}'.")
        mapped[mapped_name] = column_index
    return mapped


def _load_openpyxl_workbook(file_storage: Any, data_only: bool) -> Any:
    payload = file_storage.read()
    if hasattr(file_storage, "stream"):
        file_storage.stream.seek(0)
    return _load_openpyxl_workbook_from_payload(payload, data_only=data_only)


def _load_openpyxl_workbook_from_payload(payload: bytes, data_only: bool) -> Any:
    try:
        return load_workbook(BytesIO(payload), data_only=data_only)
    except Exception:
        pass

    if xlrd is not None:
        try:
            legacy_workbook = xlrd.open_workbook(file_contents=payload)
            return _convert_xlrd_workbook(legacy_workbook)
        except Exception:
            pass

    delimited_rows = _load_delimited_rows(payload)
    if delimited_rows:
        return _workbook_from_rows(delimited_rows, SALES_ORDER_SOURCE_SHEET)

    raise WorkflowError("The uploaded workbook could not be opened.")


def _load_tabular_workbook(payload: bytes) -> tuple[str, list[list[Any]]]:
    try:
        workbook = load_workbook(BytesIO(payload), data_only=True, read_only=True)
        if TALLY_REGISTER_SOURCE_SHEET in workbook.sheetnames:
            sheet = workbook[TALLY_REGISTER_SOURCE_SHEET]
        else:
            sheet = workbook[workbook.sheetnames[0]]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        return sheet.title, rows
    except Exception:
        if xlrd is None:
            raise WorkflowError("The Tally register could not be opened.")
        try:
            workbook = xlrd.open_workbook(file_contents=payload)
        except Exception as exc:  # pragma: no cover
            raise WorkflowError("The Tally register could not be opened.") from exc

        if TALLY_REGISTER_SOURCE_SHEET in workbook.sheet_names():
            sheet = workbook.sheet_by_name(TALLY_REGISTER_SOURCE_SHEET)
        else:
            sheet = workbook.sheet_by_index(0)
        rows = [sheet.row_values(index) for index in range(sheet.nrows)]
        return sheet.name, rows


def _convert_xlrd_workbook(legacy_workbook: Any) -> Workbook:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for index in range(legacy_workbook.nsheets):
        source_sheet = legacy_workbook.sheet_by_index(index)
        target_sheet = workbook.create_sheet(source_sheet.name or f"Sheet{index + 1}")
        for row_index in range(source_sheet.nrows):
            target_sheet.append(source_sheet.row_values(row_index))
    return workbook


def _load_delimited_rows(payload: bytes) -> list[list[str]] | None:
    text = _decode_delimited_payload(payload)
    if text is None:
        return None

    sample_lines = [line for line in text.splitlines() if line.strip()]
    if not sample_lines:
        return None

    sample = "\n".join(sample_lines[:8])
    delimiter = "\t" if "\t" in sample else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="\t,;|")
        delimiter = dialect.delimiter
    except Exception:
        pass

    rows = [
        [cell.strip() for cell in row]
        for row in csv.reader(StringIO(text), delimiter=delimiter)
        if any(cell.strip() for cell in row)
    ]
    return rows or None


def _decode_delimited_payload(payload: bytes) -> str | None:
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            text = payload.decode(encoding)
        except UnicodeDecodeError:
            continue
        if "\t" in text or "," in text:
            return text
    return None


def _workbook_from_rows(rows: list[list[Any]], sheet_name: str) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(list(row))
    return workbook


def _is_tally_header_row(row: list[Any]) -> bool:
    if len(row) < 4 or row[0] in (None, ""):
        return False
    first = _string_value(row[0]).lower()
    voucher = _string_value(row[2]).lower()
    order_ref = _string_value(row[3]).lower()
    if first == "date" or voucher == "voucher no." or order_ref == "order reference no.":
        return False
    return bool(voucher and order_ref)


def _is_tally_item_row(row: list[Any]) -> bool:
    if len(row) < 8:
        return False
    particulars = _string_value(row[1])
    if not particulars:
        return False
    if particulars.lower() in {"particulars", "grand total", "printed by"}:
        return False
    if _string_value(row[2]) or _string_value(row[3]):
        return False
    return _decimal_value(row[7]) not in (None, Decimal("0"))


def _validate_sales_order_product(product: Product) -> None:
    if not product.uom:
        raise WorkflowError(f"'{product.sku_name}' is missing a base UOM in the master.")
    if not product.is_active:
        raise WorkflowError(f"'{product.sku_name}' is inactive in the master.")


def _validate_sku_automator_product(product: Product) -> None:
    if product.price is None or product.price <= Decimal("0"):
        raise WorkflowError(f"'{product.sku_name}' is missing a valid carton price in the master.")
    if not product.is_active:
        raise WorkflowError(f"'{product.sku_name}' is inactive in the master.")


def _resolve_sales_order_quantity(source_uom: str | None, source_quantity: Decimal, product: Product) -> tuple[str, Decimal]:
    if _is_case_uom(source_uom, product.uom):
        quantity_value = source_quantity
    else:
        if product.conversion is None or product.conversion <= Decimal("0"):
            raise WorkflowError(f"'{product.sku_name}' is missing a conversion value in the master.")
        quantity_value = (source_quantity / product.conversion).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    return f"{_format_decimal(quantity_value)}{product.uom}", quantity_value


def _resolve_sales_order_rate(source_uom: str | None, source_rate: Decimal, product: Product) -> Decimal:
    if _is_case_uom(source_uom, product.uom):
        carton_rate = source_rate
    else:
        if product.conversion is None or product.conversion <= Decimal("0"):
            raise WorkflowError(f"'{product.sku_name}' is missing a conversion value in the master.")
        carton_rate = source_rate * product.conversion

    if product.vatable:
        carton_rate = (carton_rate * Decimal("100") / (Decimal("100") + VAT_RATE)).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
    else:
        carton_rate = carton_rate.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return carton_rate


def _persist_alias(source_sku: str, product: Product) -> None:
    alias = db.session.scalar(select(ProductAlias).where(ProductAlias.alias_name == source_sku))
    if alias is None:
        alias = ProductAlias(
            alias_name=source_sku,
            normalized_name=normalize_sku(source_sku),
            product_id=product.id,
            match_method="approved-alias",
        )
        db.session.add(alias)
    else:
        alias.product_id = product.id
        alias.normalized_name = normalize_sku(source_sku)
        alias.match_method = "approved-alias"


def _write_styled_header(sheet: Any, headers: list[str]) -> None:
    sheet.append(headers)
    header_fill = PatternFill(fill_type="solid", fgColor="1F616C")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column_index)
        cell.fill = header_fill
        cell.font = header_font
        sheet.column_dimensions[cell.column_letter].width = min(max(len(header) + 4, 16), 34)


def _apply_sales_order_category_style(sheet: Any, row_index: int, invoice_category: str | None) -> None:
    palette = SALES_ORDER_CATEGORY_COLORS.get((invoice_category or "").upper())
    if palette is None:
        return

    row_fill = PatternFill(fill_type="solid", fgColor=palette["row_fill"])
    key_fill = PatternFill(fill_type="solid", fgColor=palette["key_fill"])
    key_font = Font(color=palette["key_font"], bold=True)

    for column_index in range(1, len(SALES_ORDER_HEADERS) + 1):
        sheet.cell(row_index, column_index).fill = row_fill

    order_cell = sheet.cell(row_index, 2)
    order_cell.fill = key_fill
    order_cell.font = key_font


def _save_workbook(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.getvalue()


def _decimal_value(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal("1") if value else Decimal("0")
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _iso_date_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _string_value(value)
    if not text:
        return ""
    for candidate in (text, text.split(" ", 1)[0] if " " in text else text):
        for date_format in (
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",
            "%d %b %Y",
            "%d %B %Y",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%d-%b-%Y",
            "%d-%B-%Y",
        ):
            try:
                return datetime.strptime(candidate, date_format).date().isoformat()
            except ValueError:
                continue
        try:
            return date.fromisoformat(candidate).isoformat()
        except ValueError:
            continue
    return text


def _excel_date_value(value: str) -> datetime | str:
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        return value


def _string_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    text = _string_value(value).replace("_", " ")
    return " ".join(text.lower().split())


def _clean_party_name(value: str) -> str:
    return " ".join(value.replace(" ,", ",").split())


def _format_decimal(value: Decimal) -> str:
    normalized = value.normalize()
    text = format(normalized, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _vat_amount(amount: Decimal) -> Decimal:
    return (amount * VAT_RATE / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _is_case_uom(source_uom: str | None, product_uom: str | None) -> bool:
    normalized_source = _normalize_uom(source_uom)
    normalized_product = _normalize_uom(product_uom)
    return normalized_source in CASE_UNITS or (normalized_source and normalized_source == normalized_product)


def _normalize_uom(value: str | None) -> str:
    return _string_value(value).lower()


def _normalize_store_key(value: str) -> str:
    return " ".join(_string_value(value).upper().split())


def _clean_filename_part(value: str, fallback: str) -> str:
    cleaned = "".join(character if character.isalnum() or character in {" ", "-", "_"} else " " for character in value)
    cleaned = " ".join(cleaned.replace("_", " ").split()).strip(" .-")
    return cleaned[:80] if cleaned else fallback
