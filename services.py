from __future__ import annotations

import csv
import re
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from uuid import uuid4
from zoneinfo import ZoneInfo

import xlwt
from flask import current_app
from openpyxl import load_workbook
from sqlalchemy import func, select, true

from audit_services import record_audit_event
from models import (
    BrandPartnerRule,
    InvoiceRoutingEntry,
    InvoiceRoutingImport,
    Product,
    ProductAlias,
    UploadLine,
    UploadRun,
    UomImport,
    UomImportReview,
    db,
)

TRACKER_SHEET = "tracker"
UOM_SHEET = "UOM"
TEMPLATE_SHEET = "Delivery Invoice"
DATE_FORMAT = "%Y-%m-%d"
TRACKER_ORDER_HEADERS = {"sales order number", "order number"}
TRACKER_STORE_HEADERS = {"stores", "store", "supermarket", "supermarket name"}
TRACKER_MIN_COLUMNS = 8
INVOICE_ROUTING_HEADERS = {
    "brand_name": {"brand name", "brand"},
    "sku_name": {"stock item name", "brand sku", "item name", "sku", "product"},
    "party_name": {"party name", "retailer", "retailer name", "customer", "party"},
    "invoice_name": {"invoice", "invoice type", "required invoice", "invoice owner"},
}
OUTPUT_HEADERS = [
    "Invoice Date",
    "Order Number",
    "Voucher Type Name",
    "Party name",
    "stock Item Name",
    "Quantity",
    "Rate",
    "Amount",
    "SalesLedger Name",
    "VAT",
    "VAT%",
    "VAT Amount",
]
VOUCHER_TYPE_NAME = "Delivery Invoice"
SALES_LEDGER_NAME = "Inventory Pool"
VAT_LABEL = "VAT"
VAT_RATE = Decimal("7.5")
SEED_UOM_PATH = Path(__file__).resolve().parent / "data" / "latest_uom_seed.csv"
INVALID_FILENAME_CHARS = r'[<>:"/\\|?*\x00-\x1f]'
INVOICE_OWNER_DALA = "DALA"
INVOICE_OWNER_BP = "BP"
TAX_BUCKET_VT = "VT"
TAX_BUCKET_NV = "NV"
INVOICE_CATEGORY_BP = "BP"
INVOICE_CATEGORY_VT = "VT"
INVOICE_CATEGORY_NV = "NV"
INVOICE_CATEGORY_BPVT = "BPVT"
INVOICE_CATEGORY_BPNV = "BPNV"
INVOICE_EXPORT_CATEGORIES = [
    INVOICE_CATEGORY_BP,
    INVOICE_CATEGORY_VT,
    INVOICE_CATEGORY_NV,
]
INVOICE_CATEGORIES = set(INVOICE_EXPORT_CATEGORIES + [INVOICE_CATEGORY_BPVT, INVOICE_CATEGORY_BPNV])


class ServiceError(Exception):
    pass


class WorkbookShapeError(ServiceError):
    pass


@dataclass
class DashboardSummary:
    product_count: int
    inactive_product_count: int
    alias_count: int
    import_count: int
    invoice_routing_import_count: int
    run_count: int
    latest_import: UomImport | None
    latest_invoice_routing_import: InvoiceRoutingImport | None
    recent_imports: list[UomImport]
    recent_runs: list[UploadRun]


@dataclass
class ProductMatch:
    product: Product
    match_method: str


@dataclass
class InvoiceClassification:
    raw_reference_no: str
    invoice_owner: str | None
    tax_bucket: str | None
    invoice_route_name: str | None
    invoice_category: str | None
    prefixed_reference_no: str | None
    classification_source: str | None
    bp_rule_reason: str | None


@dataclass
class UnresolvedGroup:
    source_sku: str
    occurrences: int
    order_numbers: list[str]
    supermarkets: list[str]
    suggestions: list[Product]


@dataclass
class IgnoredGroup:
    source_sku: str
    occurrences: int
    order_numbers: list[str]
    supermarkets: list[str]
    reason: str


@dataclass
class RunSummary:
    run: UploadRun
    unresolved_groups: list[UnresolvedGroup]
    ignored_groups: list[IgnoredGroup]
    ready_lines: int
    ignored_lines: int
    category_counts: dict[str, int]


@dataclass
class IgnoredHistorySummary:
    run: UploadRun
    ignored_groups: list[IgnoredGroup]
    ignored_lines: int
    ignored_runs_for_sku: list[dict[str, Any]]


@dataclass
class UomImportOutcome:
    import_log: UomImport | None
    review: UomImportReview | None


@dataclass
class InvoiceRoutingSummary:
    latest_import: InvoiceRoutingImport | None
    entry_count: int
    preview_entries: list[InvoiceRoutingEntry]


def build_dashboard_summary() -> DashboardSummary:
    latest_import = db.session.scalar(select(UomImport).order_by(UomImport.created_at.desc()).limit(1))
    latest_invoice_routing_import = db.session.scalar(
        select(InvoiceRoutingImport).order_by(InvoiceRoutingImport.created_at.desc()).limit(1)
    )
    return DashboardSummary(
        product_count=db.session.scalar(select(func.count(Product.id)).where(Product.is_active.is_(True))) or 0,
        inactive_product_count=db.session.scalar(select(func.count(Product.id)).where(Product.is_active.is_(False))) or 0,
        alias_count=db.session.scalar(select(func.count(ProductAlias.id))) or 0,
        import_count=db.session.scalar(select(func.count(UomImport.id))) or 0,
        invoice_routing_import_count=db.session.scalar(select(func.count(InvoiceRoutingImport.id))) or 0,
        run_count=db.session.scalar(select(func.count(UploadRun.id))) or 0,
        latest_import=latest_import,
        latest_invoice_routing_import=latest_invoice_routing_import,
        recent_imports=list(db.session.scalars(select(UomImport).order_by(UomImport.created_at.desc()).limit(5))),
        recent_runs=list(db.session.scalars(select(UploadRun).order_by(UploadRun.created_at.desc()).limit(8))),
    )


def build_invoice_routing_summary(*, limit: int = 8) -> InvoiceRoutingSummary:
    latest_import = db.session.scalar(select(InvoiceRoutingImport).order_by(InvoiceRoutingImport.created_at.desc()).limit(1))
    preview_entries = list(
        db.session.scalars(
            select(InvoiceRoutingEntry)
            .order_by(InvoiceRoutingEntry.party_name.asc(), InvoiceRoutingEntry.sku_name.asc(), InvoiceRoutingEntry.id.asc())
            .limit(limit)
        )
    )
    entry_count = db.session.scalar(select(func.count(InvoiceRoutingEntry.id))) or 0
    return InvoiceRoutingSummary(
        latest_import=latest_import,
        entry_count=entry_count,
        preview_entries=preview_entries,
    )


def list_brand_partner_rules() -> list[BrandPartnerRule]:
    return list(
        db.session.scalars(
            select(BrandPartnerRule)
            .where(BrandPartnerRule.is_active.is_(True))
            .order_by(BrandPartnerRule.sku_name_pattern.asc(), BrandPartnerRule.store_name_pattern.asc())
        )
    )


def load_invoice_routing_entries() -> list[InvoiceRoutingEntry]:
    return list(
        db.session.scalars(
            select(InvoiceRoutingEntry).order_by(InvoiceRoutingEntry.party_name.asc(), InvoiceRoutingEntry.sku_name.asc())
        )
    )


def preview_brand_partner_classification(
    *,
    sku_name: str,
    store_name: str | None,
    raw_reference_no: str | None,
    product_id: int | None = None,
) -> dict[str, Any]:
    product = db.session.get(Product, product_id) if product_id else None
    classification = classify_invoice_line(
        raw_reference_no=raw_reference_no,
        store_name=store_name,
        sku_name=sku_name,
        product=product,
        bp_rules=load_brand_partner_rules(),
    )
    return {
        "sku_name": sku_name,
        "store_name": store_name or "",
        "raw_reference_no": raw_reference_no or "",
        "invoice_owner": classification.invoice_owner or "",
        "tax_bucket": classification.tax_bucket or "",
        "invoice_route_name": classification.invoice_route_name or "",
        "product": product,
        "invoice_category": classification.invoice_category or "",
        "prefixed_reference_no": classification.prefixed_reference_no or "",
        "classification_source": classification.classification_source or "",
        "bp_rule_reason": classification.bp_rule_reason or "",
    }


def save_brand_partner_rule(form_data: dict[str, Any]) -> BrandPartnerRule:
    sku_name_pattern = _string_value(form_data.get("sku_name_pattern"))
    if not sku_name_pattern:
        raise ServiceError("SKU pattern is required for a Brand Partner rule.")

    store_name_pattern = _string_value(form_data.get("store_name_pattern")) or None
    rule_name = _string_value(form_data.get("rule_name")) or None
    normalized_sku_pattern = normalize_sku(sku_name_pattern)
    normalized_store_pattern = normalize_sku(store_name_pattern) if store_name_pattern else None

    duplicate_query = select(BrandPartnerRule).where(
        BrandPartnerRule.is_active.is_(True),
        BrandPartnerRule.normalized_sku_pattern == normalized_sku_pattern,
        BrandPartnerRule.normalized_store_pattern.is_(normalized_store_pattern)
        if normalized_store_pattern is None
        else BrandPartnerRule.normalized_store_pattern == normalized_store_pattern,
    )
    duplicate = db.session.scalar(duplicate_query)
    if duplicate is not None:
        raise ServiceError("That Brand Partner rule already exists.")

    rule = BrandPartnerRule(
        rule_name=rule_name,
        sku_name_pattern=sku_name_pattern,
        normalized_sku_pattern=normalized_sku_pattern,
        store_name_pattern=store_name_pattern,
        normalized_store_pattern=normalized_store_pattern,
        is_active=True,
    )
    db.session.add(rule)
    db.session.commit()
    return rule


def set_brand_partner_rule_active(rule_id: int, is_active: bool) -> BrandPartnerRule:
    rule = db.session.get(BrandPartnerRule, rule_id)
    if rule is None:
        raise ServiceError("That Brand Partner rule could not be found.")

    rule.is_active = is_active
    db.session.commit()
    return rule


def import_invoice_routing_workbook(file_storage: Any) -> InvoiceRoutingImport:
    payload = _read_upload_payload(file_storage)
    filename = file_storage.filename or "invoice-routing.xlsx"

    workbook = _try_load_workbook(payload)
    if workbook is not None:
        rows = _extract_invoice_routing_rows_from_workbook(workbook)
    else:
        rows = _extract_invoice_routing_rows_from_delimited_payload(payload)

    if not rows:
        raise WorkbookShapeError(
            "The uploaded file must contain Brand Name, Stock item Name, Party Name, and Invoice columns."
        )

    return import_invoice_routing_rows(rows, filename)


def import_invoice_routing_rows(rows: list[dict[str, str]], filename: str) -> InvoiceRoutingImport:
    import_log = InvoiceRoutingImport(filename=filename, row_count=0)
    db.session.add(import_log)
    db.session.flush()

    db.session.query(InvoiceRoutingEntry).delete()

    seen_keys: set[tuple[str, str, str]] = set()
    for row in rows:
        sku_name = _string_value(row.get("sku_name"))
        party_name = _string_value(row.get("party_name"))
        invoice_name = _string_value(row.get("invoice_name"))
        if not sku_name or not party_name or not invoice_name:
            continue

        normalized_sku_name = normalize_sku(sku_name)
        normalized_party_name = normalize_sku(party_name)
        normalized_invoice_name = normalize_sku(invoice_name)
        key = (normalized_sku_name, normalized_party_name, normalized_invoice_name)
        if key in seen_keys:
            continue
        seen_keys.add(key)

        db.session.add(
            InvoiceRoutingEntry(
                import_id=import_log.id,
                brand_name=_string_value(row.get("brand_name")) or None,
                normalized_brand_name=normalize_sku(_string_value(row.get("brand_name"))) or None,
                sku_name=sku_name,
                normalized_sku_name=normalized_sku_name,
                party_name=party_name,
                normalized_party_name=normalized_party_name,
                invoice_name=invoice_name,
                normalized_invoice_name=normalized_invoice_name,
            )
        )
        import_log.row_count += 1

    db.session.commit()
    return import_log


def import_uom_workbook(file_storage: Any) -> UomImportOutcome:
    payload = _read_upload_payload(file_storage)
    filename = file_storage.filename or "uom.xlsx"

    workbook = _try_load_workbook(payload)
    if workbook is not None:
        rows = _extract_uom_workbook_rows(workbook)
        if rows is not None:
            review = create_uom_import_review(rows, filename)
            if review is not None:
                return UomImportOutcome(import_log=None, review=review)
            return UomImportOutcome(import_log=import_uom_rows(rows, filename, mode="replace"), review=None)

        stock_rows = _extract_stock_category_summary_rows(workbook)
        if stock_rows is not None:
            review = create_uom_import_review(stock_rows, filename)
            if review is not None:
                return UomImportOutcome(import_log=None, review=review)
            return UomImportOutcome(import_log=import_uom_rows(stock_rows, filename, mode="replace"), review=None)

        raise WorkbookShapeError(
            f"The workbook must contain a '{UOM_SHEET}' sheet, a recognized UOM header layout, or a recognized 'Stock Category Summary' sheet."
        )

    item_rows = _extract_item_list_rows(payload)
    if item_rows is not None:
        return UomImportOutcome(import_log=import_uom_rows(item_rows, filename, mode="merge"), review=None)

    raise WorkbookShapeError(
        "The uploaded file must be a UOM workbook, a workbook with UOM-style headers, a Stock Category Summary workbook, or an item list export."
    )


def import_uom_rows(
    rows: list[list[Any]],
    filename: str,
    mode: str = "replace",
    *,
    keep_active_product_ids: set[int] | None = None,
    forced_product_matches: dict[str, int] | None = None,
) -> UomImport:
    import_log = UomImport(filename=filename, product_count=0)
    db.session.add(import_log)
    db.session.flush()

    protected_ids = keep_active_product_ids or set()
    forced_matches = {str(key): value for key, value in (forced_product_matches or {}).items()}
    existing_products = {
        product.sku_name: product for product in db.session.scalars(select(Product))
    }
    existing_by_normalized: dict[str, list[Product]] = {}
    existing_by_sync_key: dict[str, list[Product]] = {}
    for product in existing_products.values():
        existing_by_normalized.setdefault(product.normalized_name, []).append(product)
        if product.source_import_id is not None:
            existing_by_sync_key.setdefault(normalize_uom_sync_key(product.sku_name), []).append(product)
        if mode == "replace" and product.source_import_id is not None and product.id not in protected_ids:
            product.is_active = False

    imported = 0
    skipped = 0
    deactivated = 0
    for row in rows:
        sku_name = _string_value(row[0])
        if not sku_name:
            continue

        product = None
        forced_product_id = forced_matches.get(sku_name)
        if forced_product_id is not None:
            product = db.session.get(Product, forced_product_id)

        if product is None:
            product = existing_products.get(sku_name)
        normalized_name = normalize_sku(sku_name)
        sync_key = normalize_uom_sync_key(sku_name)
        is_active_row = True if len(row) < 7 else bool(row[6])
        if product is None:
            normalized_matches = existing_by_normalized.get(normalized_name, [])
            if len(normalized_matches) == 1:
                product = normalized_matches[0]

        if product is None and mode == "replace":
            sync_matches = existing_by_sync_key.get(sync_key, [])
            if len(sync_matches) == 1:
                product = sync_matches[0]

        if product is None:
            product = Product(sku_name=sku_name, normalized_name=normalized_name)
            db.session.add(product)
            existing_products[sku_name] = product
            existing_by_normalized.setdefault(normalized_name, []).append(product)
        elif mode == "merge":
            if is_active_row:
                skipped += 1
                continue
            product.is_active = False
            product.source_import_id = import_log.id
            deactivated += 1
            continue

        old_sku_name = product.sku_name
        old_normalized_name = product.normalized_name
        if old_sku_name != sku_name:
            if existing_products.get(old_sku_name) is product:
                del existing_products[old_sku_name]
            existing_products[sku_name] = product
            product.sku_name = sku_name

        if old_normalized_name != normalized_name:
            previous_matches = existing_by_normalized.get(old_normalized_name, [])
            existing_by_normalized[old_normalized_name] = [item for item in previous_matches if item is not product]
            if not existing_by_normalized[old_normalized_name]:
                del existing_by_normalized[old_normalized_name]
            existing_by_normalized.setdefault(normalized_name, []).append(product)

        product.normalized_name = normalized_name
        product.uom = _string_value(row[1]) or None
        product.alt_uom = _string_value(row[2]) or None
        product.conversion = _decimal_value(row[3])
        product.vatable = _string_value(row[4]).lower() == "yes"
        product.price = _decimal_value(row[5])
        product.is_active = is_active_row
        product.source_import_id = import_log.id
        if is_active_row:
            imported += 1
        else:
            deactivated += 1

    import_log.product_count = imported
    import_log.skipped_count = skipped
    import_log.import_mode = mode
    import_log.deactivated_count = deactivated
    db.session.commit()
    return import_log


def create_uom_import_review(rows: list[list[Any]], filename: str) -> UomImportReview | None:
    serialized_rows = [_serialize_uom_row(row) for row in rows if _string_value(row[0])]
    if not serialized_rows:
        return None

    active_source_products = list(
        db.session.scalars(
            select(Product).where(Product.is_active.is_(True), Product.source_import_id.is_not(None)).order_by(Product.sku_name.asc())
        )
    )
    if not active_source_products:
        return None

    matched_product_ids: set[int] = set()
    unmatched_rows: list[dict[str, Any]] = []
    existing_by_name = {product.sku_name: product for product in db.session.scalars(select(Product))}
    existing_by_normalized: dict[str, list[Product]] = {}
    existing_by_sync_key: dict[str, list[Product]] = {}
    for product in db.session.scalars(select(Product)):
        existing_by_normalized.setdefault(product.normalized_name, []).append(product)
        if product.source_import_id is not None:
            existing_by_sync_key.setdefault(normalize_uom_sync_key(product.sku_name), []).append(product)

    for row in serialized_rows:
        product = _match_existing_uom_product(
            row["sku_name"],
            existing_by_name=existing_by_name,
            existing_by_normalized=existing_by_normalized,
            existing_by_sync_key=existing_by_sync_key,
        )
        if product is not None:
            matched_product_ids.add(product.id)
            continue
        unmatched_rows.append(row)

    missing_products = [product for product in active_source_products if product.id not in matched_product_ids]
    if not missing_products:
        return None

    suggestions = _suggest_uom_review_matches(missing_products, unmatched_rows)
    missing_payload = []
    for product in missing_products:
        suggestion = suggestions.get(product.id)
        missing_payload.append(
            {
                "product_id": product.id,
                "sku_name": product.sku_name,
                "price": _decimal_string(product.price),
                "uom": product.uom or "",
                "suggested_incoming_sku": suggestion["sku_name"] if suggestion else "",
                "suggested_reason": suggestion["reason"] if suggestion else "",
                "suggested_score": suggestion["score"] if suggestion else 0,
            }
        )

    review = UomImportReview(
        id=uuid4().hex,
        filename=filename,
        status="pending",
        row_count=len(serialized_rows),
        matched_count=len(matched_product_ids),
        new_count=len(unmatched_rows),
        missing_count=len(missing_products),
        rename_candidate_count=sum(1 for item in missing_payload if item["suggested_incoming_sku"]),
        rows_json=serialized_rows,
        unmatched_rows_json=unmatched_rows,
        missing_products_json=missing_payload,
    )
    db.session.add(review)
    record_audit_event(
        module_name="Database",
        event_type="uom_review_created",
        entity_type="uom_review",
        entity_id=review.id,
        entity_name=filename,
        summary_text=(
            f"UOM review created for {filename} with {review.missing_count} at-risk product"
            f"{'' if review.missing_count == 1 else 's'}."
        ),
        details={
            "row_count": review.row_count,
            "matched_count": review.matched_count,
            "new_count": review.new_count,
            "missing_count": review.missing_count,
            "rename_candidate_count": review.rename_candidate_count,
        },
    )
    db.session.commit()
    return review


def get_uom_import_review(review_id: str) -> UomImportReview | None:
    return db.session.get(UomImportReview, review_id)


def get_pending_uom_import_review() -> UomImportReview | None:
    return db.session.scalar(
        select(UomImportReview)
        .where(UomImportReview.status == "pending")
        .order_by(UomImportReview.created_at.desc())
        .limit(1)
    )


def apply_uom_import_review(review_id: str, decisions: dict[str, str]) -> tuple[UomImportReview, UomImport]:
    review = db.session.get(UomImportReview, review_id)
    if review is None:
        raise ServiceError("That UOM review could not be found.")
    if review.status != "pending":
        raise ServiceError("That UOM review has already been applied or dismissed.")

    keep_active_ids: set[int] = set()
    forced_matches: dict[str, int] = {}
    merged_count = 0
    kept_count = 0
    inactive_count = 0

    for item in review.missing_products_json or []:
        product_id = int(item["product_id"])
        decision = (decisions.get(str(product_id)) or "").strip()
        if not decision:
            if item.get("suggested_incoming_sku"):
                decision = f"merge::{item['suggested_incoming_sku']}"
            else:
                decision = "inactive"

        if decision.startswith("merge::"):
            incoming_sku = decision.split("::", 1)[1].strip()
            if not incoming_sku:
                raise ServiceError("A rename merge decision was missing its incoming SKU.")
            forced_matches[incoming_sku] = product_id
            merged_count += 1
        elif decision == "keep":
            keep_active_ids.add(product_id)
            kept_count += 1
        else:
            inactive_count += 1

    import_log = import_uom_rows(
        [_deserialize_uom_row(item) for item in review.rows_json or []],
        review.filename,
        mode="replace",
        keep_active_product_ids=keep_active_ids,
        forced_product_matches=forced_matches,
    )
    review.status = "applied"
    review.import_log_id = import_log.id
    review.applied_at = datetime.now(UTC)
    record_audit_event(
        module_name="Database",
        event_type="uom_review_applied",
        entity_type="uom_review",
        entity_id=review.id,
        entity_name=review.filename,
        summary_text=(
            f"Applied reviewed UOM refresh for {review.filename}: {merged_count} merge"
            f"{'' if merged_count == 1 else 's'}, {kept_count} kept active, {inactive_count} inactivated."
        ),
        details={
            "import_log_id": import_log.id,
            "merged_count": merged_count,
            "kept_count": kept_count,
            "inactive_count": inactive_count,
        },
    )
    db.session.commit()
    return review, import_log


def discard_uom_import_review(review_id: str) -> UomImportReview:
    review = db.session.get(UomImportReview, review_id)
    if review is None:
        raise ServiceError("That UOM review could not be found.")
    if review.status != "pending":
        raise ServiceError("Only pending UOM reviews can be dismissed.")
    review.status = "discarded"
    record_audit_event(
        module_name="Database",
        event_type="uom_review_discarded",
        entity_type="uom_review",
        entity_id=review.id,
        entity_name=review.filename,
        summary_text=f"Dismissed the pending UOM review for {review.filename}.",
    )
    db.session.commit()
    return review


def bootstrap_seed_uom_if_empty() -> UomImport | None:
    if not SEED_UOM_PATH.exists():
        return None

    if db.session.scalar(select(func.count(UomImport.id))) not in (None, 0):
        return None

    rows: list[list[Any]] = []
    with SEED_UOM_PATH.open("r", encoding="utf-8", newline="") as seed_file:
        reader = csv.reader(seed_file)
        next(reader, None)
        for row in reader:
            rows.append(row[:6])

    if not rows:
        return None

    return import_uom_rows(rows, "LT to DN system 1.xlsx (seed)")


def save_product_master_entry(form_data: dict[str, Any], product_id: int | None = None) -> Product:
    sku_name = _string_value(form_data.get("sku_name"))
    if not sku_name:
        raise ServiceError("Product name is required.")

    price = _decimal_value(form_data.get("price"))
    if price is None:
        raise ServiceError("Rate/price is required.")

    product = db.session.get(Product, product_id) if product_id is not None else None

    duplicate = db.session.scalar(select(Product).where(Product.sku_name == sku_name))
    if duplicate is not None and (product is None or duplicate.id != product.id):
        raise ServiceError(f"'{sku_name}' already exists in the product master.")

    if product is None:
        product = Product(
            sku_name=sku_name,
            normalized_name=normalize_sku(sku_name),
            is_active=True,
            source_import_id=None,
        )
        db.session.add(product)

    product.sku_name = sku_name
    product.normalized_name = normalize_sku(sku_name)
    product.uom = _string_value(form_data.get("uom")) or None
    product.alt_uom = _string_value(form_data.get("alt_uom")) or None
    product.conversion = _decimal_value(form_data.get("conversion"))
    product.price = price
    product.vatable = _string_value(form_data.get("vatable")).lower() in {"yes", "true", "1", "on"}
    product.is_active = True
    if product.source_import_id is None:
        product.source_import_id = None

    db.session.commit()
    return product


def set_product_active(product_id: int, is_active: bool) -> Product:
    product = db.session.get(Product, product_id)
    if product is None:
        raise ServiceError("That product could not be found.")

    product.is_active = is_active
    db.session.commit()
    return product


def create_tracker_run(file_storage: Any, timezone_name: str) -> UploadRun:
    workbook = _load_workbook_from_upload(file_storage)
    sheet = _resolve_tracker_sheet(workbook)
    product_headers = []
    for column_index in range(3, sheet.max_column + 1):
        value = _string_value(sheet.cell(1, column_index).value)
        if value:
            product_headers.append((column_index, value))

    run = UploadRun(
        id=uuid4().hex,
        original_filename=file_storage.filename or "tracker.xlsx",
        invoice_date=tomorrow_in_timezone(timezone_name).strftime(DATE_FORMAT),
        status="needs_review",
        rows_detected=0,
        rows_ready=0,
        rows_needing_review=0,
    )
    db.session.add(run)
    db.session.flush()

    products = list(db.session.scalars(select(Product)))
    aliases = list(db.session.scalars(select(ProductAlias)))
    bp_rules = load_brand_partner_rules()
    invoice_routing_entries = load_invoice_routing_entries()

    for row_index in range(2, sheet.max_row + 1):
        order_number = _string_value(sheet.cell(row_index, 1).value)
        supermarket = _string_value(sheet.cell(row_index, 2).value)
        if not order_number and not supermarket:
            continue

        for column_index, sku_name in product_headers:
            quantity = _decimal_value(sheet.cell(row_index, column_index).value)
            if quantity is None or quantity <= Decimal("0"):
                continue

            run.rows_detected += 1
            parsed_category, parsed_reference = split_prefixed_reference(order_number)
            _, parsed_owner, parsed_tax_bucket = invoice_category_parts(parsed_category)
            line = UploadLine(
                run_id=run.id,
                order_number=order_number,
                supermarket_name=supermarket,
                source_sku=sku_name,
                normalized_source_sku=normalize_sku(sku_name),
                quantity=quantity,
                raw_reference_no=parsed_reference or order_number,
                invoice_owner=parsed_owner,
                tax_bucket=parsed_tax_bucket,
                invoice_category=parsed_category,
                prefixed_reference_no=order_number if parsed_category else None,
                classification_source="prefixed_reference" if parsed_category else None,
            )

            match = resolve_product_match(sku_name, products, aliases)
            if match is None:
                line.status = "needs_review"
                run.rows_needing_review += 1
            elif not match.product.is_active:
                line.status = "ignored"
                line.matched_by = "inactive"
                line.product_id = match.product.id
                line.resolved_sku_name = match.product.sku_name
            elif match.product.price is None:
                line.status = "needs_review"
                run.rows_needing_review += 1
            else:
                apply_product_to_line(
                    line,
                    match.product,
                    match.match_method,
                    bp_rules=bp_rules,
                    invoice_routing_entries=invoice_routing_entries,
                )
                run.rows_ready += 1

            db.session.add(line)

    run.status = "ready" if run.rows_needing_review == 0 else "needs_review"
    db.session.commit()
    return run


def build_run_summary(run_id: str) -> RunSummary | None:
    run = db.session.get(UploadRun, run_id)
    if run is None:
        return None

    unresolved_lines = list(
        db.session.scalars(
            select(UploadLine).where(UploadLine.run_id == run_id, UploadLine.status == "needs_review")
        )
    )

    grouped: dict[str, list[UploadLine]] = {}
    for line in unresolved_lines:
        grouped.setdefault(line.source_sku, []).append(line)

    products = list(db.session.scalars(select(Product).where(Product.is_active.is_(True))))
    groups = []
    for source_sku, lines in grouped.items():
        groups.append(
            UnresolvedGroup(
                source_sku=source_sku,
                occurrences=len(lines),
                order_numbers=sorted({line.order_number for line in lines}),
                supermarkets=sorted({line.supermarket_name for line in lines}),
                suggestions=suggest_products(source_sku, products),
            )
        )

    groups.sort(key=lambda item: item.source_sku)
    ignored_line_items = list(
        db.session.scalars(
            select(UploadLine).where(UploadLine.run_id == run_id, UploadLine.status == "ignored")
        )
    )
    ignored_grouped: dict[str, list[UploadLine]] = {}
    for line in ignored_line_items:
        ignored_grouped.setdefault(line.source_sku, []).append(line)

    ignored_groups = _build_ignored_groups(ignored_grouped)
    ready_lines = db.session.scalar(
        select(func.count(UploadLine.id)).where(UploadLine.run_id == run_id, UploadLine.status == "ready")
    ) or 0
    ignored_lines = db.session.scalar(
        select(func.count(UploadLine.id)).where(UploadLine.run_id == run_id, UploadLine.status == "ignored")
    ) or 0
    ready_line_items = list(
        db.session.scalars(select(UploadLine).where(UploadLine.run_id == run_id, UploadLine.status == "ready"))
    )
    normalized_categories = [invoice_category_parts(line.invoice_category)[0] or line.invoice_category for line in ready_line_items]
    category_counts = {
        category: sum(1 for line_category in normalized_categories if line_category == category)
        for category in INVOICE_EXPORT_CATEGORIES
    }
    return RunSummary(
        run=run,
        unresolved_groups=groups,
        ignored_groups=ignored_groups,
        ready_lines=ready_lines,
        ignored_lines=ignored_lines,
        category_counts=category_counts,
    )


def build_ignored_history_summary(run_id: str) -> IgnoredHistorySummary | None:
    run = db.session.get(UploadRun, run_id)
    if run is None:
        return None

    ignored_line_items = list(
        db.session.scalars(
            select(UploadLine).where(UploadLine.run_id == run_id, UploadLine.status == "ignored")
        )
    )
    ignored_grouped: dict[str, list[UploadLine]] = {}
    for line in ignored_line_items:
        ignored_grouped.setdefault(line.source_sku, []).append(line)
    ignored_groups = _build_ignored_groups(ignored_grouped)

    ignored_source_skus = list(ignored_grouped.keys())
    ignored_runs_for_sku: list[dict[str, Any]] = []
    if ignored_source_skus:
        ignored_lines_history = list(
            db.session.scalars(
                select(UploadLine)
                .join(UploadRun, UploadRun.id == UploadLine.run_id)
                .where(UploadLine.status == "ignored", UploadLine.source_sku.in_(ignored_source_skus))
                .order_by(UploadRun.created_at.desc(), UploadLine.id.asc())
            )
        )
        by_sku_counter = Counter(line.source_sku for line in ignored_lines_history)
        by_run_counter = Counter(line.run_id for line in ignored_lines_history)
        for line in ignored_lines_history:
            ignored_runs_for_sku.append(
                {
                    "source_sku": line.source_sku,
                    "run_id": line.run_id,
                    "run_filename": line.run.original_filename if line.run else "",
                    "invoice_date": line.run.invoice_date if line.run else "",
                    "supermarket_name": line.supermarket_name,
                    "order_number": line.order_number,
                    "occurrences_for_sku": by_sku_counter[line.source_sku],
                    "occurrences_for_run": by_run_counter[line.run_id],
                }
            )

    return IgnoredHistorySummary(
        run=run,
        ignored_groups=ignored_groups,
        ignored_lines=len(ignored_line_items),
        ignored_runs_for_sku=ignored_runs_for_sku,
    )


def apply_review_decisions(run_id: str, mapping: dict[str, int]) -> UploadRun:
    run = db.session.get(UploadRun, run_id)
    if run is None:
        raise WorkbookShapeError("This upload run could not be found.")

    bp_rules = load_brand_partner_rules()
    invoice_routing_entries = load_invoice_routing_entries()
    for source_sku, product_id in mapping.items():
        product = db.session.get(Product, product_id)
        if product is None:
            raise WorkbookShapeError(f"Selected product for '{source_sku}' could not be found.")
        if product.price is None:
            raise WorkbookShapeError(f"'{product.sku_name}' still has no price in the UOM master.")

        alias = db.session.scalar(select(ProductAlias).where(ProductAlias.alias_name == source_sku))
        if alias is None:
            alias = ProductAlias(
                alias_name=source_sku,
                normalized_name=normalize_sku(source_sku),
                product_id=product.id,
                match_method="approved-alias",
            )
            db.session.add(alias)
        else:
            alias.product_id = product.id
            alias.normalized_name = normalize_sku(source_sku)
            alias.match_method = "approved-alias"

        lines = list(
            db.session.scalars(
                select(UploadLine).where(UploadLine.run_id == run_id, UploadLine.source_sku == source_sku)
            )
        )
        for line in lines:
            apply_product_to_line(
                line,
                product,
                "approved-alias",
                bp_rules=bp_rules,
                invoice_routing_entries=invoice_routing_entries,
            )

    _refresh_run_totals(run)
    db.session.commit()
    return run


def mark_source_sku_inactive(run_id: str, source_sku: str) -> tuple[UploadRun, Product]:
    run = db.session.get(UploadRun, run_id)
    if run is None:
        raise WorkbookShapeError("This upload run could not be found.")

    sku_name = _string_value(source_sku)
    if not sku_name:
        raise WorkbookShapeError("The source SKU to mark inactive was empty.")

    product = db.session.scalar(select(Product).where(Product.sku_name == sku_name))
    if product is None:
        product = Product(
            sku_name=sku_name,
            normalized_name=normalize_sku(sku_name),
            is_active=False,
            source_import_id=None,
        )
        db.session.add(product)
        db.session.flush()
    else:
        product.is_active = False

    lines = list(
        db.session.scalars(
            select(UploadLine).where(UploadLine.run_id == run_id, UploadLine.source_sku == sku_name)
        )
    )
    if not lines:
        raise WorkbookShapeError(f"No review lines were found for '{sku_name}'.")

    for line in lines:
        line.status = "ignored"
        line.matched_by = "inactive"
        line.product_id = product.id
        line.resolved_sku_name = product.sku_name
        line.resolved_rate = None
        line.resolved_vatable = False
        line.invoice_owner = None
        line.tax_bucket = None
        line.invoice_route_name = None
        line.invoice_category = None
        line.prefixed_reference_no = None
        line.classification_source = None
        line.bp_rule_reason = None

    _refresh_run_totals(run)
    db.session.commit()
    return run, product


def export_run_to_xls(run_id: str, invoice_category: str | None = None) -> tuple[str, bytes]:
    run = db.session.get(UploadRun, run_id)
    if run is None:
        raise WorkbookShapeError("This upload run could not be found.")
    if run.rows_needing_review > 0:
        raise WorkbookShapeError("Resolve all review items before downloading the final file.")

    selected_category = _string_value(invoice_category).upper()
    if selected_category and selected_category not in INVOICE_CATEGORIES:
        raise WorkbookShapeError("The selected invoice category is not supported.")

    allowed_categories: set[str] | None = None
    if selected_category:
        if selected_category == INVOICE_CATEGORY_BP:
            allowed_categories = {INVOICE_CATEGORY_BP, INVOICE_CATEGORY_BPVT, INVOICE_CATEGORY_BPNV}
        else:
            allowed_categories = {selected_category}

    lines = list(
        db.session.scalars(
            select(UploadLine)
            .where(
                UploadLine.run_id == run_id,
                UploadLine.status == "ready",
                UploadLine.invoice_category.in_(allowed_categories) if allowed_categories else true(),
            )
            .order_by(UploadLine.id.asc())
        )
    )
    if not lines:
        label = selected_category or "selected"
        raise WorkbookShapeError(f"There are no ready {label} lines to export.")
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet(TEMPLATE_SHEET)

    header_style = xlwt.easyxf(
        "font: bold on, height 240;"
        "pattern: pattern solid, fore_colour ocean_blue;"
        "align: horiz center;"
        "borders: left thin, right thin, top thin, bottom thin;"
    )
    text_style = xlwt.easyxf("borders: left thin, right thin, top thin, bottom thin;")
    date_style = xlwt.easyxf(
        "borders: left thin, right thin, top thin, bottom thin;",
        num_format_str="yyyy-mm-dd",
    )
    decimal_style = xlwt.easyxf(
        "borders: left thin, right thin, top thin, bottom thin;",
        num_format_str="0.00##",
    )

    for col_index, header in enumerate(OUTPUT_HEADERS):
        sheet.write(0, col_index, header, header_style)
        sheet.col(col_index).width = min(max(len(header) + 3, 14) * 256, 35 * 256)

    for row_index, line in enumerate(lines, start=1):
        amount = line.quantity * line.resolved_rate
        vat_amount = (amount * VAT_RATE / Decimal("100")).quantize(Decimal("0.01")) if line.resolved_vatable else ""
        display_reference = build_prefixed_reference(line.invoice_category, line.raw_reference_no) or line.prefixed_reference_no or line.order_number
        row = [
            run.invoice_date,
            display_reference,
            VOUCHER_TYPE_NAME,
            line.supermarket_name,
            line.resolved_sku_name,
            float(line.quantity),
            float(line.resolved_rate),
            float(amount),
            SALES_LEDGER_NAME,
            VAT_LABEL if line.resolved_vatable else "",
            float(VAT_RATE) if line.resolved_vatable else "",
            float(vat_amount) if vat_amount != "" else "",
        ]

        for col_index, value in enumerate(row):
            if col_index == 0:
                parsed = datetime.strptime(str(value), DATE_FORMAT)
                sheet.write(row_index, col_index, parsed, date_style)
            elif isinstance(value, (int, float)):
                sheet.write(row_index, col_index, value, decimal_style)
            else:
                sheet.write(row_index, col_index, value, text_style)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    run.status = "exported"
    run.exported_at = datetime.now(ZoneInfo(current_app.config["APP_TIMEZONE"]))
    db.session.commit()
    return _build_export_filename(run, selected_category or None), stream.getvalue()


def export_ignored_history_to_xls(run_id: str) -> tuple[str, bytes]:
    summary = build_ignored_history_summary(run_id)
    if summary is None:
        raise WorkbookShapeError("This upload run could not be found.")
    if not summary.ignored_groups:
        raise WorkbookShapeError("There are no ignored inactive items in this run yet.")

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Ignored Items")

    header_style = xlwt.easyxf(
        "font: bold on, height 220;"
        "pattern: pattern solid, fore_colour ocean_blue;"
        "align: horiz center;"
        "borders: left thin, right thin, top thin, bottom thin;"
    )
    text_style = xlwt.easyxf("borders: left thin, right thin, top thin, bottom thin;")
    number_style = xlwt.easyxf(
        "borders: left thin, right thin, top thin, bottom thin;",
        num_format_str="0",
    )

    summary_headers = [
        "Source SKU",
        "Occurrences In This Run",
        "Stores",
        "Order Numbers",
        "Reason",
        "Run File",
        "Invoice Date",
    ]
    for col_index, header in enumerate(summary_headers):
        sheet.write(0, col_index, header, header_style)
        sheet.col(col_index).width = min(max(len(header) + 4, 16) * 256, 42 * 256)

    for row_index, group in enumerate(summary.ignored_groups, start=1):
        row = [
            group.source_sku,
            group.occurrences,
            ", ".join(group.supermarkets),
            ", ".join(group.order_numbers),
            group.reason,
            summary.run.original_filename,
            summary.run.invoice_date,
        ]
        for col_index, value in enumerate(row):
            style = number_style if isinstance(value, int) else text_style
            sheet.write(row_index, col_index, value, style)

    history_start = len(summary.ignored_groups) + 3
    history_headers = [
        "Source SKU",
        "Run Id",
        "Run File",
        "Invoice Date",
        "Store",
        "Order Number",
        "SKU Ignore Count",
        "Run Ignore Count",
    ]
    for col_index, header in enumerate(history_headers):
        sheet.write(history_start, col_index, header, header_style)

    for offset, item in enumerate(summary.ignored_runs_for_sku, start=1):
        row_index = history_start + offset
        row = [
            item["source_sku"],
            item["run_id"],
            item["run_filename"],
            item["invoice_date"],
            item["supermarket_name"],
            item["order_number"],
            item["occurrences_for_sku"],
            item["occurrences_for_run"],
        ]
        for col_index, value in enumerate(row):
            style = number_style if isinstance(value, int) else text_style
            sheet.write(row_index, col_index, value, style)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    run_label = _clean_filename_part(Path(summary.run.original_filename or "tracker.xlsx").stem, "Tracker")
    filename = f"DALA Ignored Items - {summary.run.invoice_date} - {run_label}.xls"
    return filename, stream.getvalue()


def resolve_product_match(source_sku: str, products: list[Product], aliases: list[ProductAlias]) -> ProductMatch | None:
    exact_lookup = {product.sku_name: product for product in products}
    if source_sku in exact_lookup:
        return ProductMatch(exact_lookup[source_sku], "exact")

    alias_lookup = {alias.alias_name: alias for alias in aliases}
    alias = alias_lookup.get(source_sku)
    if alias is not None:
        product = db.session.get(Product, alias.product_id)
        if product is not None:
            return ProductMatch(product, "approved-alias")

    normalized = normalize_sku(source_sku)
    normalized_products = [product for product in products if product.normalized_name == normalized]
    if len(normalized_products) == 1:
        return ProductMatch(normalized_products[0], "normalized")

    normalized_aliases = [alias for alias in aliases if alias.normalized_name == normalized]
    if len(normalized_aliases) == 1:
        product = db.session.get(Product, normalized_aliases[0].product_id)
        if product is not None:
            return ProductMatch(product, "approved-alias")

    sync_key = normalize_uom_sync_key(source_sku)
    sync_products = [product for product in products if normalize_uom_sync_key(product.sku_name) == sync_key]
    if len(sync_products) == 1:
        return ProductMatch(sync_products[0], "sync-key")

    sync_aliases = [alias for alias in aliases if normalize_uom_sync_key(alias.alias_name) == sync_key]
    if len(sync_aliases) == 1:
        product = db.session.get(Product, sync_aliases[0].product_id)
        if product is not None:
            return ProductMatch(product, "approved-alias")

    best_match = _best_high_confidence_product_match(source_sku, products)
    if best_match is not None:
        return ProductMatch(best_match, "high-confidence")

    return None


def suggest_products(source_sku: str, products: list[Product], limit: int = 5) -> list[Product]:
    normalized = normalize_sku(source_sku)
    sync_key = normalize_uom_sync_key(source_sku)
    ranked = []
    for product in products:
        product_sync_key = normalize_uom_sync_key(product.sku_name)
        normalized_score = SequenceMatcher(None, normalized, product.normalized_name).ratio()
        sync_score = SequenceMatcher(None, sync_key, product_sync_key).ratio()
        score = max(normalized_score, sync_score)
        if score >= 0.42:
            ranked.append((score, product))

    ranked.sort(key=lambda item: (-item[0], item[1].sku_name))
    suggestions = []
    seen = set()
    for _, product in ranked:
        if product.id in seen:
            continue
        suggestions.append(product)
        seen.add(product.id)
        if len(suggestions) == limit:
            break
    return suggestions


def normalize_sku(value: str) -> str:
    text = value.upper().replace("'S", "S")
    for source, target in (
        ("LITRES", "LITRE"),
        ("LITERS", "LITRE"),
        ("LTRS", "LITRE"),
        ("LTR", "LITRE"),
    ):
        text = text.replace(source, target)
    cleaned = []
    for character in text:
        cleaned.append(character if character.isalnum() else " ")
    return " ".join("".join(cleaned).split())


def _best_high_confidence_product_match(source_sku: str, products: list[Product]) -> Product | None:
    normalized = normalize_sku(source_sku)
    sync_key = normalize_uom_sync_key(source_sku)
    scored: list[tuple[float, float, float, Product]] = []
    for product in products:
        product_sync_key = normalize_uom_sync_key(product.sku_name)
        normalized_score = SequenceMatcher(None, normalized, product.normalized_name).ratio()
        sync_score = SequenceMatcher(None, sync_key, product_sync_key).ratio()
        score = max(normalized_score, sync_score)
        scored.append((score, sync_score, normalized_score, product))

    if not scored:
        return None

    scored.sort(key=lambda item: (-item[0], -item[1], -item[2], item[3].sku_name.upper()))
    top_score, top_sync_score, _, top_product = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0

    if top_score < 0.86:
        return None
    if top_score - second_score < 0.12 and top_sync_score < 0.97:
        return None
    return top_product


def _contains_match(left: str, right: str) -> bool:
    return bool(left and right and (left in right or right in left))


def _routing_match_score(
    candidate_party: str,
    target_party: str,
    candidate_sku: str,
    target_sku: str,
) -> tuple[int, int, int, int]:
    exact_party = int(candidate_party == target_party)
    exact_sku = int(candidate_sku == target_sku)
    return (
        exact_party + exact_sku,
        exact_party,
        exact_sku,
        min(len(candidate_party), len(target_party)) + min(len(candidate_sku), len(target_sku)),
    )


def normalize_uom_sync_key(value: str) -> str:
    tokens = [token for token in normalize_sku(value).split() if token and token != "X000D"]
    if len(tokens) > 1 and tokens[0].isalpha() and len(tokens[0]) <= 4:
        tokens = tokens[1:]

    core_tokens = [token for token in tokens if re.fullmatch(r"\d+X", token) is None]
    if core_tokens:
        tokens = core_tokens

    return " ".join(sorted(tokens))


def _match_existing_uom_product(
    sku_name: str,
    *,
    existing_by_name: dict[str, Product],
    existing_by_normalized: dict[str, list[Product]],
    existing_by_sync_key: dict[str, list[Product]],
) -> Product | None:
    product = existing_by_name.get(sku_name)
    if product is not None:
        return product

    normalized_name = normalize_sku(sku_name)
    normalized_matches = existing_by_normalized.get(normalized_name, [])
    if len(normalized_matches) == 1:
        return normalized_matches[0]

    sync_key = normalize_uom_sync_key(sku_name)
    sync_matches = existing_by_sync_key.get(sync_key, [])
    if len(sync_matches) == 1:
        return sync_matches[0]
    return None


def _serialize_uom_row(row: list[Any]) -> dict[str, Any]:
    return {
        "sku_name": _string_value(row[0]),
        "uom": _string_value(row[1]),
        "alt_uom": _string_value(row[2]),
        "conversion": _decimal_string(_decimal_value(row[3])),
        "vatable": _string_value(row[4]),
        "price": _decimal_string(_decimal_value(row[5])),
        "is_active": True if len(row) < 7 else bool(row[6]),
    }


def _deserialize_uom_row(row: dict[str, Any]) -> list[Any]:
    return [
        row.get("sku_name", ""),
        row.get("uom", ""),
        row.get("alt_uom", ""),
        row.get("conversion", ""),
        row.get("vatable", ""),
        row.get("price", ""),
        bool(row.get("is_active", True)),
    ]


def _suggest_uom_review_matches(
    missing_products: list[Product],
    unmatched_rows: list[dict[str, Any]],
) -> dict[int, dict[str, Any]]:
    scored_pairs: list[tuple[float, int, str, str]] = []
    for product in missing_products:
        product_sync_key = normalize_uom_sync_key(product.sku_name)
        product_normalized = normalize_sku(product.sku_name)
        for row in unmatched_rows:
            row_sku = row["sku_name"]
            row_sync_key = normalize_uom_sync_key(row_sku)
            row_normalized = normalize_sku(row_sku)
            sync_score = SequenceMatcher(None, product_sync_key, row_sync_key).ratio()
            name_score = SequenceMatcher(None, product_normalized, row_normalized).ratio()
            exact_sync = product_sync_key and product_sync_key == row_sync_key
            if exact_sync:
                score = max(sync_score, name_score, 0.99)
                reason = "Matching sync key after source-name cleanup"
            else:
                score = max(sync_score, name_score)
                reason = "High name similarity"
            if score < 0.62:
                continue
            scored_pairs.append((score, product.id, row_sku, reason))

    row_scores: dict[str, list[tuple[float, int]]] = {}
    for score, product_id, row_sku, _ in scored_pairs:
        row_scores.setdefault(row_sku, []).append((score, product_id))

    suggestions: dict[int, dict[str, Any]] = {}
    used_rows: set[str] = set()
    used_products: set[int] = set()
    for score, product_id, row_sku, reason in sorted(scored_pairs, key=lambda item: (-item[0], item[1], item[2])):
        if product_id in used_products or row_sku in used_rows:
            continue
        competing_scores = sorted(row_scores.get(row_sku, []), key=lambda item: item[0], reverse=True)
        if len(competing_scores) > 1 and competing_scores[0][0] - competing_scores[1][0] < 0.08:
            continue
        suggestions[product_id] = {
            "sku_name": row_sku,
            "reason": reason,
            "score": round(score, 2),
        }
        used_products.add(product_id)
        used_rows.add(row_sku)
    return suggestions


def tomorrow_in_timezone(timezone_name: str) -> datetime:
    return datetime.now(ZoneInfo(timezone_name)) + timedelta(days=1)


def _build_export_filename(run: UploadRun, invoice_category: str | None = None) -> str:
    source_label = _clean_filename_part(Path(run.original_filename or "tracker.xlsx").stem, "Tracker")
    invoice_label = _clean_filename_part(run.invoice_date, "Invoice Date")
    category_label = _clean_filename_part(invoice_category or "All", "All")
    return f"DALA Delivery Note - {invoice_label} - {category_label} - {source_label}.xls"


def _clean_filename_part(value: str, fallback: str) -> str:
    cleaned = re.sub(INVALID_FILENAME_CHARS, " ", str(value))
    cleaned = cleaned.replace("_", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .-")
    if not cleaned:
        return fallback
    return cleaned[:90].rstrip(" .-")


def _build_ignored_groups(grouped_lines: dict[str, list[UploadLine]]) -> list[IgnoredGroup]:
    ignored_groups = []
    for source_sku, lines in grouped_lines.items():
        ignored_groups.append(
            IgnoredGroup(
                source_sku=source_sku,
                occurrences=len(lines),
                order_numbers=sorted({line.order_number for line in lines}),
                supermarkets=sorted({line.supermarket_name for line in lines}),
                reason="Inactive product",
            )
        )
    ignored_groups.sort(key=lambda item: item.source_sku)
    return ignored_groups


def load_brand_partner_rules() -> list[BrandPartnerRule]:
    return list_brand_partner_rules()


def match_invoice_routing_entry(
    *,
    store_name: str | None,
    sku_name: str | None,
    entries: list[InvoiceRoutingEntry] | None = None,
) -> InvoiceRoutingEntry | None:
    normalized_store = normalize_sku(_string_value(store_name))
    normalized_sku = normalize_sku(_string_value(sku_name))
    if not normalized_store or not normalized_sku:
        return None

    routing_entries = entries if entries is not None else load_invoice_routing_entries()
    exact_matches = [
        entry
        for entry in routing_entries
        if entry.normalized_party_name == normalized_store and entry.normalized_sku_name == normalized_sku
    ]
    if len(exact_matches) == 1:
        return exact_matches[0]
    if exact_matches:
        unique_invoices = {entry.normalized_invoice_name for entry in exact_matches}
        if len(unique_invoices) == 1:
            return sorted(exact_matches, key=lambda item: (item.invoice_name.upper(), item.party_name.upper(), item.sku_name.upper()))[0]
        return None

    fallback_matches = [
        entry
        for entry in routing_entries
        if _contains_match(entry.normalized_party_name, normalized_store)
        and _contains_match(entry.normalized_sku_name, normalized_sku)
    ]
    if len(fallback_matches) == 1:
        return fallback_matches[0]
    if not fallback_matches:
        return None

    fallback_matches.sort(
        key=lambda item: (
            -_routing_match_score(item.normalized_party_name, normalized_store, item.normalized_sku_name, normalized_sku),
            item.invoice_name.upper(),
            item.party_name.upper(),
            item.sku_name.upper(),
        )
    )
    best = fallback_matches[0]
    if len(fallback_matches) > 1:
        first_score = _routing_match_score(best.normalized_party_name, normalized_store, best.normalized_sku_name, normalized_sku)
        second = fallback_matches[1]
        second_score = _routing_match_score(second.normalized_party_name, normalized_store, second.normalized_sku_name, normalized_sku)
        if first_score == second_score:
            return None
    return best


def split_prefixed_reference(value: str | None) -> tuple[str | None, str]:
    text = _string_value(value)
    if not text:
        return None, ""
    upper = text.upper()
    for category in [INVOICE_CATEGORY_BPVT, INVOICE_CATEGORY_BPNV, INVOICE_CATEGORY_BP, INVOICE_CATEGORY_VT, INVOICE_CATEGORY_NV]:
        prefix = f"{category}-"
        if upper.startswith(prefix):
            return category, text[len(prefix) :].strip()
    return None, text


def build_prefixed_reference(invoice_category: str | None, raw_reference_no: str | None) -> str | None:
    normalized_category, _, _ = invoice_category_parts(invoice_category)
    category = _string_value(normalized_category or invoice_category).upper()
    raw_reference = _string_value(raw_reference_no)
    if not category or category not in INVOICE_CATEGORIES or not raw_reference:
        return None
    return f"{category}-{raw_reference}"


def invoice_category_parts(invoice_category: str | None) -> tuple[str | None, str | None, str | None]:
    category = _string_value(invoice_category).upper()
    if category == INVOICE_CATEGORY_VT:
        return category, INVOICE_OWNER_DALA, TAX_BUCKET_VT
    if category == INVOICE_CATEGORY_NV:
        return category, INVOICE_OWNER_DALA, TAX_BUCKET_NV
    if category == INVOICE_CATEGORY_BPVT:
        return INVOICE_CATEGORY_BP, INVOICE_OWNER_BP, TAX_BUCKET_VT
    if category == INVOICE_CATEGORY_BPNV:
        return INVOICE_CATEGORY_BP, INVOICE_OWNER_BP, TAX_BUCKET_NV
    if category == INVOICE_CATEGORY_BP:
        return category, INVOICE_OWNER_BP, None
    return None, None, None


def build_invoice_category(invoice_owner: str | None, tax_bucket: str | None) -> str | None:
    owner = _string_value(invoice_owner).upper()
    bucket = _string_value(tax_bucket).upper()
    if owner == INVOICE_OWNER_BP:
        return INVOICE_CATEGORY_BP
    if owner == INVOICE_OWNER_DALA:
        if bucket == TAX_BUCKET_VT:
            return INVOICE_CATEGORY_VT
        if bucket == TAX_BUCKET_NV:
            return INVOICE_CATEGORY_NV
    return None


def normalize_invoice_rule_key(value: str | None) -> list[str]:
    return [token for token in normalize_sku(_string_value(value)).split() if token and token != "X000D"]


def _tokens_match_rule(target_value: str | None, pattern_value: str | None) -> bool:
    pattern_tokens = normalize_invoice_rule_key(pattern_value)
    if not pattern_tokens:
        return True
    target_tokens = set(normalize_invoice_rule_key(target_value))
    return all(token in target_tokens for token in pattern_tokens)


def classify_invoice_line(
    *,
    raw_reference_no: str | None,
    store_name: str | None,
    sku_name: str | None,
    source_sku_name: str | None = None,
    product: Product | None,
    bp_rules: list[BrandPartnerRule] | None = None,
    invoice_routing_entries: list[InvoiceRoutingEntry] | None = None,
    existing_category: str | None = None,
    existing_prefixed_reference_no: str | None = None,
    existing_invoice_route_name: str | None = None,
) -> InvoiceClassification:
    raw_reference = _string_value(raw_reference_no)
    existing = _string_value(existing_category).upper()
    existing_category_name, existing_owner, existing_tax_bucket = invoice_category_parts(existing)
    if existing_category_name == INVOICE_CATEGORY_BP and product is not None and not existing_tax_bucket:
        existing_tax_bucket = TAX_BUCKET_VT if product.vatable else TAX_BUCKET_NV
    if existing_category_name in INVOICE_CATEGORIES:
        normalized_prefixed_reference = build_prefixed_reference(existing_category_name, raw_reference)
        prefixed_reference = (
            normalized_prefixed_reference
            if existing and existing != existing_category_name
            else existing_prefixed_reference_no or normalized_prefixed_reference
        )
        return InvoiceClassification(
            raw_reference_no=raw_reference,
            invoice_owner=existing_owner,
            tax_bucket=existing_tax_bucket,
            invoice_route_name=_string_value(existing_invoice_route_name) or None,
            invoice_category=existing_category_name,
            prefixed_reference_no=prefixed_reference,
            classification_source="prefixed_reference",
            bp_rule_reason=None,
        )

    resolved_sku_name = _string_value(sku_name or (product.sku_name if product is not None else ""))
    invoice_route = None
    for candidate_sku_name in [source_sku_name, resolved_sku_name]:
        if not _string_value(candidate_sku_name):
            continue
        invoice_route = match_invoice_routing_entry(
            store_name=store_name,
            sku_name=candidate_sku_name,
            entries=invoice_routing_entries,
        )
        if invoice_route is not None:
            break
    if invoice_route is not None:
        normalized_invoice_name = invoice_route.normalized_invoice_name or normalize_sku(invoice_route.invoice_name)
        invoice_owner = INVOICE_OWNER_DALA if normalized_invoice_name == INVOICE_OWNER_DALA else INVOICE_OWNER_BP
        tax_bucket = TAX_BUCKET_VT if product is not None and product.vatable else TAX_BUCKET_NV if product is not None else None
        category = build_invoice_category(invoice_owner, tax_bucket)
        return InvoiceClassification(
            raw_reference_no=raw_reference,
            invoice_owner=invoice_owner,
            tax_bucket=tax_bucket,
            invoice_route_name=invoice_route.invoice_name,
            invoice_category=category,
            prefixed_reference_no=build_prefixed_reference(category, raw_reference),
            classification_source="invoice_routing_db",
            bp_rule_reason=invoice_route.invoice_name,
        )

    rules = bp_rules if bp_rules is not None else load_brand_partner_rules()
    matched_bp_reason: str | None = None
    for rule in rules:
        if not rule.is_active:
            continue
        if not _tokens_match_rule(resolved_sku_name, rule.sku_name_pattern):
            continue
        if rule.store_name_pattern and not _tokens_match_rule(store_name, rule.store_name_pattern):
            continue
        matched_bp_reason = rule.rule_name or rule.store_name_pattern or rule.sku_name_pattern
        break

    if product is not None:
        tax_bucket = TAX_BUCKET_VT if product.vatable else TAX_BUCKET_NV
        invoice_owner = INVOICE_OWNER_BP if matched_bp_reason else INVOICE_OWNER_DALA
        category = build_invoice_category(invoice_owner, tax_bucket)
        return InvoiceClassification(
            raw_reference_no=raw_reference,
            invoice_owner=invoice_owner,
            tax_bucket=tax_bucket,
            invoice_route_name=None,
            invoice_category=category,
            prefixed_reference_no=build_prefixed_reference(category, raw_reference),
            classification_source="bp_rule" if matched_bp_reason else "product_vat",
            bp_rule_reason=matched_bp_reason,
        )

    if matched_bp_reason:
        return InvoiceClassification(
            raw_reference_no=raw_reference,
            invoice_owner=INVOICE_OWNER_BP,
            tax_bucket=None,
            invoice_route_name=None,
            invoice_category=INVOICE_CATEGORY_BP,
            prefixed_reference_no=build_prefixed_reference(INVOICE_CATEGORY_BP, raw_reference),
            classification_source="bp_rule",
            bp_rule_reason=matched_bp_reason,
        )

    return InvoiceClassification(
        raw_reference_no=raw_reference,
        invoice_owner=None,
        tax_bucket=None,
        invoice_route_name=None,
        invoice_category=None,
        prefixed_reference_no=existing_prefixed_reference_no or None,
        classification_source=None,
        bp_rule_reason=None,
    )


def apply_invoice_classification_to_record(
    record: Any,
    *,
    product: Product | None,
    store_name: str | None,
    sku_name: str | None,
    source_sku_name: str | None = None,
    raw_reference_no: str | None,
    bp_rules: list[BrandPartnerRule] | None = None,
    invoice_routing_entries: list[InvoiceRoutingEntry] | None = None,
) -> InvoiceClassification:
    parsed_category, stripped_reference = split_prefixed_reference(
        getattr(record, "prefixed_reference_no", None) or getattr(record, "order_reference_no", None)
    )
    classification = classify_invoice_line(
        raw_reference_no=raw_reference_no or stripped_reference,
        store_name=store_name,
        sku_name=sku_name,
        source_sku_name=source_sku_name or getattr(record, "source_sku", None),
        product=product,
        bp_rules=bp_rules,
        invoice_routing_entries=invoice_routing_entries,
        existing_category=parsed_category or getattr(record, "invoice_category", None),
        existing_prefixed_reference_no=getattr(record, "prefixed_reference_no", None)
        or getattr(record, "order_reference_no", None),
        existing_invoice_route_name=getattr(record, "invoice_route_name", None),
    )
    record.raw_reference_no = classification.raw_reference_no or None
    record.invoice_owner = classification.invoice_owner
    record.tax_bucket = classification.tax_bucket
    record.invoice_route_name = classification.invoice_route_name
    record.invoice_category = classification.invoice_category
    record.prefixed_reference_no = classification.prefixed_reference_no
    record.classification_source = classification.classification_source
    record.bp_rule_reason = classification.bp_rule_reason
    return classification


def apply_product_to_line(
    line: UploadLine,
    product: Product,
    match_method: str,
    *,
    bp_rules: list[BrandPartnerRule] | None = None,
    invoice_routing_entries: list[InvoiceRoutingEntry] | None = None,
) -> None:
    line.product_id = product.id
    line.status = "ready"
    line.matched_by = match_method
    line.resolved_sku_name = product.sku_name
    line.resolved_rate = product.price
    line.resolved_vatable = bool(product.vatable)
    classification = apply_invoice_classification_to_record(
        line,
        product=product,
        store_name=line.supermarket_name,
        sku_name=product.sku_name,
        source_sku_name=getattr(line, "source_sku", None),
        raw_reference_no=getattr(line, "raw_reference_no", None) or line.order_number,
        bp_rules=bp_rules,
        invoice_routing_entries=invoice_routing_entries,
    )
    line.order_number = classification.prefixed_reference_no or classification.raw_reference_no or line.order_number


def _refresh_run_totals(run: UploadRun) -> None:
    run.rows_ready = db.session.scalar(
        select(func.count(UploadLine.id)).where(UploadLine.run_id == run.id, UploadLine.status == "ready")
    ) or 0
    run.rows_needing_review = db.session.scalar(
        select(func.count(UploadLine.id)).where(UploadLine.run_id == run.id, UploadLine.status == "needs_review")
    ) or 0
    run.status = "ready" if run.rows_needing_review == 0 else "needs_review"


def _load_workbook_from_upload(file_storage: Any):
    payload = _read_upload_payload(file_storage)
    workbook = _try_load_workbook(payload)
    if workbook is not None:
        return workbook
    raise WorkbookShapeError("The uploaded file could not be read as an Excel workbook.")


def _read_upload_payload(file_storage: Any) -> bytes:
    payload = file_storage.read()
    file_storage.stream.seek(0)
    return payload


def _try_load_workbook(payload: bytes):
    try:
        return load_workbook(BytesIO(payload), data_only=True)
    except Exception:  # pragma: no cover
        return None


def _extract_uom_workbook_rows(workbook: Any) -> list[list[Any]] | None:
    if UOM_SHEET in workbook.sheetnames:
        return _extract_uom_sheet_rows(workbook[UOM_SHEET])

    required_headers = {"item", "uom", "alt uom", "conversion", "vatable", "prices"}
    for sheet in workbook.worksheets:
        headers = [_normalize_header(sheet.cell(1, column_index).value) for column_index in range(1, sheet.max_column + 1)]
        if not required_headers.issubset(set(headers)):
            continue
        return _extract_uom_sheet_rows(sheet)

    return None


def _extract_uom_sheet_rows(sheet: Any) -> list[list[Any]]:
    headers = [_normalize_header(sheet.cell(1, column_index).value) for column_index in range(1, sheet.max_column + 1)]
    header_map = {header: index + 1 for index, header in enumerate(headers) if header}
    required_headers = {"item", "uom", "alt uom", "conversion", "vatable", "prices"}

    if required_headers.issubset(set(header_map)):
        return [
            [
                sheet.cell(row_index, header_map["item"]).value,
                sheet.cell(row_index, header_map["uom"]).value,
                sheet.cell(row_index, header_map["alt uom"]).value,
                sheet.cell(row_index, header_map["conversion"]).value,
                sheet.cell(row_index, header_map["vatable"]).value,
                sheet.cell(row_index, header_map["prices"]).value,
            ]
            for row_index in range(2, sheet.max_row + 1)
        ]

    return [
        [
            sheet.cell(row_index, 1).value,
            sheet.cell(row_index, 2).value,
            sheet.cell(row_index, 3).value,
            sheet.cell(row_index, 4).value,
            sheet.cell(row_index, 5).value,
            sheet.cell(row_index, 6).value,
        ]
        for row_index in range(2, sheet.max_row + 1)
    ]


def _extract_stock_category_summary_rows(workbook: Any) -> list[list[Any]] | None:
    required_headers = {"sku", "quantity", "(alt. units)", "rate"}
    existing_products = list(db.session.scalars(select(Product)))
    by_exact_name = {product.sku_name: product for product in existing_products}
    by_normalized_name = {product.normalized_name: product for product in existing_products}

    for sheet in workbook.worksheets:
        headers = [_normalize_header(sheet.cell(1, column_index).value) for column_index in range(1, sheet.max_column + 1)]
        if not required_headers.issubset(set(headers)):
            continue

        header_map = {header: index + 1 for index, header in enumerate(headers) if header}
        rows: list[list[Any]] = []
        for row_index in range(2, sheet.max_row + 1):
            sku_name = _string_value(sheet.cell(row_index, header_map["sku"]).value)
            if not sku_name:
                continue

            uom = _string_value(sheet.cell(row_index, header_map["quantity"]).value) or "ctn"
            alt_uom = _string_value(sheet.cell(row_index, header_map["(alt. units)"]).value) or "unt"
            price = _decimal_value(sheet.cell(row_index, header_map["rate"]).value)
            if price is None:
                continue

            existing_product = by_exact_name.get(sku_name) or by_normalized_name.get(normalize_sku(sku_name))
            conversion = _extract_pack_conversion(sku_name)
            if conversion is None and existing_product is not None:
                conversion = existing_product.conversion
            vatable = "Yes" if existing_product and existing_product.vatable else "No"

            rows.append([sku_name, uom, alt_uom, conversion, vatable, price, True])

        return rows or None

    return None


def _extract_invoice_routing_rows_from_workbook(workbook: Any) -> list[dict[str, str]] | None:
    for sheet in workbook.worksheets:
        rows = [
            [sheet.cell(row_index, column_index).value for column_index in range(1, sheet.max_column + 1)]
            for row_index in range(1, sheet.max_row + 1)
        ]
        extracted = _extract_invoice_routing_rows_from_rows(rows)
        if extracted is not None:
            return extracted
    return None


def _extract_invoice_routing_rows_from_delimited_payload(payload: bytes) -> list[dict[str, str]] | None:
    text = _decode_text_payload(payload)
    if text is None:
        return None

    rows = [row for row in csv.reader(StringIO(text), delimiter="\t") if any(_string_value(cell) for cell in row)]
    if not rows:
        return None
    return _extract_invoice_routing_rows_from_rows(rows)


def _extract_invoice_routing_rows_from_rows(rows: list[list[Any]]) -> list[dict[str, str]] | None:
    if not rows:
        return None
    header_map = _map_invoice_routing_headers(rows[0])
    if header_map is None:
        return None

    extracted: list[dict[str, str]] = []
    for row in rows[1:]:
        brand_name = _string_value(row[header_map["brand_name"]]) if header_map["brand_name"] < len(row) else ""
        sku_name = _string_value(row[header_map["sku_name"]]) if header_map["sku_name"] < len(row) else ""
        party_name = _string_value(row[header_map["party_name"]]) if header_map["party_name"] < len(row) else ""
        invoice_name = _string_value(row[header_map["invoice_name"]]) if header_map["invoice_name"] < len(row) else ""
        if not sku_name or not party_name or not invoice_name:
            continue
        extracted.append(
            {
                "brand_name": brand_name,
                "sku_name": sku_name,
                "party_name": party_name,
                "invoice_name": invoice_name,
            }
        )
    return extracted or None


def _map_invoice_routing_headers(header_row: list[Any]) -> dict[str, int] | None:
    normalized_headers = {_normalize_header(value): index for index, value in enumerate(header_row)}
    mapped: dict[str, int] = {}
    for field_name, aliases in INVOICE_ROUTING_HEADERS.items():
        index = next((normalized_headers[alias] for alias in aliases if alias in normalized_headers), None)
        if index is None:
            return None
        mapped[field_name] = index
    return mapped


def _extract_item_list_rows(payload: bytes) -> list[list[Any]] | None:
    text = _decode_text_payload(payload)
    if text is None:
        return None

    reader = csv.DictReader(StringIO(text), delimiter="\t")
    if reader.fieldnames is None:
        return None

    normalized_headers = {(_normalize_header(name) if name is not None else "") for name in reader.fieldnames}
    required_headers = {"item name", "cases size", "item ptr"}
    if not required_headers.issubset(normalized_headers):
        return None

    rows: list[list[Any]] = []
    for row in reader:
        cleaned = {_normalize_header(key): value for key, value in row.items() if key is not None}
        sku_name = _string_value(cleaned.get("item name"))
        if not sku_name:
            continue

        status = _string_value(cleaned.get("status")).lower()
        is_active_row = status != "inactive" and status != "deactivated"

        price = _decimal_value(cleaned.get("item ptr"))
        if price is None and is_active_row:
            continue

        tax_rate = _decimal_value(cleaned.get("tax rate"))
        rows.append(
            [
                sku_name,
                "ctn",
                "unt",
                _decimal_value(cleaned.get("cases size")),
                "Yes" if tax_rate is not None and tax_rate > 0 else "No",
                price,
                is_active_row,
            ]
        )

    return rows or None


def _decode_text_payload(payload: bytes) -> str | None:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return None


def _extract_pack_conversion(sku_name: str) -> Decimal | None:
    match = re.search(r"\((\d+(?:\.\d+)?)\s*[xX]\)", sku_name)
    if match is None:
        return None
    return _decimal_value(match.group(1))


def _normalize_header(value: Any) -> str:
    return " ".join(_string_value(value).lower().split())


def _string_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _decimal_value(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _decimal_string(value: Decimal | None) -> str:
    if value is None:
        return ""
    return format(value.normalize(), "f") if value != value.to_integral() else str(value.quantize(Decimal("1")))


def _resolve_tracker_sheet(workbook: Any):
    best_sheet = None
    best_score = 0

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        score = _tracker_sheet_score(sheet)
        if _normalize_header(sheet_name) == TRACKER_SHEET:
            score += 2
        if score > best_score:
            best_score = score
            best_sheet = sheet

    if best_sheet is not None and best_score >= 8:
        return best_sheet

    raise WorkbookShapeError(
        "We could not find the sheet that contains order numbers, store names, and product quantities."
    )


def _looks_like_tracker_sheet(sheet: Any) -> bool:
    return _tracker_sheet_score(sheet) >= 8


def _tracker_sheet_score(sheet: Any) -> int:
    if sheet.max_row < 2 or sheet.max_column < 3:
        return 0

    first_header = _normalize_header(sheet.cell(1, 1).value)
    second_header = _normalize_header(sheet.cell(1, 2).value)
    product_headers = sum(1 for column_index in range(3, sheet.max_column + 1) if _string_value(sheet.cell(1, column_index).value))
    sample_rows = min(sheet.max_row, 8)
    row_identity_hits = 0
    quantity_hits = 0

    for row_index in range(2, sample_rows + 1):
        if _string_value(sheet.cell(row_index, 1).value) and _string_value(sheet.cell(row_index, 2).value):
            row_identity_hits += 1

        for column_index in range(3, min(sheet.max_column, 20) + 1):
            value = _decimal_value(sheet.cell(row_index, column_index).value)
            if value is not None:
                quantity_hits += 1
                break

    score = 0
    if sheet.max_column >= TRACKER_MIN_COLUMNS:
        score += 4
    if first_header in TRACKER_ORDER_HEADERS:
        score += 4
    if second_header in TRACKER_STORE_HEADERS:
        score += 4
    if product_headers >= 5:
        score += 2
    if row_identity_hits >= 1:
        score += 2
    if quantity_hits >= 1:
        score += 4
    return score


def _normalize_header(value: Any) -> str:
    return " ".join(_string_value(value).lower().split())
