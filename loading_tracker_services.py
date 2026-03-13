from __future__ import annotations

import csv
import shutil
import smtplib
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from decimal import Decimal, InvalidOperation
from email.message import EmailMessage
from io import BytesIO
from pathlib import Path
from typing import Any, Callable
from uuid import uuid4

from flask import current_app
from openpyxl import load_workbook
from sqlalchemy import and_, func, or_, select, update

from models import (
    LoadingTrackerDay,
    LoadingTrackerDailyCount,
    LoadingTrackerEvent,
    LoadingTrackerFeeItem,
    LoadingTrackerImport,
    LoadingTrackerImportJob,
    LoadingTrackerInventoryItem,
    LoadingTrackerRow,
    LoadingTrackerRowItem,
    LoadingTrackerTemplate,
    LoadingTrackerTemplateDay,
    LoadingTrackerTemplateFeeItem,
    LoadingTrackerTemplateInventoryItem,
    LoadingTrackerTemplateRow,
    LoadingTrackerTemplateRowItem,
    Product,
    ProductAlias,
    SkuAutomatorLine,
    SkuAutomatorRun,
    UploadLine,
    UploadRun,
    db,
)
from services import (
    DATE_FORMAT,
    apply_invoice_classification_to_record,
    apply_product_to_line,
    build_prefixed_reference,
    invoice_category_parts,
    load_brand_partner_rules,
    resolve_product_match,
    split_prefixed_reference,
    tomorrow_in_timezone,
)

DAY_SHEET_NAMES = ["Mon", "Tues", "Wed", "Thurs", "Fri", "Sat"]
LL_SHEET_MAP = {
    "Mon": "LL Mon",
    "Tues": "LL Tue",
    "Wed": "LL Wed",
    "Thurs": "LLThurs",
    "Fri": "LL Fri",
    "Sat": "LL Sat",
}
PLANNING_SUMMARY_LABELS = {
    "g2g_total": "Expected in G2G For Loading",
    "loaded_total": "TOTAL LOADED OUT FOR DELIVERY",
    "remaining_total": "Remaining Inventory After Loading",
    "expected_store_total": "Expected in Store For Loading",
}
PLANNED_STATE = "planned"
PENDING_STATE = "pending"
PENDING_SENTINEL = "__pending__"
INVALID_FILENAME_CHARS = r'[<>:"/\\|?*\x00-\x1f]'
DEFAULT_BATCHES = ["Load 1", "Load 2", "Load 3", "Load 4", "Unassigned"]
RUNNING_IMPORT_STATUSES = {"queued", "running"}
IMPORT_JOB_STALE_AFTER = timedelta(minutes=15)
PENDING_REASON_OPTIONS = [
    {"code": "stock_shortage", "label": "Insufficient stock"},
    {"code": "receiving_day", "label": "Store does not receive today"},
    {"code": "credit_hold", "label": "Credit or payment concern"},
    {"code": "store_hold", "label": "Store asked us to wait"},
    {"code": "route_priority", "label": "Route or load priority decision"},
    {"code": "manual_hold", "label": "Manual operations hold"},
]
PENDING_REASON_LABELS = {item["code"]: item["label"] for item in PENDING_REASON_OPTIONS}
NON_PRODUCT_HEADERS = {
    "contact",
    "lp",
    "tier",
    "region",
    "weight",
    "value",
    "date",
    "date assigned",
    "customer's name",
    "customers name",
    "location",
    "cartons delivered",
    "no of delivery notes",
    "wjc invoice",
    "no of receipts",
    "products descriptions",
    "load for external delivery",
}


class LoadingTrackerError(Exception):
    pass


@dataclass
class LoadingTrackerSummary:
    import_count: int
    latest_import: LoadingTrackerImport | None
    recent_imports: list[LoadingTrackerImport]
    total_batches: int
    total_active_stores: int
    total_pending_stores: int
    total_fee_rows: int


@dataclass
class LoadingTrackerTemplateSummary:
    active_template: LoadingTrackerTemplate | None
    template_count: int
    total_template_rows: int
    total_template_skus: int


def get_pending_reason_options() -> list[dict[str, str]]:
    return list(PENDING_REASON_OPTIONS)


def create_loading_tracker_import_job(filename: str) -> LoadingTrackerImportJob:
    job = LoadingTrackerImportJob(
        id=uuid4().hex,
        filename=filename,
        status="queued",
        progress_percent=2,
        stage_label="Queued",
    )
    db.session.add(job)
    db.session.commit()
    return job


def get_loading_tracker_import_job(job_id: str) -> LoadingTrackerImportJob | None:
    return db.session.get(LoadingTrackerImportJob, job_id)


def get_active_loading_tracker_import_job() -> LoadingTrackerImportJob | None:
    return db.session.scalar(
        select(LoadingTrackerImportJob)
        .where(LoadingTrackerImportJob.status.in_(tuple(RUNNING_IMPORT_STATUSES)))
        .order_by(LoadingTrackerImportJob.created_at.desc())
        .limit(1)
    )


def serialize_loading_tracker_import_job(job: LoadingTrackerImportJob | None) -> dict[str, Any] | None:
    if job is None:
        return None
    return {
        "id": job.id,
        "filename": job.filename,
        "status": job.status,
        "progress_percent": job.progress_percent or 0,
        "stage_label": job.stage_label or "",
        "error_message": job.error_message or "",
        "tracker_import_id": job.tracker_import_id,
        "created_at": job.created_at.isoformat() if job.created_at else "",
        "updated_at": job.updated_at.isoformat() if job.updated_at else "",
    }


def update_loading_tracker_import_job(
    job_id: str,
    *,
    status: str | None = None,
    progress_percent: int | None = None,
    stage_label: str | None = None,
    error_message: str | None = None,
    tracker_import_id: str | None = None,
) -> None:
    values: dict[str, Any] = {"updated_at": datetime.now(UTC)}
    if status is not None:
        values["status"] = status
    if progress_percent is not None:
        values["progress_percent"] = max(0, min(int(progress_percent), 100))
    if stage_label is not None:
        values["stage_label"] = stage_label
    if error_message is not None:
        values["error_message"] = error_message
    if tracker_import_id is not None:
        values["tracker_import_id"] = tracker_import_id

    with db.engine.begin() as connection:
        connection.execute(
            update(LoadingTrackerImportJob)
            .where(LoadingTrackerImportJob.id == job_id)
            .values(**values)
        )

def claim_loading_tracker_import_job(job_id: str) -> bool:
    stale_cutoff = datetime.now(UTC) - IMPORT_JOB_STALE_AFTER
    values = {
        "status": "running",
        "progress_percent": 6,
        "stage_label": "Opening workbook",
        "error_message": "",
        "updated_at": datetime.now(UTC),
    }
    with db.engine.begin() as connection:
        result = connection.execute(
            update(LoadingTrackerImportJob)
            .where(LoadingTrackerImportJob.id == job_id)
            .where(
                or_(
                    LoadingTrackerImportJob.status == "queued",
                    and_(
                        LoadingTrackerImportJob.status == "running",
                        LoadingTrackerImportJob.updated_at < stale_cutoff,
                        LoadingTrackerImportJob.tracker_import_id.is_(None),
                    ),
                )
            )
            .values(**values)
        )
    return bool(result.rowcount)


def run_loading_tracker_import_job(
    job_id: str,
    workbook_path: str | Path | None = None,
    filename: str | None = None,
) -> None:
    job = get_loading_tracker_import_job(job_id)
    if job is None:
        return
    if not claim_loading_tracker_import_job(job_id):
        refreshed = get_loading_tracker_import_job(job_id)
        if refreshed is None or refreshed.status in {"completed", "failed"}:
            return
        if refreshed.status == "running" and refreshed.updated_at:
            stale_cutoff = datetime.now(UTC) - IMPORT_JOB_STALE_AFTER
            if refreshed.updated_at >= stale_cutoff:
                return
        return

    workbook_path = Path(workbook_path) if workbook_path is not None else _resolve_loading_tracker_job_upload(job_id)
    filename = filename or job.filename
    if workbook_path is None or not workbook_path.exists():
        update_loading_tracker_import_job(
            job_id,
            status="failed",
            progress_percent=100,
            stage_label="Import failed",
            error_message="The queued workbook upload could not be found. Please upload the weekly tracker again.",
        )
        db.session.remove()
        return

    try:
        with workbook_path.open("rb") as handle:
            tracker_import = import_loading_tracker_workbook(
                handle,
                filename=filename,
                progress_callback=lambda progress, stage: update_loading_tracker_import_job(
                    job_id,
                    status="running",
                    progress_percent=progress,
                    stage_label=stage,
                ),
            )
        update_loading_tracker_import_job(
            job_id,
            status="completed",
            progress_percent=100,
            stage_label="Live planning ready",
            tracker_import_id=tracker_import.id,
        )
    except Exception as exc:  # pragma: no cover - background thread safety
        update_loading_tracker_import_job(
            job_id,
            status="failed",
            progress_percent=100,
            stage_label="Import failed",
            error_message=str(exc),
        )
    finally:
        try:
            workbook_path.unlink(missing_ok=True)
        except OSError:
            pass
        db.session.remove()


def reset_loading_tracker_workspace(instance_path: str | Path) -> dict[str, int]:
    active_job = get_active_loading_tracker_import_job()
    if active_job is not None:
        raise LoadingTrackerError("Please wait for the current loading tracker import to finish before clearing the workspace.")

    tracker_imports = list(
        db.session.scalars(select(LoadingTrackerImport).order_by(LoadingTrackerImport.created_at.desc()))
    )
    import_jobs = list(
        db.session.scalars(select(LoadingTrackerImportJob).order_by(LoadingTrackerImportJob.created_at.desc()))
    )
    summary = {
        "imports": len(tracker_imports),
        "days": sum(len(item.days) for item in tracker_imports),
        "rows": sum(len(item.planning_rows) for item in tracker_imports),
        "events": sum(len(item.events) for item in tracker_imports),
        "jobs": len(import_jobs),
    }

    for job in import_jobs:
        db.session.delete(job)
    for tracker_import in tracker_imports:
        db.session.delete(tracker_import)
    db.session.commit()

    upload_root = Path(instance_path) / "loading_tracker_jobs"
    if upload_root.exists():
        for child in upload_root.iterdir():
            if child.is_dir():
                shutil.rmtree(child, ignore_errors=True)
            else:
                child.unlink(missing_ok=True)

    return summary


def build_loading_tracker_summary() -> LoadingTrackerSummary:
    latest_import = db.session.scalar(
        select(LoadingTrackerImport).order_by(LoadingTrackerImport.created_at.desc()).limit(1)
    )
    total_batches = 0
    total_active_stores = 0
    total_pending_stores = 0
    total_fee_rows = 0

    if latest_import is not None:
        if latest_import.planning_rows:
            total_batches = sum(
                len({row.batch_name for row in day.planning_rows if row.row_state == PLANNED_STATE})
                for day in latest_import.days
            )
            total_active_stores = sum(
                len([row for row in day.planning_rows if row.row_state == PLANNED_STATE])
                for day in latest_import.days
            )
            total_pending_stores = len([row for row in latest_import.planning_rows if row.row_state == PENDING_STATE])
            total_fee_rows = len(latest_import.fee_items)
        else:
            total_batches = sum(day.batch_count for day in latest_import.days)
            total_active_stores = sum(day.active_store_count for day in latest_import.days)
            total_pending_stores = len(latest_import.pending_rows_json or [])
            total_fee_rows = latest_import.fees_row_count or 0

    return LoadingTrackerSummary(
        import_count=db.session.scalar(select(func.count(LoadingTrackerImport.id))) or 0,
        latest_import=latest_import,
        recent_imports=list(
            db.session.scalars(select(LoadingTrackerImport).order_by(LoadingTrackerImport.created_at.desc()).limit(5))
        ),
        total_batches=total_batches,
        total_active_stores=total_active_stores,
        total_pending_stores=total_pending_stores,
        total_fee_rows=total_fee_rows,
    )


def get_loading_tracker_template(template_id: str | None = None) -> LoadingTrackerTemplate | None:
    if template_id:
        return db.session.get(LoadingTrackerTemplate, template_id)
    return db.session.scalar(
        select(LoadingTrackerTemplate)
        .where(LoadingTrackerTemplate.is_active.is_(True))
        .order_by(LoadingTrackerTemplate.updated_at.desc(), LoadingTrackerTemplate.created_at.desc())
        .limit(1)
    )


def build_loading_tracker_template_summary() -> LoadingTrackerTemplateSummary:
    active_template = get_loading_tracker_template()
    return LoadingTrackerTemplateSummary(
        active_template=active_template,
        template_count=db.session.scalar(select(func.count(LoadingTrackerTemplate.id))) or 0,
        total_template_rows=len(active_template.rows) if active_template is not None else 0,
        total_template_skus=len(active_template.inventory_items) if active_template is not None else 0,
    )


def build_loading_tracker_template_context(template: LoadingTrackerTemplate | None) -> dict[str, Any]:
    if template is None:
        return {
            "template": None,
            "day_count": 0,
            "row_count": 0,
            "pending_baseline_count": 0,
            "inventory_seed_count": 0,
            "fee_row_count": 0,
            "day_cards": [],
            "future_only": True,
        }

    day_cards = []
    for day in template.days:
        rows = [row for row in day.rows if row.row_state == PLANNED_STATE]
        day_cards.append(
            {
                "day_name": day.day_name,
                "row_count": len(rows),
                "store_count": len(rows),
                "sku_count": len({item.sku_name for row in rows for item in row.items}),
            }
        )

    return {
        "template": template,
        "day_count": len(template.days),
        "row_count": len([row for row in template.rows if row.row_state == PLANNED_STATE]),
        "pending_baseline_count": len([row for row in template.rows if row.row_state == PENDING_STATE]),
        "inventory_seed_count": len(template.inventory_items),
        "fee_row_count": len(template.fee_items),
        "day_cards": day_cards,
        "future_only": True,
    }


def capture_loading_tracker_template(source_import_id: str | None = None, *, name: str | None = None) -> LoadingTrackerTemplate:
    source_import = get_loading_tracker_import(source_import_id)
    if source_import is None:
        raise LoadingTrackerError("There is no live loading week to save as a backend template yet.")

    db.session.execute(update(LoadingTrackerTemplate).values(is_active=False))
    template = LoadingTrackerTemplate(
        id=uuid4().hex,
        name=name or f"{source_import.week_label} template",
        description="Future weeks use this structure by default. Current live weeks stay unchanged unless a planner chooses otherwise.",
        is_active=True,
        source_import_label=source_import.week_label,
        assumptions_sku_count=source_import.assumptions_sku_count,
        assumptions_store_count=source_import.assumptions_store_count,
        fees_row_count=source_import.fees_row_count,
        notes_count=source_import.notes_count,
    )
    db.session.add(template)
    db.session.flush()

    template_days: dict[str, LoadingTrackerTemplateDay] = {}
    for day in source_import.days:
        template_day = LoadingTrackerTemplateDay(
            template_id=template.id,
            day_name=day.day_name,
            day_order=day.day_order,
        )
        db.session.add(template_day)
        db.session.flush()
        template_days[day.day_name] = template_day

    for row in source_import.planning_rows:
        if row.row_state != PLANNED_STATE or row.day is None:
            continue
        template_row = LoadingTrackerTemplateRow(
            template_id=template.id,
            day_id=template_days[row.day.day_name].id,
            row_state=PLANNED_STATE,
            batch_name=row.batch_name,
            store_name=row.store_name,
            contact=row.contact,
            lp=row.lp,
            tier=row.tier,
            region=row.region,
            delivery_date=row.delivery_date,
            reason_text=row.reason_text,
            total_weight=row.total_weight,
            total_value=row.total_value,
            sort_order=row.sort_order,
        )
        db.session.add(template_row)
        db.session.flush()
        for item in row.items:
            db.session.add(
                LoadingTrackerTemplateRowItem(
                    row_id=template_row.id,
                    sku_name=item.sku_name,
                    quantity=item.quantity,
                )
            )

    inventory_seed_names = {
        item.sku_name
        for item in source_import.inventory_items
    } | {
        item.sku_name
        for row in source_import.planning_rows
        if row.row_state == PLANNED_STATE
        for item in row.items
    }
    for sort_order, sku_name in enumerate(sorted(inventory_seed_names), start=1):
        db.session.add(
            LoadingTrackerTemplateInventoryItem(
                template_id=template.id,
                sku_name=sku_name,
                sort_order=sort_order,
            )
        )

    for fee_item in source_import.fee_items:
        db.session.add(
            LoadingTrackerTemplateFeeItem(
                template_id=template.id,
                brand_partner=fee_item.brand_partner,
                sku_name=fee_item.sku_name,
                vatable_text=fee_item.vatable_text,
                retail_delivery_value=fee_item.retail_delivery_value,
                payment_collection_value=fee_item.payment_collection_value,
            )
        )

    db.session.commit()
    return template


def create_loading_tracker_week_from_template(
    template_id: str | None = None,
    *,
    source_import_id: str | None = None,
    week_label: str | None = None,
    include_template_rows: bool = True,
) -> LoadingTrackerImport:
    template = get_loading_tracker_template(template_id)
    if template is None:
        raise LoadingTrackerError("There is no backend planning template yet. Save a live week as template first.")

    source_import = get_loading_tracker_import(source_import_id)
    remaining_map = _closing_inventory_by_import(source_import) if source_import is not None else {}
    pending_rows = [row for row in source_import.planning_rows if row.row_state == PENDING_STATE] if source_import else []

    tracker_import = LoadingTrackerImport(
        id=uuid4().hex,
        filename=f"Template week - {template.name}",
        week_label=week_label or f"Week of {datetime.now().date()}",
        assumptions_sku_count=template.assumptions_sku_count,
        assumptions_store_count=template.assumptions_store_count,
        opening_g2g_total=_decimal_or_none(sum(remaining_map.values())),
        opening_remaining_total=_decimal_or_none(sum(remaining_map.values())),
        fees_row_count=template.fees_row_count,
        notes_count=template.notes_count,
    )
    db.session.add(tracker_import)
    db.session.flush()

    day_lookup: dict[str, LoadingTrackerDay] = {}
    for template_day in template.days:
        day_record = LoadingTrackerDay(
            tracker_import_id=tracker_import.id,
            day_name=template_day.day_name,
            day_order=template_day.day_order,
        )
        db.session.add(day_record)
        db.session.flush()
        day_lookup[template_day.day_name] = day_record

    inventory_seed_names = {item.sku_name for item in template.inventory_items} | set(remaining_map)
    if not inventory_seed_names:
        inventory_seed_names = {product.sku_name for product in db.session.query(Product).filter(Product.is_active.is_(True))}
    for sort_order, sku_name in enumerate(sorted(inventory_seed_names), start=1):
        qty = round(remaining_map.get(sku_name, 0.0), 4)
        db.session.add(
            LoadingTrackerInventoryItem(
                tracker_import_id=tracker_import.id,
                sku_name=sku_name,
                opening_g2g_qty=_decimal_or_none(qty) or Decimal("0"),
                opening_remaining_qty=_decimal_or_none(qty) or Decimal("0"),
                added_qty=Decimal("0"),
                sort_order=sort_order,
            )
        )

    for fee_item in template.fee_items:
        db.session.add(
            LoadingTrackerFeeItem(
                tracker_import_id=tracker_import.id,
                brand_partner=fee_item.brand_partner,
                sku_name=fee_item.sku_name,
                vatable_text=fee_item.vatable_text,
                retail_delivery_value=fee_item.retail_delivery_value,
                payment_collection_value=fee_item.payment_collection_value,
            )
        )

    day_sort_orders: dict[str, int] = {day_name: 0 for day_name in day_lookup}
    if include_template_rows:
        for template_row in template.rows:
            if template_row.row_state != PLANNED_STATE or template_row.day is None:
                continue
            row_data = _serialize_template_row(template_row)
            day_name = template_row.day.day_name
            day_sort_orders[day_name] += 1
            _save_row_record(
                tracker_import=tracker_import,
                row_data=row_data,
                row_state=PLANNED_STATE,
                day=day_lookup[day_name],
                sort_order=day_sort_orders[day_name],
                source_kind="template",
            )

    for sort_order, row in enumerate(pending_rows, start=1):
        _save_row_record(
            tracker_import=tracker_import,
            row_data=_serialize_row(row),
            row_state=PENDING_STATE,
            day=None,
            sort_order=sort_order,
            source_kind="carry_forward",
            reason_text=row.reason_text,
        )

    _log_tracker_event(
        tracker_import=tracker_import,
        day=None,
        row=None,
        event_type="started_week_from_template",
        entity_name=template.name,
        details={
            "template_id": template.id,
            "template_name": template.name,
            "carried_pending_rows": len(pending_rows),
            "carried_inventory_skus": len(remaining_map),
        },
    )
    db.session.commit()
    return tracker_import


def create_loading_tracker_week_from_sku_automator_run(
    run_id: str,
    *,
    target_day_name: str | None = None,
    template_id: str | None = None,
    source_import_id: str | None = None,
    week_label: str | None = None,
) -> LoadingTrackerImport:
    run = db.session.get(SkuAutomatorRun, run_id)
    if run is None:
        raise LoadingTrackerError("The selected SKU Automator run could not be found.")
    if run.rows_needing_review > 0:
        raise LoadingTrackerError("Resolve all SKU Automator review items before opening it in Loading Tracker.")

    ready_lines = list(
        db.session.scalars(
            select(SkuAutomatorLine)
            .where(SkuAutomatorLine.run_id == run_id, SkuAutomatorLine.status == "ready")
            .order_by(SkuAutomatorLine.store_name.asc(), SkuAutomatorLine.id.asc())
        )
    )
    if not ready_lines:
        raise LoadingTrackerError("This SKU Automator run has no ready rows to seed into Loading Tracker.")

    latest_import = get_loading_tracker_import(source_import_id) if source_import_id else get_loading_tracker_import()
    tracker_import = create_loading_tracker_week_from_template(
        template_id,
        source_import_id=latest_import.id if latest_import is not None else None,
        week_label=week_label or f"Planner week from {Path(run.original_filename).stem}",
        include_template_rows=False,
    )
    tracker_import.filename = f"SKU Automator - {run.original_filename}"

    target_day = get_loading_tracker_day(tracker_import.id, target_day_name or "")
    if target_day is None:
        target_day = tracker_import.days[0] if tracker_import.days else None
    if target_day is None:
        raise LoadingTrackerError("The Loading Tracker template does not contain any planning days yet.")

    existing_inventory = {item.sku_name for item in tracker_import.inventory_items}
    next_inventory_order = max((item.sort_order for item in tracker_import.inventory_items), default=0)
    grouped_rows = _group_sku_automator_store_rows(ready_lines)
    next_sort_order = max(
        (row.sort_order for row in tracker_import.planning_rows if row.row_state == PLANNED_STATE and row.day_id == target_day.id),
        default=0,
    )

    for store_row in grouped_rows:
        for sku_name in {item["sku"] for item in store_row["items"]}:
            if sku_name in existing_inventory:
                continue
            next_inventory_order += 1
            existing_inventory.add(sku_name)
            db.session.add(
                LoadingTrackerInventoryItem(
                    tracker_import_id=tracker_import.id,
                    sku_name=sku_name,
                    opening_g2g_qty=Decimal("0"),
                    opening_remaining_qty=Decimal("0"),
                    added_qty=Decimal("0"),
                    sort_order=next_inventory_order,
                )
            )

        next_sort_order += 1
        _save_row_record(
            tracker_import=tracker_import,
            row_data=store_row,
            row_state=PLANNED_STATE,
            day=target_day,
            sort_order=next_sort_order,
            source_kind="sku_automator",
        )

    _log_tracker_event(
        tracker_import=tracker_import,
        day=target_day,
        row=None,
        event_type="started_week_from_sku_automator",
        entity_name=run.original_filename,
        details={
            "run_id": run.id,
            "store_rows_added": len(grouped_rows),
            "target_day_name": target_day.day_name,
            "carried_forward_from_import_id": latest_import.id if latest_import is not None else None,
        },
    )
    db.session.commit()
    return tracker_import


def import_loading_tracker_workbook(
    file_storage: Any,
    *,
    filename: str | None = None,
    progress_callback: Callable[[int, str], None] | None = None,
) -> LoadingTrackerImport:
    _report_import_progress(progress_callback, 8, "Opening workbook")
    workbook = _load_workbook_from_upload(file_storage)
    filename = filename or getattr(file_storage, "filename", None) or "loading-tracker.xlsx"
    imported_pending_count = 0

    present_day_sheets = [name for name in DAY_SHEET_NAMES if name in workbook.sheetnames]
    if not present_day_sheets:
        raise LoadingTrackerError("The workbook must include at least one day sheet like Mon or Tues.")

    tracker_import = LoadingTrackerImport(
        id=uuid4().hex,
        filename=filename,
        week_label=_clean_filename_part(Path(filename).stem, "Loading Tracker"),
    )
    db.session.add(tracker_import)

    assumptions_sheet = workbook["Assumptions"] if "Assumptions" in workbook.sheetnames else None
    if assumptions_sheet is not None:
        _report_import_progress(progress_callback, 16, "Reading assumptions")
        assumptions = _parse_assumptions_sheet(assumptions_sheet)
        tracker_import.assumptions_sku_count = assumptions["sku_count"]
        tracker_import.assumptions_store_count = assumptions["store_count"]

    opening_sheet = workbook["Opening Inventory"] if "Opening Inventory" in workbook.sheetnames else None
    if opening_sheet is not None:
        _report_import_progress(progress_callback, 24, "Reading opening inventory")
        opening = _parse_inventory_sheet(opening_sheet)
        tracker_import.opening_g2g_total = opening["g2g_total"]
        tracker_import.opening_remaining_total = opening["remaining_total"]
        tracker_import.opening_top_products_json = opening["top_products"]
        for sort_order, item in enumerate(opening["inventory_items"], start=1):
            db.session.add(
                LoadingTrackerInventoryItem(
                    tracker_import_id=tracker_import.id,
                    sku_name=item["sku"],
                    opening_g2g_qty=_decimal_or_none(item["opening_g2g_qty"]),
                    opening_remaining_qty=_decimal_or_none(item["opening_remaining_qty"]),
                    added_qty=Decimal("0"),
                    sort_order=sort_order,
                )
            )

    pending_sheet = workbook["Pending Orders"] if "Pending Orders" in workbook.sheetnames else None
    if pending_sheet is not None:
        _report_import_progress(progress_callback, 34, "Reading pending lines")
        pending = _parse_support_sheet(pending_sheet)
        tracker_import.pending_g2g_total = pending["g2g_total"]
        tracker_import.pending_loaded_total = pending["loaded_total"]
        tracker_import.pending_remaining_total = pending["remaining_total"]
        tracker_import.pending_rows_json = pending["store_rows"]
        tracker_import.pending_top_products_json = pending["top_products"]
        imported_pending_count = len(pending["store_rows"])
        for sort_order, row_data in enumerate(pending["store_rows"], start=1):
            _save_row_record(
                tracker_import=tracker_import,
                row_data=row_data,
                row_state=PENDING_STATE,
                day=None,
                sort_order=sort_order,
                source_kind="import",
                reason_text=row_data.get("reason_text") or "Imported from Pending Orders",
            )

    fee_sheet = _resolve_fee_sheet(workbook)
    if fee_sheet is not None:
        _report_import_progress(progress_callback, 44, "Reading fee controls")
        fees = _parse_fee_sheet(fee_sheet)
        tracker_import.fees_row_count = fees["row_count"]
        tracker_import.fees_total_delivery_value = fees["total_delivery_value"]
        tracker_import.fees_total_payment_value = fees["total_payment_value"]
        tracker_import.fee_rows_json = fees["top_rows"]
        for row in fees["rows"]:
            db.session.add(
                LoadingTrackerFeeItem(
                    tracker_import_id=tracker_import.id,
                    brand_partner=row["brand_partner"] or None,
                    sku_name=row["sku"],
                    vatable_text=row["vatable"] or None,
                    retail_delivery_value=_decimal_or_none(row["retail_delivery_value"]),
                    payment_collection_value=_decimal_or_none(row["payment_collection_value"]),
                )
            )

    notes_sheet = workbook["NOTES FOR USER"] if "NOTES FOR USER" in workbook.sheetnames else None
    if notes_sheet is not None:
        tracker_import.notes_count = _count_note_lines(notes_sheet)

    total_days = max(len(present_day_sheets), 1)
    for order, day_name in enumerate(present_day_sheets, start=1):
        day_progress = 50 + int(((order - 1) / total_days) * 38)
        _report_import_progress(progress_callback, day_progress, f"Reading {day_name} planning")
        planning_sheet = workbook[day_name]
        load_sheet = workbook[LL_SHEET_MAP[day_name]] if LL_SHEET_MAP.get(day_name) in workbook.sheetnames else None
        parsed_day = _parse_day_sheet(planning_sheet, day_name, load_sheet)
        day_record = LoadingTrackerDay(
            tracker_import=tracker_import,
            day_name=day_name,
            day_order=order,
            g2g_total=parsed_day["g2g_total"],
            loaded_total=parsed_day["loaded_total"],
            remaining_total=parsed_day["remaining_total"],
            expected_store_total=parsed_day["expected_store_total"],
            batch_count=parsed_day["batch_count"],
            active_store_count=parsed_day["active_store_count"],
            total_weight=parsed_day["total_weight"],
            total_value=parsed_day["total_value"],
            load_1_total=parsed_day["load_1_total"],
            load_2_total=parsed_day["load_2_total"],
            load_3_total=parsed_day["load_3_total"],
            load_4_total=parsed_day["load_4_total"],
            load_total=parsed_day["load_total"],
            store_rows_json=parsed_day["store_rows"],
            top_products_json=parsed_day["top_products"],
            load_rows_json=parsed_day["load_rows"],
        )
        db.session.add(day_record)

        for sort_order, row_data in enumerate(parsed_day["store_rows"], start=1):
            _save_row_record(
                tracker_import=tracker_import,
                row_data=row_data,
                row_state=PLANNED_STATE,
                day=day_record,
                sort_order=sort_order,
                source_kind="import",
            )

    _report_import_progress(progress_callback, 96, "Saving live planning")
    _log_tracker_event(
        tracker_import=tracker_import,
        day=None,
        row=None,
        event_type="imported_weekly_tracker",
        entity_name=tracker_import.week_label,
        details={
            "filename": tracker_import.filename,
            "day_count": len(present_day_sheets),
            "pending_rows": imported_pending_count,
        },
    )
    db.session.commit()
    _report_import_progress(progress_callback, 99, "Finalizing")
    return tracker_import


def get_loading_tracker_import(import_id: str | None = None) -> LoadingTrackerImport | None:
    if import_id:
        return db.session.get(LoadingTrackerImport, import_id)
    return db.session.scalar(select(LoadingTrackerImport).order_by(LoadingTrackerImport.created_at.desc()).limit(1))


def get_loading_tracker_day(import_id: str, day_name: str) -> LoadingTrackerDay | None:
    return db.session.scalar(
        select(LoadingTrackerDay)
        .where(LoadingTrackerDay.tracker_import_id == import_id, LoadingTrackerDay.day_name == day_name)
        .limit(1)
    )


def get_loading_tracker_row(row_id: int) -> LoadingTrackerRow | None:
    return db.session.get(LoadingTrackerRow, row_id)


def build_loading_tracker_overview(tracker_import: LoadingTrackerImport | None) -> dict[str, Any]:
    if tracker_import is None:
        return {
            "day_count": 0,
            "total_batches": 0,
            "total_active_stores": 0,
            "total_pending_rows": 0,
            "total_load_value": Decimal("0"),
            "largest_day_name": None,
            "largest_day_value": Decimal("0"),
            "live_ready": False,
        }

    if tracker_import.planning_rows:
        day_contexts = [build_loading_tracker_day_context(day) for day in tracker_import.days]
        largest_day = max(day_contexts, key=lambda item: item["metrics"]["total_value"], default=None)
        return {
            "day_count": len(tracker_import.days),
            "total_batches": sum(item["metrics"]["batch_count"] for item in day_contexts),
            "total_active_stores": sum(item["metrics"]["active_store_count"] for item in day_contexts),
            "total_pending_rows": len([row for row in tracker_import.planning_rows if row.row_state == PENDING_STATE]),
            "total_load_value": sum((item["metrics"]["total_value_decimal"] for item in day_contexts), Decimal("0")),
            "largest_day_name": largest_day["day"].day_name if largest_day else None,
            "largest_day_value": largest_day["metrics"]["total_value_decimal"] if largest_day else Decimal("0"),
            "live_ready": True,
        }

    days = tracker_import.days or []
    largest_day = max(days, key=lambda day: day.total_value or 0, default=None)
    return {
        "day_count": len(days),
        "total_batches": sum(day.batch_count for day in days),
        "total_active_stores": sum(day.active_store_count for day in days),
        "total_pending_rows": len(tracker_import.pending_rows_json or []),
        "total_load_value": sum((day.total_value or Decimal("0")) for day in days),
        "largest_day_name": largest_day.day_name if largest_day else None,
        "largest_day_value": largest_day.total_value or Decimal("0") if largest_day else Decimal("0"),
        "live_ready": False,
    }


def build_loading_tracker_day_context(day: LoadingTrackerDay) -> dict[str, Any]:
    tracker_import = day.tracker_import
    planned_rows = [row for row in day.planning_rows if row.row_state == PLANNED_STATE]
    serialized_rows = [_serialize_row(row) for row in planned_rows]
    grouped_batches = group_store_rows_by_batch(serialized_rows)
    expected_start = _expected_start_by_day(tracker_import, day)
    for sku_name in _aggregate_row_item_totals(planned_rows):
        expected_start.setdefault(sku_name, 0.0)
    count_lookup, discrepancy_rows, count_ready = _day_count_snapshot(day, expected_start)
    available_start = count_lookup if count_ready else expected_start
    consumed_today = _aggregate_row_item_totals(planned_rows)
    remaining_after = _subtract_maps(available_start, consumed_today)
    ll_rows, load_totals = _build_ll_rows(serialized_rows)
    pack_breaker_rows, pack_breaker_batches = _build_pack_breaker_rows(serialized_rows)
    top_products = _sorted_top_products(consumed_today, 12)
    inventory_warnings = [
        {"sku": sku, "remaining": round(quantity, 2)}
        for sku, quantity in sorted(remaining_after.items(), key=lambda item: item[1])
        if quantity <= 0
    ][:8]
    suggestions = _build_day_suggestions(planned_rows, remaining_after)

    total_weight = round(sum(row["weight"] for row in serialized_rows), 2)
    total_value = round(sum(row["value"] for row in serialized_rows), 2)
    total_quantity = round(sum(row["total_quantity"] for row in serialized_rows), 2)

    return {
        "day": day,
        "metrics": {
            "g2g_total": round(sum(available_start.values()), 2),
            "loaded_total": total_quantity,
            "remaining_total": round(sum(remaining_after.values()), 2),
            "expected_store_total": total_quantity,
            "batch_count": len(grouped_batches),
            "active_store_count": len(serialized_rows),
            "total_weight": total_weight,
            "total_value": total_value,
            "total_value_decimal": Decimal(f"{total_value:.4f}"),
            "load_1_total": round(load_totals.get("load_1", 0.0), 2),
            "load_2_total": round(load_totals.get("load_2", 0.0), 2),
            "load_3_total": round(load_totals.get("load_3", 0.0), 2),
            "load_4_total": round(load_totals.get("load_4", 0.0), 2),
            "load_total": round(load_totals.get("total", 0.0), 2),
        },
        "rows": serialized_rows,
        "grouped_batches": grouped_batches,
        "ll_rows": ll_rows,
        "pack_breaker_rows": pack_breaker_rows,
        "pack_breaker_batches": pack_breaker_batches,
        "top_products": top_products,
        "inventory_warnings": inventory_warnings,
        "count_rows": discrepancy_rows,
        "count_ready": count_ready,
        "count_discrepancy_total": round(sum(abs(row["discrepancy_qty"]) for row in discrepancy_rows), 2),
        "count_discrepancy_rows": len([row for row in discrepancy_rows if abs(row["discrepancy_qty"]) > 0.0001]),
        "day_options": [{"value": item.day_name, "label": item.day_name} for item in tracker_import.days],
        "bulk_target_options": [{"value": item.day_name, "label": item.day_name} for item in tracker_import.days]
        + [{"value": PENDING_SENTINEL, "label": "Pending"}],
        "bulk_reason_options": get_pending_reason_options(),
        "suggestions": suggestions,
    }


def build_loading_tracker_pending_context(tracker_import: LoadingTrackerImport) -> dict[str, Any]:
    pending_rows = [row for row in tracker_import.planning_rows if row.row_state == PENDING_STATE]
    serialized_rows = [_serialize_row(row) for row in pending_rows]
    grouped_batches = group_store_rows_by_batch(serialized_rows)
    top_products = _sorted_top_products(_aggregate_row_item_totals(pending_rows), 12)
    reasons: dict[str, int] = {}
    for row in pending_rows:
        key = row.reason_text or "Waiting for planning decision"
        reasons[key] = reasons.get(key, 0) + 1
    return {
        "rows": serialized_rows,
        "grouped_batches": grouped_batches,
        "top_products": top_products,
        "reason_breakdown": [{"reason": reason, "count": count} for reason, count in sorted(reasons.items(), key=lambda item: (-item[1], item[0]))],
        "total_rows": len(serialized_rows),
        "total_quantity": round(sum(row["total_quantity"] for row in serialized_rows), 2),
        "total_value": round(sum(row["value"] for row in serialized_rows), 2),
        "day_options": [{"value": item.day_name, "label": item.day_name} for item in tracker_import.days],
        "reason_options": get_pending_reason_options(),
    }


def build_loading_tracker_inventory_context(tracker_import: LoadingTrackerImport) -> dict[str, Any]:
    inventory_rows: list[dict[str, Any]] = []
    inventory_totals = _inventory_totals_by_sku(tracker_import)
    planned_totals = _aggregate_row_item_totals(_rows_for_days(tracker_import.days))
    all_skus = sorted(set(inventory_totals) | set(planned_totals))
    inventory_lookup = {item.sku_name: item for item in tracker_import.inventory_items}
    closing_inventory = _closing_inventory_by_import(tracker_import)
    day_count_status = [
        {
            "day_name": day.day_name,
            "captured": bool(day.inventory_counts),
            "discrepancy_total": round(
                sum(abs(_float_value(item.discrepancy_qty)) for item in day.inventory_counts),
                2,
            ),
        }
        for day in tracker_import.days
    ]

    for sku in all_skus:
        record = inventory_lookup.get(sku)
        opening = _float_value(record.opening_g2g_qty) if record else 0.0
        added = _float_value(record.added_qty) if record else 0.0
        planned = round(planned_totals.get(sku, 0.0), 2)
        remaining = round(closing_inventory.get(sku, opening + added - planned), 2)
        inventory_rows.append(
            {
                "id": record.id if record else None,
                "sku": sku,
                "opening_g2g_qty": round(opening, 2),
                "opening_remaining_qty": round(_float_value(record.opening_remaining_qty), 2) if record else 0.0,
                "added_qty": round(added, 2),
                "planned_qty": planned,
                "remaining_qty": remaining,
                "status": "tight" if remaining <= 0 else "healthy",
            }
        )

    inventory_rows.sort(key=lambda item: (item["remaining_qty"], item["sku"]))
    return {
        "rows": inventory_rows,
        "g2g_total": round(sum(row["opening_g2g_qty"] + row["added_qty"] for row in inventory_rows), 2),
        "planned_total": round(sum(row["planned_qty"] for row in inventory_rows), 2),
        "remaining_total": round(sum(row["remaining_qty"] for row in inventory_rows), 2),
        "tight_count": len([row for row in inventory_rows if row["remaining_qty"] <= 0]),
        "day_count_status": day_count_status,
    }


def build_loading_tracker_fees_context(tracker_import: LoadingTrackerImport) -> dict[str, Any]:
    fee_rows = [
        {
            "brand_partner": item.brand_partner or "",
            "sku": item.sku_name,
            "vatable": item.vatable_text or "",
            "retail_delivery_value": round(_float_value(item.retail_delivery_value), 2),
            "payment_collection_value": round(_float_value(item.payment_collection_value), 2),
        }
        for item in tracker_import.fee_items
    ]
    fee_rows.sort(key=lambda row: (-row["retail_delivery_value"], row["sku"]))
    return {
        "rows": fee_rows,
        "row_count": len(fee_rows),
        "total_delivery_value": round(sum(row["retail_delivery_value"] for row in fee_rows), 2),
        "total_payment_value": round(sum(row["payment_collection_value"] for row in fee_rows), 2),
    }


def build_loading_tracker_row_editor(
    tracker_import: LoadingTrackerImport,
    row: LoadingTrackerRow | None = None,
    selected_day_name: str | None = None,
) -> dict[str, Any]:
    if row is not None:
        current_day_name = row.day.day_name if row.day is not None else PENDING_SENTINEL
        selected_day_name = selected_day_name or current_day_name
    else:
        selected_day_name = selected_day_name or tracker_import.days[0].day_name

    return {
        "row": row,
        "selected_day_name": selected_day_name,
        "day_options": [{"value": day.day_name, "label": day.day_name} for day in tracker_import.days]
        + [{"value": PENDING_SENTINEL, "label": "Pending"}],
        "batch_options": DEFAULT_BATCHES,
        "items_text": _row_items_text(row) if row is not None else "",
        "reason_options": get_pending_reason_options(),
    }


def save_loading_tracker_row(
    tracker_import_id: str,
    form_data: dict[str, Any],
    row_id: int | None = None,
) -> LoadingTrackerRow:
    tracker_import = get_loading_tracker_import(tracker_import_id)
    if tracker_import is None:
        raise LoadingTrackerError("The selected loading tracker import could not be found.")

    row = get_loading_tracker_row(row_id) if row_id is not None else None
    if row_id is not None and (row is None or row.tracker_import_id != tracker_import_id):
        raise LoadingTrackerError("The planning row could not be found.")

    store_name = _string_value(form_data.get("store_name"))
    if not store_name:
        raise LoadingTrackerError("Store name is required.")

    items = _parse_items_text(_string_value(form_data.get("items_text")))
    if not items:
        raise LoadingTrackerError("Add at least one SKU quantity in the planner.")

    selected_day_name = _string_value(form_data.get("target_day_name")) or PENDING_SENTINEL
    reason_code = _string_value(form_data.get("reason_code"))
    reason_note = _string_value(form_data.get("reason_note"))
    day = None
    row_state = PENDING_STATE
    if selected_day_name != PENDING_SENTINEL:
        day = get_loading_tracker_day(tracker_import_id, selected_day_name)
        if day is None:
            raise LoadingTrackerError("The selected planning day could not be found.")
        row_state = PLANNED_STATE
    elif not reason_code and not (row is not None and row.reason_text):
        raise LoadingTrackerError("Choose a pending reason before sending a line out of today's plan.")

    if row is None:
        row = LoadingTrackerRow(tracker_import_id=tracker_import_id, source_kind="manual")
        db.session.add(row)
        db.session.flush()

    row.day_id = day.id if day is not None else None
    row.row_state = row_state
    row.batch_name = _string_value(form_data.get("batch_name")) or "Unassigned"
    row.store_name = store_name
    row.contact = _string_value(form_data.get("contact")) or None
    row.lp = _string_value(form_data.get("lp")) or None
    row.tier = _string_value(form_data.get("tier")) or None
    row.region = _string_value(form_data.get("region")) or None
    row.delivery_date = _string_value(form_data.get("delivery_date")) or None
    if row_state == PENDING_STATE:
        row.reason_text = _compose_pending_reason(reason_code, reason_note) or row.reason_text
    else:
        row.reason_text = None
    row.total_weight = _decimal_value(form_data.get("total_weight"))
    row.total_value = _decimal_value(form_data.get("total_value"))
    row.sort_order = _next_sort_order(tracker_import, day, row_state, row.id)

    row.items.clear()
    for item in items:
        row.items.append(
            LoadingTrackerRowItem(
                sku_name=item["sku"],
                quantity=_decimal_or_none(item["quantity"]) or Decimal("0"),
                raw_reference_no=item.get("raw_reference_no") or None,
                invoice_category=item.get("invoice_category") or None,
                prefixed_reference_no=item.get("prefixed_reference_no") or None,
                classification_source="planner_text" if item.get("prefixed_reference_no") else None,
                bp_rule_reason=None,
            )
        )

    db.session.flush()
    _log_tracker_event(
        tracker_import=tracker_import,
        day=day,
        row=row,
        event_type="planner_row_saved",
        entity_name=row.store_name,
        reason_code=reason_code or None,
        reason_text=row.reason_text,
        details={
            "row_state": row.row_state,
            "batch_name": row.batch_name,
            "product_count": len(items),
            "target_day_name": day.day_name if day is not None else PENDING_SENTINEL,
        },
    )
    db.session.commit()
    return row


def move_loading_tracker_row(
    tracker_import_id: str,
    row_id: int,
    target_day_name: str,
    reason_code: str | None = None,
    reason_note: str | None = None,
) -> LoadingTrackerRow:
    row = get_loading_tracker_row(row_id)
    if row is None or row.tracker_import_id != tracker_import_id:
        raise LoadingTrackerError("The planning row could not be found.")
    _move_loading_tracker_row_record(row, tracker_import_id, target_day_name, reason_code, reason_note)

    db.session.commit()
    return row


def bulk_move_loading_tracker_rows(
    tracker_import_id: str,
    row_ids: list[int],
    target_day_name: str,
    reason_code: str | None = None,
    reason_note: str | None = None,
) -> list[LoadingTrackerRow]:
    unique_row_ids = list(dict.fromkeys(row_ids))
    if not unique_row_ids:
        raise LoadingTrackerError("Choose at least one planner row first.")

    rows = list(
        db.session.scalars(
            select(LoadingTrackerRow)
            .where(
                LoadingTrackerRow.tracker_import_id == tracker_import_id,
                LoadingTrackerRow.id.in_(unique_row_ids),
            )
            .order_by(LoadingTrackerRow.sort_order.asc(), LoadingTrackerRow.id.asc())
        )
    )
    if len(rows) != len(unique_row_ids):
        raise LoadingTrackerError("One or more selected planner rows could not be found.")

    for row in rows:
        _move_loading_tracker_row_record(row, tracker_import_id, target_day_name, reason_code, reason_note)

    db.session.commit()
    return rows


def _move_loading_tracker_row_record(
    row: LoadingTrackerRow,
    tracker_import_id: str,
    target_day_name: str,
    reason_code: str | None = None,
    reason_note: str | None = None,
) -> None:
    previous_day_name = row.day.day_name if row.day is not None else None

    if target_day_name == PENDING_SENTINEL:
        if not reason_code:
            raise LoadingTrackerError("Choose a pending reason before moving a line into Pending.")
        row.day_id = None
        row.row_state = PENDING_STATE
        row.reason_text = _compose_pending_reason(reason_code, reason_note)
        row.sort_order = _next_sort_order(row.tracker_import, None, PENDING_STATE, row.id)
        _log_tracker_event(
            tracker_import=row.tracker_import,
            day=None,
            row=row,
            event_type="moved_to_pending",
            entity_name=row.store_name,
            reason_code=reason_code,
            reason_text=row.reason_text,
            details={"from_day": previous_day_name},
        )
        return

    day = get_loading_tracker_day(tracker_import_id, target_day_name)
    if day is None:
        raise LoadingTrackerError("The selected planning day could not be found.")
    row.day_id = day.id
    row.row_state = PLANNED_STATE
    row.reason_text = None
    row.sort_order = _next_sort_order(row.tracker_import, day, PLANNED_STATE, row.id)
    _log_tracker_event(
        tracker_import=row.tracker_import,
        day=day,
        row=row,
        event_type="moved_into_day",
        entity_name=row.store_name,
        details={"from_day": previous_day_name, "to_day": day.day_name},
    )


def save_inventory_adjustment(tracker_import_id: str, form_data: dict[str, Any]) -> LoadingTrackerInventoryItem:
    tracker_import = get_loading_tracker_import(tracker_import_id)
    if tracker_import is None:
        raise LoadingTrackerError("The selected loading tracker import could not be found.")

    sku_name = _string_value(form_data.get("sku_name"))
    if not sku_name:
        raise LoadingTrackerError("SKU name is required for the inventory adjustment.")

    item = next((entry for entry in tracker_import.inventory_items if entry.sku_name == sku_name), None)
    if item is None:
        item = LoadingTrackerInventoryItem(
            tracker_import_id=tracker_import_id,
            sku_name=sku_name,
            sort_order=(max((entry.sort_order for entry in tracker_import.inventory_items), default=0) + 1),
        )
        db.session.add(item)

    opening_g2g = _decimal_value(form_data.get("opening_g2g_qty"))
    opening_remaining = _decimal_value(form_data.get("opening_remaining_qty"))
    added_qty = _decimal_value(form_data.get("added_qty"))
    item.opening_g2g_qty = opening_g2g or Decimal("0")
    item.opening_remaining_qty = opening_remaining or Decimal("0")
    item.added_qty = added_qty or Decimal("0")

    _log_tracker_event(
        tracker_import=tracker_import,
        day=None,
        row=None,
        event_type="inventory_adjusted",
        entity_name=item.sku_name,
        details={
            "opening_g2g_qty": _float_value(item.opening_g2g_qty),
            "opening_remaining_qty": _float_value(item.opening_remaining_qty),
            "added_qty": _float_value(item.added_qty),
        },
    )
    db.session.commit()
    return item


def build_loading_tracker_count_context(day: LoadingTrackerDay) -> dict[str, Any]:
    expected_start = _expected_start_by_day(day.tracker_import, day)
    for sku_name in _aggregate_row_item_totals([row for row in day.planning_rows if row.row_state == PLANNED_STATE]):
        expected_start.setdefault(sku_name, 0.0)
    count_lookup, discrepancy_rows, count_ready = _day_count_snapshot(day, expected_start)
    rows = []
    for row in discrepancy_rows:
        rows.append(
            {
                "sku": row["sku"],
                "expected_qty": row["expected_qty"],
                "physical_qty": row["physical_qty"],
                "discrepancy_qty": row["discrepancy_qty"],
                "status": row["status"],
            }
        )

    return {
        "rows": rows,
        "count_ready": count_ready,
        "expected_total": round(sum(item["expected_qty"] for item in rows), 2),
        "physical_total": round(sum(item["physical_qty"] for item in rows), 2),
        "discrepancy_total": round(sum(item["discrepancy_qty"] for item in rows), 2),
        "variance_rows": len([item for item in rows if abs(item["discrepancy_qty"]) > 0.0001]),
    }


def save_loading_tracker_day_counts(import_id: str, day_name: str, form_data: dict[str, Any]) -> LoadingTrackerDay:
    day = get_loading_tracker_day(import_id, day_name)
    if day is None:
        raise LoadingTrackerError("The selected planning day could not be found.")

    expected_start = _expected_start_by_day(day.tracker_import, day)
    incoming: dict[str, float] = {}
    for key, value in form_data.items():
        if not key.startswith("count::"):
            continue
        sku_name = key.split("count::", 1)[1]
        incoming[sku_name] = _float_value(value)

    if not incoming:
        raise LoadingTrackerError("Enter the physical inventory count for the day before saving.")

    existing = {item.sku_name: item for item in day.inventory_counts}
    all_skus = sorted(set(expected_start) | set(incoming) | set(existing))
    for sku_name in all_skus:
        expected_qty = round(expected_start.get(sku_name, 0.0), 4)
        physical_qty = round(incoming.get(sku_name, existing.get(sku_name).physical_qty if sku_name in existing else expected_qty), 4)
        discrepancy_qty = round(physical_qty - expected_qty, 4)
        record = existing.get(sku_name)
        if record is None:
            record = LoadingTrackerDailyCount(day_id=day.id, sku_name=sku_name)
            db.session.add(record)
        record.expected_qty = _decimal_or_none(expected_qty) or Decimal("0")
        record.physical_qty = _decimal_or_none(physical_qty) or Decimal("0")
        record.discrepancy_qty = _decimal_or_none(discrepancy_qty) or Decimal("0")

    _log_tracker_event(
        tracker_import=day.tracker_import,
        day=day,
        row=None,
        event_type="captured_day_count",
        entity_name=day.day_name,
        details={
            "sku_count": len(all_skus),
            "variance_rows": len([sku for sku in all_skus if abs(incoming.get(sku, expected_start.get(sku, 0.0)) - expected_start.get(sku, 0.0)) > 0.0001]),
        },
    )
    discrepancy_rows = [
        {
            "sku": sku_name,
            "expected_qty": round(expected_start.get(sku_name, 0.0), 4),
            "physical_qty": round(incoming.get(sku_name, expected_start.get(sku_name, 0.0)), 4),
            "discrepancy_qty": round(incoming.get(sku_name, expected_start.get(sku_name, 0.0)) - expected_start.get(sku_name, 0.0), 4),
        }
        for sku_name in all_skus
        if incoming.get(sku_name, expected_start.get(sku_name, 0.0)) < expected_start.get(sku_name, 0.0)
    ]
    if discrepancy_rows:
        _log_inventory_discrepancy_alert(day, discrepancy_rows)
    db.session.commit()
    return day


def build_loading_tracker_history_context(tracker_import: LoadingTrackerImport) -> dict[str, Any]:
    events = []
    reason_counts: dict[str, int] = {}
    for event in tracker_import.events:
        reason_text = event.reason_text or ""
        if event.reason_code:
            reason_counts[PENDING_REASON_LABELS.get(event.reason_code, event.reason_code)] = (
                reason_counts.get(PENDING_REASON_LABELS.get(event.reason_code, event.reason_code), 0) + 1
            )
        events.append(
            {
                "created_at": event.created_at,
                "event_type": event.event_type.replace("_", " "),
                "entity_name": event.entity_name,
                "day_name": event.day.day_name if event.day is not None else "",
                "reason_text": reason_text,
                "details": event.details_json or {},
            }
        )

    return {
        "events": events,
        "event_count": len(events),
        "reason_counts": [{"reason": reason, "count": count} for reason, count in sorted(reason_counts.items(), key=lambda item: (-item[1], item[0]))],
    }


def export_loading_tracker_history_csv(import_id: str) -> tuple[str, bytes]:
    tracker_import = get_loading_tracker_import(import_id)
    if tracker_import is None:
        raise LoadingTrackerError("The selected loading tracker import could not be found.")

    import io

    buffer = BytesIO()
    string_io = io.StringIO()
    writer = csv.writer(string_io)
    writer.writerow(["Created At", "Event Type", "Week", "Day", "Entity", "Reason", "Details"])
    for event in tracker_import.events:
        writer.writerow(
            [
                event.created_at.isoformat() if event.created_at else "",
                event.event_type,
                tracker_import.week_label,
                event.day.day_name if event.day is not None else "",
                event.entity_name,
                event.reason_text or "",
                event.details_json or {},
            ]
        )
    buffer.write(string_io.getvalue().encode("utf-8"))
    return f"{_clean_filename_part(tracker_import.week_label, 'loading-tracker-history')} - history.csv", buffer.getvalue()


def carry_forward_loading_tracker_week(source_import_id: str | None = None) -> LoadingTrackerImport:
    source_import = get_loading_tracker_import(source_import_id)
    if source_import is None:
        raise LoadingTrackerError("There is no loading tracker week to carry forward yet.")

    remaining_map = _closing_inventory_by_import(source_import)
    pending_rows = [row for row in source_import.planning_rows if row.row_state == PENDING_STATE]
    carried_import = LoadingTrackerImport(
        id=uuid4().hex,
        filename=f"Carry Forward - {source_import.filename}",
        week_label=f"Week of {datetime.now().date()}",
        assumptions_sku_count=source_import.assumptions_sku_count,
        assumptions_store_count=source_import.assumptions_store_count,
        opening_g2g_total=_decimal_or_none(sum(remaining_map.values())),
        opening_remaining_total=_decimal_or_none(sum(remaining_map.values())),
        fees_row_count=source_import.fees_row_count,
        fees_total_delivery_value=source_import.fees_total_delivery_value,
        fees_total_payment_value=source_import.fees_total_payment_value,
        notes_count=source_import.notes_count,
    )
    db.session.add(carried_import)
    db.session.flush()

    for order, day_name in enumerate(DAY_SHEET_NAMES, start=1):
        db.session.add(
            LoadingTrackerDay(
                tracker_import_id=carried_import.id,
                day_name=day_name,
                day_order=order,
            )
        )

    for sort_order, sku_name in enumerate(sorted(remaining_map), start=1):
        qty = round(remaining_map.get(sku_name, 0.0), 4)
        db.session.add(
            LoadingTrackerInventoryItem(
                tracker_import_id=carried_import.id,
                sku_name=sku_name,
                opening_g2g_qty=_decimal_or_none(qty) or Decimal("0"),
                opening_remaining_qty=_decimal_or_none(qty) or Decimal("0"),
                added_qty=Decimal("0"),
                sort_order=sort_order,
            )
        )

    for fee_item in source_import.fee_items:
        db.session.add(
            LoadingTrackerFeeItem(
                tracker_import_id=carried_import.id,
                brand_partner=fee_item.brand_partner,
                sku_name=fee_item.sku_name,
                vatable_text=fee_item.vatable_text,
                retail_delivery_value=fee_item.retail_delivery_value,
                payment_collection_value=fee_item.payment_collection_value,
            )
        )

    for sort_order, row in enumerate(pending_rows, start=1):
        row_data = _serialize_row(row)
        _save_row_record(
            tracker_import=carried_import,
            row_data=row_data,
            row_state=PENDING_STATE,
            day=None,
            sort_order=sort_order,
            source_kind="carry-forward",
            reason_text=row.reason_text or "Carried into the new week",
        )

    _log_tracker_event(
        tracker_import=source_import,
        day=None,
        row=None,
        event_type="carried_forward_week",
        entity_name=source_import.week_label,
        details={"next_week_id": carried_import.id},
    )
    _log_tracker_event(
        tracker_import=carried_import,
        day=None,
        row=None,
        event_type="created_from_carry_forward",
        entity_name=carried_import.week_label,
        details={"source_week_id": source_import.id, "pending_rows": len(pending_rows)},
    )
    db.session.commit()
    return carried_import


def create_delivery_note_run_from_loading_day(import_id: str, day_name: str, timezone_name: str) -> UploadRun:
    day = get_loading_tracker_day(import_id, day_name)
    if day is None:
        raise LoadingTrackerError("The selected planning day could not be found.")

    planned_rows = [row for row in day.planning_rows if row.row_state == PLANNED_STATE]
    if not planned_rows:
        raise LoadingTrackerError("This day has no planned rows to send into Delivery Note.")

    products = list(db.session.scalars(select(Product)))
    aliases = list(db.session.scalars(select(ProductAlias)))
    bp_rules = load_brand_partner_rules()

    run = UploadRun(
        id=uuid4().hex,
        original_filename=f"{day.tracker_import.week_label} - {day.day_name} adjusted plan",
        invoice_date=tomorrow_in_timezone(timezone_name).strftime(DATE_FORMAT),
        status="needs_review",
        rows_detected=0,
        rows_ready=0,
        rows_needing_review=0,
    )
    db.session.add(run)
    db.session.flush()

    for row in planned_rows:
        planning_ref = f"LT-{day.day_name.upper()}-{row.id:04d}"
        for item in row.items:
            quantity = item.quantity or Decimal("0")
            if quantity <= 0:
                continue
            run.rows_detected += 1
            item_reference = item.raw_reference_no or planning_ref
            item_category, item_owner, item_tax_bucket = invoice_category_parts(item.invoice_category)
            prefixed_reference = item.prefixed_reference_no or build_prefixed_reference(
                item_category or item.invoice_category,
                item_reference,
            )
            line = UploadLine(
                run_id=run.id,
                order_number=prefixed_reference or item_reference,
                supermarket_name=row.store_name,
                source_sku=item.sku_name,
                normalized_source_sku=_normalize_text(item.sku_name).upper(),
                quantity=quantity,
                raw_reference_no=item_reference,
                invoice_owner=item.invoice_owner or item_owner,
                tax_bucket=item.tax_bucket or item_tax_bucket,
                invoice_category=item_category or item.invoice_category or None,
                prefixed_reference_no=prefixed_reference or None,
                classification_source=item.classification_source or None,
                bp_rule_reason=item.bp_rule_reason or None,
            )
            match = resolve_product_match(item.sku_name, products, aliases)
            if match is None:
                line.status = "needs_review"
                run.rows_needing_review += 1
            elif not match.product.is_active:
                line.status = "ignored"
                line.matched_by = "inactive"
                line.product_id = match.product.id
                line.resolved_sku_name = match.product.sku_name
            elif match.product.price is None:
                line.status = "needs_review"
                run.rows_needing_review += 1
            else:
                apply_product_to_line(line, match.product, match.match_method, bp_rules=bp_rules)
                run.rows_ready += 1
            db.session.add(line)

    run.status = "ready" if run.rows_needing_review == 0 else "needs_review"
    _log_tracker_event(
        tracker_import=day.tracker_import,
        day=day,
        row=None,
        event_type="sent_to_delivery_note",
        entity_name=day.day_name,
        details={"run_id": run.id, "rows_detected": run.rows_detected},
    )
    db.session.commit()
    return run


def group_store_rows_by_batch(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows or []:
        grouped.setdefault(row.get("batch_name", "Unassigned"), []).append(row)

    ordered = []
    for batch_name in sorted(grouped, key=_batch_sort_key):
        batch_rows = grouped[batch_name]
        ordered.append(
            {
                "batch_name": batch_name,
                "store_count": len(batch_rows),
                "total_quantity": round(sum(row.get("total_quantity", 0) for row in batch_rows), 2),
                "total_weight": round(sum(row.get("weight", 0) for row in batch_rows), 2),
                "total_value": round(sum(row.get("value", 0) for row in batch_rows), 2),
                "rows": batch_rows,
            }
        )
    return ordered


def _save_row_record(
    tracker_import: LoadingTrackerImport,
    row_data: dict[str, Any],
    row_state: str,
    day: LoadingTrackerDay | None,
    sort_order: int,
    source_kind: str,
    reason_text: str | None = None,
) -> None:
    record = LoadingTrackerRow(
        tracker_import=tracker_import,
        day=day,
        row_state=row_state,
        source_kind=source_kind,
        batch_name=row_data.get("batch_name") or "Unassigned",
        store_name=row_data.get("store_name") or "Unnamed store",
        contact=row_data.get("contact") or None,
        lp=row_data.get("lp") or None,
        tier=row_data.get("tier") or None,
        region=row_data.get("region") or None,
        delivery_date=row_data.get("delivery_date") or None,
        reason_text=reason_text,
        total_weight=_decimal_or_none(row_data.get("weight", 0.0)),
        total_value=_decimal_or_none(row_data.get("value", 0.0)),
        sort_order=sort_order,
    )
    db.session.add(record)

    for item in row_data.get("items", []):
        item_category, item_owner, item_tax_bucket = invoice_category_parts(item.get("invoice_category"))
        record.items.append(
            LoadingTrackerRowItem(
                sku_name=item["sku"],
                quantity=_decimal_or_none(item["quantity"]) or Decimal("0"),
                raw_reference_no=item.get("raw_reference_no") or None,
                invoice_owner=item.get("invoice_owner") or item_owner,
                tax_bucket=item.get("tax_bucket") or item_tax_bucket,
                invoice_category=item_category or item.get("invoice_category") or None,
                prefixed_reference_no=item.get("prefixed_reference_no") or None,
                classification_source=item.get("classification_source") or None,
                bp_rule_reason=item.get("bp_rule_reason") or None,
            )
        )


def _rows_for_days(days: list[LoadingTrackerDay]) -> list[LoadingTrackerRow]:
    rows: list[LoadingTrackerRow] = []
    for day in days:
        rows.extend([row for row in day.planning_rows if row.row_state == PLANNED_STATE])
    return rows


def _aggregate_row_item_totals(rows: list[LoadingTrackerRow]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for row in rows:
        for item in row.items:
            quantity = _float_value(item.quantity)
            if quantity <= 0:
                continue
            totals[item.sku_name] = totals.get(item.sku_name, 0.0) + quantity
    return totals


def _inventory_totals_by_sku(tracker_import: LoadingTrackerImport) -> dict[str, float]:
    totals: dict[str, float] = {}
    for item in tracker_import.inventory_items:
        totals[item.sku_name] = round(_float_value(item.opening_g2g_qty) + _float_value(item.added_qty), 4)
    return totals


def _subtract_maps(left: dict[str, float], right: dict[str, float]) -> dict[str, float]:
    keys = set(left) | set(right)
    return {key: round(left.get(key, 0.0) - right.get(key, 0.0), 4) for key in keys}


def _build_day_suggestions(
    planned_rows: list[LoadingTrackerRow],
    remaining_after: dict[str, float],
) -> list[dict[str, Any]]:
    suggestions: list[dict[str, Any]] = []
    for sku_name, remaining_qty in sorted(remaining_after.items(), key=lambda item: item[1]):
        if remaining_qty >= 0:
            continue
        shortfall = round(abs(remaining_qty), 2)
        candidate_rows = [
            row for row in planned_rows if _row_quantity_for_sku(row, sku_name) > 0
        ]
        candidate_rows.sort(key=lambda row: (-_row_quantity_for_sku(row, sku_name), row.sort_order, row.id))
        selected_rows: list[dict[str, Any]] = []
        covered_qty = 0.0
        selected_ids: list[int] = []
        for row in candidate_rows:
            sku_qty = round(_row_quantity_for_sku(row, sku_name), 2)
            if sku_qty <= 0:
                continue
            selected_ids.append(row.id)
            selected_rows.append({"id": row.id, "store_name": row.store_name, "quantity": sku_qty})
            covered_qty += sku_qty
            if covered_qty >= shortfall:
                break
        if not selected_ids:
            continue
        suggestions.append(
            {
                "sku": sku_name,
                "shortfall": shortfall,
                "selected_count": len(selected_ids),
                "row_ids": selected_ids,
                "rows": selected_rows,
                "reason_code": "stock_shortage",
                "reason_label": PENDING_REASON_LABELS["stock_shortage"],
                "reason_note": f"{sku_name} short by {shortfall:.2f}",
            }
        )
    return suggestions


def _row_quantity_for_sku(row: LoadingTrackerRow, sku_name: str) -> float:
    total = 0.0
    for item in row.items:
        if item.sku_name == sku_name:
            total += _float_value(item.quantity)
    return round(total, 4)


def _serialize_row(row: LoadingTrackerRow) -> dict[str, Any]:
    items = [
        {
            "sku": item.sku_name,
            "quantity": round(_float_value(item.quantity), 2),
            "invoice_owner": item.invoice_owner or "",
            "tax_bucket": item.tax_bucket or "",
            "invoice_category": item.invoice_category or "",
            "raw_reference_no": item.raw_reference_no or "",
            "prefixed_reference_no": build_prefixed_reference(item.invoice_category, item.raw_reference_no)
            or item.prefixed_reference_no
            or "",
            "classification_source": item.classification_source or "",
            "bp_rule_reason": item.bp_rule_reason or "",
        }
        for item in row.items
        if _float_value(item.quantity) > 0
    ]
    items.sort(
        key=lambda item: (
            -item["quantity"],
            item["sku"],
            item["prefixed_reference_no"] or item["raw_reference_no"],
            item["invoice_category"],
        )
    )
    return {
        "id": row.id,
        "row_state": row.row_state,
        "batch_name": row.batch_name or "Unassigned",
        "store_name": row.store_name,
        "contact": row.contact or "",
        "lp": row.lp or "",
        "tier": row.tier or "",
        "region": row.region or "",
        "delivery_date": row.delivery_date or "",
        "reason_text": row.reason_text or "",
        "weight": round(_float_value(row.total_weight), 2),
        "value": round(_float_value(row.total_value), 2),
        "total_quantity": round(sum(item["quantity"] for item in items), 2),
        "product_count": len(items),
        "items": items,
        "top_items": items[:5],
    }


def _serialize_template_row(row: LoadingTrackerTemplateRow) -> dict[str, Any]:
    items = [
        {"sku": item.sku_name, "quantity": round(_float_value(item.quantity), 2)}
        for item in row.items
        if _float_value(item.quantity) > 0
    ]
    items.sort(key=lambda item: (-item["quantity"], item["sku"]))
    return {
        "batch_name": row.batch_name or "Unassigned",
        "store_name": row.store_name,
        "contact": row.contact or "",
        "lp": row.lp or "",
        "tier": row.tier or "",
        "region": row.region or "",
        "delivery_date": row.delivery_date or "",
        "reason_text": row.reason_text or "",
        "weight": round(_float_value(row.total_weight), 2),
        "value": round(_float_value(row.total_value), 2),
        "total_quantity": round(sum(item["quantity"] for item in items), 2),
        "product_count": len(items),
        "items": items,
        "top_items": items[:5],
    }


def _group_sku_automator_store_rows(lines: list[SkuAutomatorLine]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for line in lines:
        store_key = _normalize_text(line.store_name).upper() or line.store_name.upper()
        entry = grouped.setdefault(
            store_key,
            {
                "batch_name": "Unassigned",
                "store_name": line.store_name,
                "contact": "",
                "lp": "",
                "tier": "",
                "region": "",
                "delivery_date": line.order_date or "",
                "weight": 0.0,
                "value": 0.0,
                "items": {},
                "order_references": set(),
            },
        )
        if line.order_date and (not entry["delivery_date"] or str(line.order_date) < str(entry["delivery_date"])):
            entry["delivery_date"] = line.order_date
        if line.order_reference_no:
            entry["order_references"].add(
                build_prefixed_reference(line.invoice_category, line.raw_reference_no)
                or line.prefixed_reference_no
                or line.order_reference_no
            )
        sku_name = line.resolved_sku_name or line.source_sku
        quantity = _float_value(line.resolved_quantity)
        if quantity > 0:
            normalized_prefixed_reference = (
                build_prefixed_reference(line.invoice_category, line.raw_reference_no)
                or line.prefixed_reference_no
                or ""
            )
            item_key = (
                sku_name,
                line.invoice_category or "",
                normalized_prefixed_reference,
                line.raw_reference_no or "",
            )
            item_entry = entry["items"].setdefault(
                item_key,
                {
                    "sku": sku_name,
                    "quantity": 0.0,
                    "invoice_owner": line.invoice_owner or "",
                    "tax_bucket": line.tax_bucket or "",
                    "invoice_category": line.invoice_category or "",
                    "prefixed_reference_no": normalized_prefixed_reference,
                    "raw_reference_no": line.raw_reference_no or "",
                    "classification_source": line.classification_source or "",
                    "bp_rule_reason": line.bp_rule_reason or "",
                },
            )
            item_entry["quantity"] = round(item_entry["quantity"] + quantity, 4)
        entry["value"] = round(entry["value"] + _float_value(line.source_value), 2)

    rows: list[dict[str, Any]] = []
    for entry in sorted(grouped.values(), key=lambda item: str(item["store_name"]).upper()):
        items = [
            {
                "sku": item["sku"],
                "quantity": round(item["quantity"], 2),
                "invoice_owner": item["invoice_owner"],
                "tax_bucket": item["tax_bucket"],
                "invoice_category": item["invoice_category"],
                "prefixed_reference_no": item["prefixed_reference_no"],
                "raw_reference_no": item["raw_reference_no"],
                "classification_source": item["classification_source"],
                "bp_rule_reason": item["bp_rule_reason"],
            }
            for _, item in sorted(
                entry["items"].items(),
                key=lambda pair: (
                    -pair[1]["quantity"],
                    pair[1]["sku"],
                    pair[1]["prefixed_reference_no"] or pair[1]["raw_reference_no"],
                ),
            )
            if item["quantity"] > 0
        ]
        if not items:
            continue
        rows.append(
            {
                "batch_name": entry["batch_name"],
                "store_name": entry["store_name"],
                "contact": "",
                "lp": "",
                "tier": "",
                "region": "",
                "delivery_date": entry["delivery_date"],
                "weight": 0.0,
                "value": round(entry["value"], 2),
                "items": items,
                "order_references": sorted(entry["order_references"]),
            }
        )
    return rows


def _build_ll_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, float]]:
    sku_totals: dict[str, dict[str, float]] = {}
    totals = {"load_1": 0.0, "load_2": 0.0, "load_3": 0.0, "load_4": 0.0, "total": 0.0}
    for row in rows:
        batch_key = _batch_to_key(row["batch_name"])
        for item in row["items"]:
            sku_totals.setdefault(
                item["sku"],
                {"load_1": 0.0, "load_2": 0.0, "load_3": 0.0, "load_4": 0.0, "total": 0.0},
            )
            sku_totals[item["sku"]][batch_key] += item["quantity"]
            sku_totals[item["sku"]]["total"] += item["quantity"]
            totals[batch_key] += item["quantity"]
            totals["total"] += item["quantity"]

    ll_rows = []
    for sku_name, item_totals in sorted(sku_totals.items(), key=lambda entry: (-entry[1]["total"], entry[0])):
        ll_rows.append(
            {
                "sku": sku_name,
                "load_1": round(item_totals["load_1"], 2),
                "load_2": round(item_totals["load_2"], 2),
                "load_3": round(item_totals["load_3"], 2),
                "load_4": round(item_totals["load_4"], 2),
                "total": round(item_totals["total"], 2),
            }
        )

    return ll_rows, totals


def _build_pack_breaker_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    breaker_map: dict[tuple[str, str], dict[str, Any]] = {}
    batch_summary: dict[str, dict[str, Any]] = {}
    for row in rows:
        batch_name = row["batch_name"]
        batch_summary.setdefault(
            batch_name,
            {"batch_name": batch_name, "half_packs": 0, "quarter_packs": 0, "rows": 0},
        )
        for item in row["items"]:
            fraction = round(item["quantity"] - int(item["quantity"]), 2)
            if fraction <= 0:
                continue
            half_count, quarter_count = _fraction_breakdown(fraction)
            if half_count == 0 and quarter_count == 0:
                continue
            key = (batch_name, item["sku"])
            entry = breaker_map.setdefault(
                key,
                {
                    "batch_name": batch_name,
                    "sku": item["sku"],
                    "half_packs": 0,
                    "quarter_packs": 0,
                    "fractional_total": 0.0,
                },
            )
            entry["half_packs"] += half_count
            entry["quarter_packs"] += quarter_count
            entry["fractional_total"] = round(entry["fractional_total"] + fraction, 2)
            batch_summary[batch_name]["half_packs"] += half_count
            batch_summary[batch_name]["quarter_packs"] += quarter_count
            batch_summary[batch_name]["rows"] += 1

    rows_list = sorted(
        breaker_map.values(),
        key=lambda item: (_batch_sort_key(item["batch_name"]), item["sku"]),
    )
    batch_list = sorted(batch_summary.values(), key=lambda item: _batch_sort_key(item["batch_name"]))
    return rows_list, batch_list


def _fraction_breakdown(fraction: float) -> tuple[int, int]:
    quarter_units = int(round(fraction * 4))
    half_count = 1 if quarter_units in (2, 3) else 0
    quarter_count = 1 if quarter_units in (1, 3) else 0
    return half_count, quarter_count


def _batch_to_key(batch_name: str) -> str:
    lowered = (batch_name or "").lower()
    if "2" in lowered:
        return "load_2"
    if "3" in lowered:
        return "load_3"
    if "4" in lowered:
        return "load_4"
    return "load_1"


def _row_items_text(row: LoadingTrackerRow | None) -> str:
    if row is None:
        return ""
    items = []
    for item in row.items:
        quantity = round(_float_value(item.quantity), 2)
        if quantity <= 0:
            continue
        line = f"{item.sku_name} = {quantity:g}"
        item_category, _, _ = invoice_category_parts(item.invoice_category)
        reference = build_prefixed_reference(
            item_category or item.invoice_category,
            item.raw_reference_no,
        ) or item.prefixed_reference_no
        if reference:
            line += f" | {reference}"
        items.append(line)
    return "\n".join(items)


def _parse_items_text(text: str) -> list[dict[str, Any]]:
    parsed: list[dict[str, Any]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        metadata_parts = [part.strip() for part in line.split("|")]
        line = metadata_parts[0]
        separator = None
        for candidate in ("=", "\t", "|", ":"):
            if candidate in line:
                separator = candidate
                break
        if separator is not None:
            sku_name, quantity_text = line.rsplit(separator, 1)
        else:
            parts = line.rsplit(" ", 1)
            if len(parts) != 2:
                raise LoadingTrackerError(f"Could not understand planner line '{line}'. Use 'SKU = qty'.")
            sku_name, quantity_text = parts
        sku_name = sku_name.strip().strip("-")
        quantity = _float_value(quantity_text)
        if not sku_name or quantity <= 0:
            continue
        invoice_category = ""
        raw_reference_no = ""
        prefixed_reference_no = ""
        if len(metadata_parts) >= 2:
            category, stripped_reference = split_prefixed_reference(metadata_parts[1])
            if category:
                invoice_category = category
                raw_reference_no = stripped_reference
                prefixed_reference_no = metadata_parts[1]
            elif metadata_parts[1].upper() in {"BP", "VT", "NV", "BPVT", "BPNV"}:
                invoice_category = metadata_parts[1].upper()
                if len(metadata_parts) >= 3:
                    raw_reference_no = metadata_parts[2]
                    prefixed_reference_no = build_prefixed_reference(invoice_category, raw_reference_no) or ""
            else:
                raw_reference_no = metadata_parts[1]
        parsed.append(
            {
                "sku": sku_name,
                "quantity": round(quantity, 4),
                "invoice_category": invoice_category,
                "raw_reference_no": raw_reference_no,
                "prefixed_reference_no": prefixed_reference_no,
            }
        )
    return parsed


def _compose_pending_reason(reason_code: str | None, reason_note: str | None) -> str | None:
    if not reason_code and not reason_note:
        return None
    label = PENDING_REASON_LABELS.get(reason_code or "", "Manual operations hold")
    note = _string_value(reason_note)
    if note:
        return f"{label}: {note}"
    return label


def _expected_start_by_day(tracker_import: LoadingTrackerImport, target_day: LoadingTrackerDay) -> dict[str, float]:
    inventory_map = _inventory_totals_by_sku(tracker_import)
    day_sequence = [day for day in tracker_import.days if day.day_order < target_day.day_order]
    running = dict(inventory_map)
    for day in day_sequence:
        counted_map = _day_count_map(day)
        available = counted_map if counted_map else running
        consumed = _aggregate_row_item_totals([row for row in day.planning_rows if row.row_state == PLANNED_STATE])
        running = _subtract_maps(available, consumed)
    return running


def _closing_inventory_by_import(tracker_import: LoadingTrackerImport) -> dict[str, float]:
    running = dict(_inventory_totals_by_sku(tracker_import))
    for day in tracker_import.days:
        counted_map = _day_count_map(day)
        available = counted_map if counted_map else running
        consumed = _aggregate_row_item_totals([row for row in day.planning_rows if row.row_state == PLANNED_STATE])
        running = _subtract_maps(available, consumed)
    return running


def _day_count_map(day: LoadingTrackerDay) -> dict[str, float]:
    return {item.sku_name: _float_value(item.physical_qty) for item in day.inventory_counts}


def _day_count_snapshot(day: LoadingTrackerDay, expected_start: dict[str, float]) -> tuple[dict[str, float], list[dict[str, Any]], bool]:
    count_lookup = _day_count_map(day)
    all_skus = sorted(set(expected_start) | set(count_lookup))
    rows = []
    for sku_name in all_skus:
        expected_qty = round(expected_start.get(sku_name, 0.0), 2)
        physical_qty = round(count_lookup.get(sku_name, expected_qty), 2)
        discrepancy_qty = round(physical_qty - expected_qty, 2)
        if discrepancy_qty > 0:
            status = "higher"
        elif discrepancy_qty < 0:
            status = "lower"
        else:
            status = "matched"
        rows.append(
            {
                "sku": sku_name,
                "expected_qty": expected_qty,
                "physical_qty": physical_qty,
                "discrepancy_qty": discrepancy_qty,
                "status": status,
            }
        )
    return count_lookup, rows, bool(day.inventory_counts)


def _log_tracker_event(
    tracker_import: LoadingTrackerImport,
    day: LoadingTrackerDay | None,
    row: LoadingTrackerRow | None,
    event_type: str,
    entity_name: str,
    reason_code: str | None = None,
    reason_text: str | None = None,
    details: dict[str, Any] | None = None,
) -> None:
    db.session.add(
        LoadingTrackerEvent(
            tracker_import_id=tracker_import.id,
            day_id=day.id if day is not None else None,
            row_id=row.id if row is not None else None,
            event_type=event_type,
            entity_name=entity_name,
            reason_code=reason_code,
            reason_text=reason_text,
            details_json=details or {},
        )
    )


def _log_inventory_discrepancy_alert(day: LoadingTrackerDay, discrepancy_rows: list[dict[str, float]]) -> None:
    recipients = [
        email.strip()
        for email in str(current_app.config.get("ALERT_EMAILS", "")).split(",")
        if email.strip()
    ]
    details = {
        "recipient_count": len(recipients),
        "recipients": recipients,
        "variance_count": len(discrepancy_rows),
        "items": discrepancy_rows[:20],
    }
    alert_status = "not_configured"
    if recipients:
        alert_status = _send_inventory_discrepancy_email(day, recipients, discrepancy_rows)
    details["delivery_status"] = alert_status
    _log_tracker_event(
        tracker_import=day.tracker_import,
        day=day,
        row=None,
        event_type="inventory_discrepancy_alert",
        entity_name=day.day_name,
        details=details,
    )


def _send_inventory_discrepancy_email(
    day: LoadingTrackerDay,
    recipients: list[str],
    discrepancy_rows: list[dict[str, float]],
) -> str:
    host = str(current_app.config.get("MAIL_HOST", "")).strip()
    port = int(current_app.config.get("MAIL_PORT", 587) or 587)
    username = str(current_app.config.get("MAIL_USERNAME", "")).strip()
    password = str(current_app.config.get("MAIL_PASSWORD", "")).strip()
    sender = str(current_app.config.get("MAIL_FROM", username or "noreply@dala-operations.local")).strip()
    use_tls = bool(current_app.config.get("MAIL_USE_TLS", True))
    if not host:
        return "not_configured"

    body_lines = [
        f"Inventory discrepancies were recorded for {day.day_name}.",
        "",
        "Items below expected stock:",
    ]
    for row in discrepancy_rows:
        body_lines.append(
            f"- {row['sku']}: expected {row['expected_qty']}, counted {row['physical_qty']}, delta {row['discrepancy_qty']}"
        )
    message = EmailMessage()
    message["Subject"] = f"DALA OPERATIONS inventory discrepancy alert - {day.day_name}"
    message["From"] = sender
    message["To"] = ", ".join(recipients)
    message.set_content("\n".join(body_lines))

    try:
        with smtplib.SMTP(host, port, timeout=15) as smtp:
            if use_tls:
                smtp.starttls()
            if username:
                smtp.login(username, password)
            smtp.send_message(message)
    except Exception:
        return "failed"
    return "sent"


def _next_sort_order(
    tracker_import: LoadingTrackerImport,
    day: LoadingTrackerDay | None,
    row_state: str,
    current_row_id: int | None = None,
) -> int:
    rows = [
        row
        for row in tracker_import.planning_rows
        if row.id != current_row_id
        and row.row_state == row_state
        and ((row.day_id == day.id) if day is not None else row.day_id is None)
    ]
    return max((row.sort_order for row in rows), default=0) + 1


def _parse_day_sheet(sheet: Any, day_name: str, load_sheet: Any | None) -> dict[str, Any]:
    grid = _locate_day_grid(sheet)
    max_row, _ = _sheet_bounds(sheet)
    cell = sheet.cell
    product_headers = grid["product_headers"]
    store_col = grid["store_col"]
    summary_rows = {
        key: _parse_metric_row(sheet, product_headers, label)
        for key, label in PLANNING_SUMMARY_LABELS.items()
    }

    batch_rows = []
    active_product_totals: dict[str, float] = {}
    current_batch = ""
    row_index = grid["header_row_index"] + 1

    while row_index <= max_row:
        batch_value = _string_value(cell(row_index, 1).value)
        if batch_value.lower().startswith("load "):
            current_batch = batch_value
            row_index += 1
            continue

        if _normalize_text(cell(row_index, store_col).value) == "load for external delivery":
            row_index += 1
            continue

        store_name = _string_value(cell(row_index, store_col).value)
        quantities: list[dict[str, Any]] = []
        total_quantity = 0.0
        for column_index, sku_name in product_headers:
            quantity = _float_value(cell(row_index, column_index).value)
            if quantity <= 0:
                continue
            total_quantity += quantity
            quantities.append({"sku": sku_name, "quantity": round(quantity, 2)})
            active_product_totals[sku_name] = active_product_totals.get(sku_name, 0.0) + quantity

        weight = _float_value(cell(row_index, grid["weight_col"]).value) if grid["weight_col"] else 0.0
        value = _float_value(cell(row_index, grid["value_col"]).value) if grid["value_col"] else 0.0

        if not store_name and total_quantity == 0 and weight == 0 and value == 0:
            row_index += 1
            continue

        if not store_name or store_name.lower() in {"item", "load for external delivery"}:
            row_index += 1
            continue

        if grid["value_col"] and value <= 0:
            row_index += 1
            continue

        if total_quantity <= 0 and value <= 0 and weight <= 0:
            row_index += 1
            continue

        quantities.sort(key=lambda item: (-item["quantity"], item["sku"]))
        batch_rows.append(
            {
                "batch_name": current_batch or "Unassigned",
                "store_name": store_name,
                "contact": _string_value(cell(row_index, grid["contact_col"]).value) if grid["contact_col"] else "",
                "lp": _string_value(cell(row_index, grid["lp_col"]).value) if grid["lp_col"] else "",
                "tier": _string_value(cell(row_index, grid["tier_col"]).value) if grid["tier_col"] else "",
                "region": _string_value(cell(row_index, grid["region_col"]).value) if grid["region_col"] else "",
                "delivery_date": _string_value(cell(row_index, grid["date_col"]).value) if grid["date_col"] else "",
                "weight": round(weight, 2),
                "value": round(value, 2),
                "total_quantity": round(total_quantity, 2),
                "product_count": len(quantities),
                "items": quantities,
                "top_items": quantities[:5],
            }
        )
        row_index += 1

    load_rows, load_totals = _parse_load_list_sheet(load_sheet) if load_sheet is not None else ([], {})
    top_products = _sorted_top_products(active_product_totals, 12)

    return {
        "day_name": day_name,
        "g2g_total": summary_rows["g2g_total"]["total"],
        "loaded_total": summary_rows["loaded_total"]["total"],
        "remaining_total": summary_rows["remaining_total"]["total"],
        "expected_store_total": summary_rows["expected_store_total"]["total"],
        "batch_count": len({row["batch_name"] for row in batch_rows}),
        "active_store_count": len(batch_rows),
        "total_weight": _decimal_or_none(sum(row["weight"] for row in batch_rows)),
        "total_value": _decimal_or_none(sum(row["value"] for row in batch_rows)),
        "store_rows": batch_rows,
        "top_products": top_products,
        "load_rows": load_rows,
        "load_1_total": _decimal_or_none(load_totals.get("load_1", 0.0)),
        "load_2_total": _decimal_or_none(load_totals.get("load_2", 0.0)),
        "load_3_total": _decimal_or_none(load_totals.get("load_3", 0.0)),
        "load_4_total": _decimal_or_none(load_totals.get("load_4", 0.0)),
        "load_total": _decimal_or_none(load_totals.get("total", 0.0)),
    }


def _parse_support_sheet(sheet: Any) -> dict[str, Any]:
    grid = _locate_support_grid(sheet)
    max_row, _ = _sheet_bounds(sheet)
    cell = sheet.cell
    product_headers = grid["product_headers"]
    store_col = grid["store_col"]
    g2g = _parse_metric_row(sheet, product_headers, PLANNING_SUMMARY_LABELS["g2g_total"])
    loaded = _parse_metric_row(sheet, product_headers, PLANNING_SUMMARY_LABELS["loaded_total"])
    remaining = _parse_metric_row(sheet, product_headers, PLANNING_SUMMARY_LABELS["remaining_total"])

    batch_rows = []
    if grid["header_row_index"] is not None and grid["store_col"]:
        current_batch = ""
        row_index = grid["header_row_index"] + 1
        while row_index <= max_row:
            batch_value = _string_value(cell(row_index, 1).value)
            if batch_value.lower().startswith("load "):
                current_batch = batch_value
                row_index += 1
                continue
            if _normalize_text(cell(row_index, store_col).value) == "load for external delivery":
                row_index += 1
                continue
            store_name = _string_value(cell(row_index, store_col).value)
            if not store_name:
                row_index += 1
                continue
            row_value = _float_value(cell(row_index, grid["value_col"]).value) if grid["value_col"] else 0.0
            if grid["value_col"] and row_value <= 0:
                row_index += 1
                continue
            items: list[dict[str, Any]] = []
            total_quantity = 0.0
            for column_index, sku_name in product_headers:
                quantity = _float_value(cell(row_index, column_index).value)
                if quantity <= 0:
                    continue
                total_quantity += quantity
                items.append({"sku": sku_name, "quantity": round(quantity, 2)})
            if total_quantity > 0:
                items.sort(key=lambda item: (-item["quantity"], item["sku"]))
                batch_rows.append(
                    {
                        "batch_name": current_batch or "Pending",
                        "store_name": store_name,
                        "contact": _string_value(cell(row_index, grid["contact_col"]).value) if grid["contact_col"] else "",
                        "region": _string_value(cell(row_index, grid["region_col"]).value) if grid["region_col"] else "",
                        "value": round(row_value, 2),
                        "total_quantity": round(total_quantity, 2),
                        "items": items,
                        "top_items": items[:5],
                        "reason_text": "Imported from Pending Orders",
                    }
                )
            row_index += 1

    return {
        "g2g_total": g2g["total"],
        "loaded_total": loaded["total"],
        "remaining_total": remaining["total"],
        "store_rows": batch_rows,
        "top_products": remaining["top_products"] or g2g["top_products"],
    }


def _parse_inventory_sheet(sheet: Any) -> dict[str, Any]:
    grid = _locate_support_grid(sheet)
    g2g = _parse_metric_row(sheet, grid["product_headers"], PLANNING_SUMMARY_LABELS["g2g_total"])
    remaining = _parse_metric_row(sheet, grid["product_headers"], PLANNING_SUMMARY_LABELS["remaining_total"])
    g2g_items = {item["sku"]: item["quantity"] for item in g2g["items"]}
    remaining_items = {item["sku"]: item["quantity"] for item in remaining["items"]}
    inventory_items = []
    for sort_order, sku_name in enumerate(sorted(set(g2g_items) | set(remaining_items)), start=1):
        inventory_items.append(
            {
                "sku": sku_name,
                "opening_g2g_qty": g2g_items.get(sku_name, 0.0),
                "opening_remaining_qty": remaining_items.get(sku_name, 0.0),
                "sort_order": sort_order,
            }
        )
    return {
        "g2g_total": g2g["total"],
        "remaining_total": remaining["total"],
        "top_products": g2g["top_products"],
        "inventory_items": inventory_items,
    }


def _parse_metric_row(sheet: Any, product_headers: list[tuple[int, str]], label: str) -> dict[str, Any]:
    row_index = _find_row_index(sheet, label)
    if row_index is None:
        return {"total": None, "top_products": [], "items": []}

    items = []
    total = 0.0
    for column_index, sku_name in product_headers:
        quantity = _float_value(sheet.cell(row_index, column_index).value)
        if quantity <= 0:
            continue
        items.append({"sku": sku_name, "quantity": round(quantity, 2)})
        total += quantity

    items.sort(key=lambda item: (-item["quantity"], item["sku"]))
    return {
        "total": _decimal_or_none(total),
        "top_products": items[:10],
        "items": items,
    }


def _parse_assumptions_sheet(sheet: Any) -> dict[str, int]:
    max_row, _ = _sheet_bounds(sheet)
    cell = sheet.cell
    sku_count = 0
    stores = set()
    for row_index in range(2, max_row + 1):
        sku_name = _string_value(cell(row_index, 2).value)
        store_name = _string_value(cell(row_index, 7).value)
        if sku_name:
            sku_count += 1
        if store_name:
            stores.add(store_name)
    return {"sku_count": sku_count, "store_count": len(stores)}


def _parse_fee_sheet(sheet: Any) -> dict[str, Any]:
    max_row, max_column = _sheet_bounds(sheet)
    cell = sheet.cell
    header_map = {}
    for column_index in range(1, max_column + 1):
        header_map[_normalize_text(cell(1, column_index).value)] = column_index

    sku_col = header_map.get("sku")
    delivery_value_col = header_map.get("retail deliveries value")
    payment_value_col = header_map.get("payment collection value")
    if not sku_col:
        return {"row_count": 0, "total_delivery_value": None, "total_payment_value": None, "top_rows": [], "rows": []}

    rows = []
    total_delivery_value = 0.0
    total_payment_value = 0.0
    for row_index in range(2, max_row + 1):
        sku_name = _string_value(cell(row_index, sku_col).value)
        if not sku_name:
            continue
        delivery_value = _float_value(cell(row_index, delivery_value_col).value) if delivery_value_col else 0.0
        payment_value = _float_value(cell(row_index, payment_value_col).value) if payment_value_col else 0.0
        total_delivery_value += delivery_value
        total_payment_value += payment_value
        rows.append(
            {
                "brand_partner": _string_value(cell(row_index, header_map.get("brand partner", 1)).value),
                "sku": sku_name,
                "vatable": _string_value(cell(row_index, header_map.get("vatable yes no", 1)).value),
                "retail_delivery_value": round(delivery_value, 2),
                "payment_collection_value": round(payment_value, 2),
            }
        )

    rows.sort(key=lambda row: (-row["retail_delivery_value"], row["sku"]))
    return {
        "row_count": len(rows),
        "total_delivery_value": _decimal_or_none(total_delivery_value),
        "total_payment_value": _decimal_or_none(total_payment_value),
        "top_rows": rows[:12],
        "rows": rows,
    }


def _parse_load_list_sheet(sheet: Any) -> tuple[list[dict[str, Any]], dict[str, float]]:
    max_row, _ = _sheet_bounds(sheet)
    cell = sheet.cell
    header_row_index = None
    for row_index in range(1, min(max_row, 12) + 1):
        if _normalize_text(cell(row_index, 1).value) == "sku":
            header_row_index = row_index
            break

    if header_row_index is None:
        return [], {}

    rows = []
    totals = {"load_1": 0.0, "load_2": 0.0, "load_3": 0.0, "load_4": 0.0, "total": 0.0}
    for row_index in range(header_row_index + 1, max_row + 1):
        sku_name = _string_value(cell(row_index, 1).value)
        if not sku_name:
            continue
        load_1 = _float_value(cell(row_index, 2).value)
        load_2 = _float_value(cell(row_index, 3).value)
        load_3 = _float_value(cell(row_index, 4).value)
        load_4 = _float_value(cell(row_index, 5).value)
        total = _float_value(cell(row_index, 6).value)
        if total <= 0 and load_1 <= 0 and load_2 <= 0 and load_3 <= 0 and load_4 <= 0:
            continue
        totals["load_1"] += load_1
        totals["load_2"] += load_2
        totals["load_3"] += load_3
        totals["load_4"] += load_4
        totals["total"] += total
        rows.append(
            {
                "sku": sku_name,
                "load_1": round(load_1, 2),
                "load_2": round(load_2, 2),
                "load_3": round(load_3, 2),
                "load_4": round(load_4, 2),
                "total": round(total, 2),
            }
        )
    rows.sort(key=lambda row: (-row["total"], row["sku"]))
    return rows, totals


def _resolve_fee_sheet(workbook: Any):
    for sheet_name in ("BP NEW FEES", "BP NEW FEES (2)"):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    return None


def _locate_day_grid(sheet: Any) -> dict[str, Any]:
    max_row, max_column = _sheet_bounds(sheet)
    cell = sheet.cell
    header_row_index = None
    store_col = None
    for row_index in range(1, min(max_row, 50) + 1):
        for column_index in range(1, min(max_column, 30) + 1):
            if _normalize_text(cell(row_index, column_index).value) == "load for external delivery":
                header_row_index = row_index
                store_col = column_index
                break
        if header_row_index is not None:
            break

    if header_row_index is None or store_col is None:
        raise LoadingTrackerError(
            f"We could not find the main planning table in the '{sheet.title}' sheet."
        )

    return _build_grid_definition(sheet, header_row_index, store_col)


def _locate_support_grid(sheet: Any) -> dict[str, Any]:
    cell = sheet.cell
    try:
        return _locate_day_grid(sheet)
    except LoadingTrackerError:
        product_header_row = _find_row_index(sheet, "PRODUCTS DESCRIPTIONS") or 2
        product_headers = _collect_product_headers(sheet, product_header_row, 10)
        return {
            "header_row_index": None,
            "store_col": None,
            "contact_col": None,
            "lp_col": None,
            "tier_col": None,
            "region_col": None,
            "weight_col": None,
            "value_col": None,
            "date_col": None,
            "product_headers": product_headers,
        }


def _build_grid_definition(sheet: Any, header_row_index: int, store_col: int) -> dict[str, Any]:
    _, max_column = _sheet_bounds(sheet)
    cell = sheet.cell
    header_values = {
        _normalize_text(cell(header_row_index, column_index).value): column_index
        for column_index in range(1, min(max_column, store_col + 20) + 1)
    }
    product_headers = _collect_product_headers(sheet, header_row_index, store_col + 1)

    return {
        "header_row_index": header_row_index,
        "store_col": store_col,
        "contact_col": header_values.get("contact"),
        "lp_col": header_values.get("lp"),
        "tier_col": header_values.get("tier"),
        "region_col": header_values.get("region"),
        "weight_col": header_values.get("weight"),
        "value_col": header_values.get("value"),
        "date_col": header_values.get("date"),
        "product_headers": product_headers,
    }


def _collect_product_headers(sheet: Any, header_row_index: int, start_column: int) -> list[tuple[int, str]]:
    _, max_column = _sheet_bounds(sheet)
    product_headers: list[tuple[int, str]] = []
    found_first_product = False
    for column_index in range(start_column, max_column + 1):
        raw_header = sheet.cell(header_row_index, column_index).value
        sku_name = _string_value(raw_header)
        if _is_product_header(sku_name):
            product_headers.append((column_index, sku_name))
            found_first_product = True
            continue
        if found_first_product:
            break
    return product_headers


def _is_product_header(value: Any) -> bool:
    text = _string_value(value)
    if not text:
        return False
    normalized = _normalize_text(text)
    if normalized in NON_PRODUCT_HEADERS:
        return False
    if normalized.isdigit():
        return False
    if _decimal_value(text) is not None:
        return False
    return any(character.isalpha() for character in text)


def _find_row_index(sheet: Any, label: str) -> int | None:
    max_row, max_column = _sheet_bounds(sheet)
    cell = sheet.cell
    target = _normalize_text(label)
    for row_index in range(1, min(max_row, 40) + 1):
        for column_index in range(1, min(max_column, 12) + 1):
            if _normalize_text(cell(row_index, column_index).value) == target:
                return row_index
    return None


def _count_note_lines(sheet: Any) -> int:
    max_row, max_column = _sheet_bounds(sheet)
    cell = sheet.cell
    count = 0
    for row_index in range(1, max_row + 1):
        values = [_string_value(cell(row_index, column_index).value) for column_index in range(1, max_column + 1)]
        if any(values):
            count += 1
    return count


def _sorted_top_products(product_totals: dict[str, float], limit: int) -> list[dict[str, Any]]:
    ranked = sorted(product_totals.items(), key=lambda item: (-item[1], item[0]))
    return [{"sku": sku, "quantity": round(quantity, 2)} for sku, quantity in ranked[:limit]]


def _batch_sort_key(batch_name: str) -> tuple[int, str]:
    label = batch_name.lower().replace("load", "").strip()
    try:
        return int(label), batch_name
    except ValueError:
        return 999, batch_name


def _resolve_loading_tracker_job_upload(job_id: str) -> Path | None:
    upload_root = Path(current_app.instance_path) / "loading_tracker_jobs"
    matches = sorted(upload_root.glob(f"{job_id}-*"))
    return matches[0] if matches else None


def _load_workbook_from_upload(file_storage: Any):
    if isinstance(file_storage, (str, Path)):
        try:
            return load_workbook(file_storage, data_only=True, keep_links=False)
        except Exception as exc:  # pragma: no cover
            raise LoadingTrackerError("The uploaded file could not be read as an Excel workbook.") from exc

    payload = file_storage.read()
    if hasattr(file_storage, "seek"):
        file_storage.seek(0)
    elif hasattr(file_storage, "stream") and hasattr(file_storage.stream, "seek"):
        file_storage.stream.seek(0)
    try:
        return load_workbook(BytesIO(payload), data_only=True, keep_links=False)
    except Exception as exc:  # pragma: no cover
        raise LoadingTrackerError("The uploaded file could not be read as an Excel workbook.") from exc


def _sheet_bounds(sheet: Any) -> tuple[int, int]:
    bounds = getattr(sheet, "_dala_bounds", None)
    if bounds is None:
        bounds = (sheet.max_row, sheet.max_column)
        setattr(sheet, "_dala_bounds", bounds)
    return bounds


def _string_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_text(value: Any) -> str:
    return " ".join(_string_value(value).lower().split())


def _float_value(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        return 0.0


def _decimal_value(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _decimal_or_none(value: float | Decimal | None) -> Decimal | None:
    if value in (None, ""):
        return None
    value_float = _float_value(value)
    if abs(value_float) < 0.000001:
        return None
    return Decimal(f"{value_float:.4f}")


def _clean_filename_part(value: str, fallback: str) -> str:
    import re

    cleaned = re.sub(INVALID_FILENAME_CHARS, " ", str(value))
    cleaned = cleaned.replace("_", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .-")
    if not cleaned:
        return fallback
    return cleaned[:90].rstrip(" .-")


def _report_import_progress(
    progress_callback: Callable[[int, str], None] | None,
    progress_percent: int,
    stage_label: str,
) -> None:
    if progress_callback is None:
        return
    progress_callback(progress_percent, stage_label)
