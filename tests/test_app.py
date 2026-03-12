from __future__ import annotations

from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from app import _database_uri, create_app
from loading_tracker_services import create_loading_tracker_import_job, run_loading_tracker_import_job
from models import (
    BrandPartnerRule,
    LoadingTrackerDailyCount,
    LoadingTrackerDay,
    LoadingTrackerEvent,
    LoadingTrackerImport,
    LoadingTrackerImportJob,
    LoadingTrackerRow,
    LoadingTrackerRowItem,
    LoadingTrackerTemplate,
    Product,
    ProductAlias,
    SalesOrderLine,
    SalesOrderRun,
    SkuAutomatorLine,
    SkuAutomatorRun,
    UploadRun,
    db,
)
from services import bootstrap_seed_uom_if_empty


def build_test_workbook(sheet_name: str = "tracker") -> BytesIO:
    workbook = Workbook()
    tracker = workbook.active
    tracker.title = sheet_name
    tracker.append(["Sales Order Number", "Stores", "SKU Alpha", "SKU Vanila"])
    tracker.append(["SO-100", "Market One", 2, 0.5])
    tracker.append(["SO-101", "Market Two", 1, 0])

    delivery = workbook.create_sheet("Delivery Invoice")
    delivery.append(
        [
            "Invoice Date",
            "Order Number",
            "Voucher Type Name",
            "Party name",
            "stock Item Name",
            "Quantity",
            "Rate",
            "Amount",
            "SalesLedger Name",
            "VAT",
            "VAT%",
            "VAT Amount",
        ]
    )

    uom = workbook.create_sheet("UOM")
    uom.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
    uom.append(["SKU Alpha", "ctn", "unt", 12, "Yes", 100])
    uom.append(["SKU Vanilla", "ctn", "unt", 12, "No", 80])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_item_list_export() -> BytesIO:
    payload = (
        "Item Id\tItem Name \t Cases Size \t Tax Rate\t Item PTR\t Status\n"
        "1\tExisting SKU Alpha\t12\t7.5\t999\tActive\n"
        "2\tFresh SKU Beta\t24\t\t3450\tActive\n"
        "3\tRetired SKU Gamma\t6\t\t1200\tInactive\n"
    )
    return BytesIO(payload.encode("utf-8"))


def build_loading_tracker_workbook() -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)

    def populate_day_sheet(sheet_name: str, store_name: str, qty_a: float, qty_b: float) -> None:
        sheet = workbook.create_sheet(sheet_name)
        sheet.cell(5, 9, "Expected in G2G For Loading")
        sheet.cell(5, 10, 10)
        sheet.cell(5, 11, 4)
        sheet.cell(6, 9, "TOTAL LOADED OUT FOR DELIVERY")
        sheet.cell(6, 10, qty_a)
        sheet.cell(6, 11, qty_b)
        sheet.cell(7, 9, "Remaining Inventory After Loading")
        sheet.cell(7, 10, 10 - qty_a)
        sheet.cell(7, 11, 4 - qty_b)
        sheet.cell(21, 9, "Expected in Store For Loading")
        sheet.cell(21, 10, 10)
        sheet.cell(21, 11, 4)
        sheet.cell(23, 1, "Load 1")
        sheet.cell(24, 2, "Contact")
        sheet.cell(24, 3, "LP")
        sheet.cell(24, 4, "Tier")
        sheet.cell(24, 5, "Region")
        sheet.cell(24, 6, "Weight")
        sheet.cell(24, 7, "Value")
        sheet.cell(24, 8, "Date")
        sheet.cell(24, 9, "LOAD FOR EXTERNAL DELIVERY")
        sheet.cell(24, 10, "SKU Alpha")
        sheet.cell(24, 11, "SKU Beta")
        sheet.cell(25, 2, "Area One")
        sheet.cell(25, 3, "Mr Tester")
        sheet.cell(25, 4, "Tier-1")
        sheet.cell(25, 5, "Reg 1")
        sheet.cell(25, 6, 5.5)
        sheet.cell(25, 7, 25000)
        sheet.cell(25, 9, store_name)
        sheet.cell(25, 10, qty_a)
        sheet.cell(25, 11, qty_b)
        sheet.cell(65, 1, "Load 2")
        sheet.cell(66, 2, "Contact")
        sheet.cell(66, 3, "LP")
        sheet.cell(66, 4, "Tier")
        sheet.cell(66, 5, "Region")
        sheet.cell(66, 6, "Weight")
        sheet.cell(66, 7, "Value")
        sheet.cell(66, 8, "Date")
        sheet.cell(66, 9, "LOAD FOR EXTERNAL DELIVERY")
        sheet.cell(67, 2, "Area Two")
        sheet.cell(67, 3, "Mr Runner")
        sheet.cell(67, 4, "Tier-2")
        sheet.cell(67, 5, "Reg 2")
        sheet.cell(67, 6, 2.0)
        sheet.cell(67, 7, 11000)
        sheet.cell(67, 9, "Second Store")
        sheet.cell(67, 10, 1)
        sheet.cell(67, 11, 0)

    def populate_load_list(sheet_name: str, qty_a: float, qty_b: float) -> None:
        sheet = workbook.create_sheet(sheet_name)
        sheet.cell(1, 1, sheet_name)
        sheet.cell(2, 1, "SKU")
        sheet.cell(2, 2, "Load 1")
        sheet.cell(2, 3, "Load 2")
        sheet.cell(2, 4, "Load 3")
        sheet.cell(2, 5, "Load 4")
        sheet.cell(2, 6, "TOTAL")
        sheet.cell(3, 1, "SKU Alpha")
        sheet.cell(3, 2, qty_a)
        sheet.cell(3, 6, qty_a)
        sheet.cell(4, 1, "SKU Beta")
        sheet.cell(4, 3, qty_b)
        sheet.cell(4, 6, qty_b)

    workbook.create_sheet("Pending Orders")
    pending = workbook["Pending Orders"]
    pending.cell(5, 9, "Expected in G2G For Loading")
    pending.cell(5, 10, 7)
    pending.cell(6, 9, "TOTAL LOADED OUT FOR DELIVERY")
    pending.cell(6, 10, 2)
    pending.cell(7, 9, "Remaining Inventory After Loading")
    pending.cell(7, 10, 5)
    pending.cell(23, 1, "Load 1")
    pending.cell(24, 2, "Contact")
    pending.cell(24, 5, "Region")
    pending.cell(24, 7, "Value")
    pending.cell(24, 9, "LOAD FOR EXTERNAL DELIVERY")
    pending.cell(24, 10, "SKU Alpha")
    pending.cell(25, 2, "Pending Area")
    pending.cell(25, 5, "Reg 8")
    pending.cell(25, 7, 15000)
    pending.cell(25, 9, "Pending Store")
    pending.cell(25, 10, 2)

    workbook.create_sheet("Opening Inventory")
    opening = workbook["Opening Inventory"]
    opening.cell(5, 9, "Expected in G2G For Loading")
    opening.cell(5, 10, 12)
    opening.cell(5, 11, 4)
    opening.cell(7, 9, "Remaining Inventory After Loading")
    opening.cell(7, 10, 10)
    opening.cell(7, 11, 4)
    opening.cell(2, 10, "SKU Alpha")
    opening.cell(2, 11, "SKU Beta")

    populate_day_sheet("Mon", "Store One", 2, 1)
    populate_day_sheet("Tues", "Store Two", 3, 0.5)
    populate_load_list("LL Mon", 3, 1)
    populate_load_list("LL Tue", 4, 0.5)

    assumptions = workbook.create_sheet("Assumptions")
    assumptions.append(
        ["Company", "SKU", "Units", "Pack Weight (kg)", "Value", "Min Inventory", "Retail Store", "Region"]
    )
    assumptions.append(["Comp A", "SKU Alpha", 12, 4, 1000, 2, "Store One", "Reg 1"])
    assumptions.append(["Comp B", "SKU Beta", 12, 3, 900, 1, "Store Two", "Reg 2"])

    fees = workbook.create_sheet("BP NEW FEES")
    fees.append(
        [
            "Brand Partner",
            "SKU",
            "Units",
            "Packaging",
            "Vatable (Yes/No)",
            "Retail Value Without VAT",
            "VAT",
            "Retail Value",
            "Retail Deliveries Value",
            "Percentage Retail Deliveries",
            "Payment Collection Value",
            "Percentage Payment Collection",
        ]
    )
    fees.append(["Comp A", "SKU Alpha", 12, "Cartons", "Yes", 900, 100, 1000, 120, 0.1, 80, 0.05])
    fees.append(["Comp B", "SKU Beta", 12, "Cartons", "No", 700, 0, 700, 90, 0.1, 50, 0.05])

    notes = workbook.create_sheet("NOTES FOR USER")
    notes.append(["Line 1"])
    notes.append(["Line 2"])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_loading_tracker_uom_workbook() -> BytesIO:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "UOM"
    ws.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
    ws.append(["SKU Alpha", "ctn", "unt", 12, "No", 100])
    ws.append(["SKU Beta", "ctn", "unt", 12, "No", 80])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_stock_category_summary_workbook() -> BytesIO:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Stock Category Summary"
    ws.append(["SKU", "Quantity", "(Alt. Units)", "Rate"])
    ws.append(["AMB- 100ml Carrot Oil (12x)", "ctn", "unt", 33600])
    ws.append(["1.5Litre Palm Oil (6X)", "ctn", "btt", 27759.69])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_stock_category_summary_uom_layout_workbook() -> BytesIO:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Stock Category Summary"
    ws.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
    ws.append(["AMB- 100ml Carrot Oil (12x)", "ctn", "unt", 12, "No", 33600])
    ws.append(["1.5Litre Palm Oil (6X)", "ctn", "btt", 6, "Yes", 27759.69])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_uom_workbook_from_rows(rows: list[list[object]]) -> BytesIO:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "UOM"
    ws.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
    for row in rows:
        ws.append(row)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_loading_tracker_workbook_with_duplicate_sections() -> BytesIO:
    workbook = Workbook()
    workbook_stream = build_loading_tracker_workbook()
    workbook_stream.seek(0)
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_stream)
    tues = workbook["Tues"]
    tues.cell(40, 4, "Value")
    tues.cell(40, 9, "Store Two")
    tues.cell(40, 10, 999999)
    tues.cell(40, 11, 888888)
    tues.cell(40, 7, None)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_pepup_orders_workbook() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Order Item List"
    sheet.append(
        [
            "Date of Order",
            "Time of Order",
            "Order Number",
            "Retailer Name",
            "Item Name",
            "UOM",
            "Quantity",
            "Price",
            "Total",
        ]
    )
    sheet.append(
        [
            "2026-03-11 00:00:00",
            "11:44:00",
            "17552475",
            "Globus Supermarket OKOTA",
            "SKU Alpha",
            "cases",
            2,
            107.5,
            215,
        ]
    )
    sheet.append(
        [
            "2026-03-11 00:00:00",
            "11:45:00",
            "17552476",
            "Value Exchange Supermarket KESHI",
            "SKU Vanilla",
            "pcs",
            6,
            10,
            60,
        ]
    )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_mixed_category_pepup_orders_workbook() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Order Item List"
    sheet.append(
        [
            "Date of Order",
            "Time of Order",
            "Order Number",
            "Retailer Name",
            "Item Name",
            "UOM",
            "Quantity",
            "Price",
            "Total",
        ]
    )
    sheet.append(
        [
            "2026-03-12 00:00:00",
            "08:15:00",
            "17575653",
            "Globus Supermarket OKOTA",
            "SKU Alpha",
            "cases",
            1,
            100,
            100,
        ]
    )
    sheet.append(
        [
            "2026-03-12 00:00:00",
            "08:16:00",
            "17575653",
            "Globus Supermarket OKOTA",
            "SKU Beta",
            "cases",
            1,
            120,
            120,
        ]
    )
    sheet.append(
        [
            "2026-03-12 00:00:00",
            "08:17:00",
            "17575653",
            "Globus Supermarket OKOTA",
            "SKU Vanilla",
            "cases",
            1,
            80,
            80,
        ]
    )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_pepup_orders_tab_delimited_export() -> BytesIO:
    rows = [
        [
            "Date of Order",
            "Time of Order",
            "Order Number",
            "Total Invoice Amount",
            "Total Number Of Item",
            "Distributor Name",
            "Distributor Code",
            "Distributor Type",
            "Zone",
            "State",
            "City",
            "Salesman Name",
            "Salesman Code",
            "Reporting To",
            "Retailer Name",
            "Retailer Code",
            "Retailer Market",
            "Retailer Address",
            "Retailer Group",
            "Retailer Channel",
            "Retailer Classification",
            "Retailer Type",
            "Retailer Phone No",
            "Display Outlet",
            "Comments",
            "Item Remark",
            "Reason",
            "Order Status",
            "Order By",
            "Item Name",
            "Item Code",
            "Item Grade",
            "Category",
            "Brand",
            "UOM",
            "Quantity",
            "Price",
            "Total",
            "Delivery Date",
            "Order Distance",
        ],
        [
            "12 Mar 2026",
            "11:16",
            "17575653",
            "105000.00",
            "1",
            "Dala",
            "32461",
            "",
            "",
            "Lagos",
            "Ikorodu Lga",
            "Folorunso Omotola",
            "15945",
            "Kehinde Ilori",
            "Grocery Bazaar IKORODU ODOGUNYAN",
            "1554545",
            "Reg 7a",
            "128 Ikorodu/Shagamu Road",
            "",
            "CORP",
            "Tier-2",
            "CORP",
            "7083376840",
            "",
            "Attached LPO",
            "YES",
            "",
            "New Adhoc Order",
            "Salesman",
            "SKU Alpha",
            "WJC3",
            "",
            "Juice",
            "BP Wilsons",
            "cases",
            "2",
            "100.00",
            "200.00",
            "",
            "1.3 KM",
        ],
        [
            "12 Mar 2026",
            "11:08",
            "17575278",
            "60.00",
            "1",
            "Dala",
            "32461",
            "",
            "",
            "Lagos",
            "Ikeja Lga",
            "Folorunso Omotola",
            "15945",
            "Kehinde Ilori",
            "Justrite OJODU",
            "1554689",
            "Reg 3a",
            "Ojodu Suit",
            "",
            "CORP",
            "Tier-1",
            "CORP",
            "8073893447",
            "",
            "Attached PO",
            "",
            "",
            "New Adhoc Order",
            "Salesman",
            "SKU Vanilla",
            "WJC7",
            "",
            "Juice",
            "BP Wilsons",
            "pcs",
            "6",
            "10.00",
            "60.00",
            "",
            "16.88 KM",
        ],
    ]
    payload = "\n".join("\t".join(row) for row in rows)
    return BytesIO(payload.encode("utf-8"))


def build_tally_export_workbook() -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales Order- Dala Register"
    sheet.append(["Sales Order- Dala Register"])
    sheet.append(["For 11-Mar-26"])
    sheet.append(
        [
            "Date",
            "Particulars",
            "Voucher No.",
            "Order Reference No.",
            "Terms of Payment",
            "Other References",
            "Terms of Delivery",
            "Value",
            "Gross Total",
            "Inventory Pool",
            "Vat on Sales- Registration No:- 31514501-0001",
        ]
    )
    sheet.append(["2026-03-11 00:00:00", "Store One", "2026/1928", "17551810", "", "", "", 200, 215, 200, 15])
    sheet.append(["", "SKU Alpha", "", "", "", "", "", 200, "", "", ""])
    sheet.append(["2026-03-11 00:00:00", "Store One", "2026/1929", "17551811", "", "", "", 60, 64.5, 60, 4.5])
    sheet.append(["", "SKU Vanilla", "", "", "", "", "", 60, "", "", ""])
    sheet.append(["2026-03-11 00:00:00", "Store Two", "2026/1930", "17551812", "", "", "", 100, 107.5, 100, 7.5])
    sheet.append(["", "SKU Alpha", "", "", "", "", "", 100, "", "", ""])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def test_end_to_end_review_and_alias_memory() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()
        workbook = build_test_workbook()

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(workbook.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert "UOM import complete" in response.get_data(as_text=True)

        response = client.post(
            "/runs/import",
            data={"tracker_workbook": (BytesIO(workbook.getvalue()), "tracker.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert response.status_code == 302
        assert "/review" in response.headers["Location"]

        run_id = response.headers["Location"].split("/runs/")[1].split("/")[0]

        with app.app_context():
            run = db.session.get(UploadRun, run_id)
            vanilla = db.session.query(Product).filter_by(sku_name="SKU Vanilla").one()
            assert run is not None
            assert run.rows_detected == 3
            assert run.rows_needing_review == 1

        response = client.post(
            f"/runs/{run_id}/review",
            data={"resolution::SKU Vanila": str(vanilla.id)},
            follow_redirects=True,
        )
        assert response.status_code == 200
        html = response.get_data(as_text=True)
        assert "rows are now ready for export" in html
        assert "Download VT" in html
        assert "Download NV" in html
        assert "Download BP" not in html

        download_vt = client.get(f"/runs/{run_id}/download?category=VT")
        assert download_vt.status_code == 200
        assert download_vt.mimetype == "application/vnd.ms-excel"
        disposition_vt = download_vt.headers["Content-Disposition"]
        assert "DALA Delivery Note -" in disposition_vt
        assert " VT " in disposition_vt or "-VT-" in disposition_vt or "- VT -" in disposition_vt

        download_nv = client.get(f"/runs/{run_id}/download?category=NV")
        assert download_nv.status_code == 200
        assert download_nv.mimetype == "application/vnd.ms-excel"
        disposition_nv = download_nv.headers["Content-Disposition"]
        assert "DALA Delivery Note -" in disposition_nv
        assert " NV " in disposition_nv or "-NV-" in disposition_nv or "- NV -" in disposition_nv

        with app.app_context():
            alias = db.session.query(ProductAlias).filter_by(alias_name="SKU Vanila").one_or_none()
            run = db.session.get(UploadRun, run_id)
            assert alias is not None
            assert run is not None
            assert run.status == "exported"
            assert run.invoice_date in disposition_vt
            assert "tracker" in disposition_vt
            db.session.remove()
            db.engine.dispose()


def test_database_url_uses_psycopg_driver(monkeypatch) -> None:
    monkeypatch.setenv("DATABASE_URL", "postgres://user:pass@db.example.com:5432/delivery")
    assert _database_uri("unused") == "postgresql+psycopg://user:pass@db.example.com:5432/delivery"

    monkeypatch.setenv("DATABASE_URL", "postgresql://user:pass@db.example.com:5432/delivery")
    assert _database_uri("unused") == "postgresql+psycopg://user:pass@db.example.com:5432/delivery"


def test_uom_import_sets_active_source_of_truth() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()

        first = Workbook()
        uom_one = first.active
        uom_one.title = "UOM"
        uom_one.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
        uom_one.append(["SKU Alpha", "ctn", "unt", 12, "Yes", 100])
        uom_one.append(["SKU Vanilla", "ctn", "unt", 12, "No", 80])
        first_bytes = BytesIO()
        first.save(first_bytes)
        first_bytes.seek(0)

        second = Workbook()
        uom_two = second.active
        uom_two.title = "UOM"
        uom_two.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
        uom_two.append(["SKU Alpha", "ctn", "unt", 12, "Yes", 120])
        second_bytes = BytesIO()
        second.save(second_bytes)
        second_bytes.seek(0)

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(first_bytes.getvalue()), "uom-one.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(second_bytes.getvalue()), "uom-two.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        with app.app_context():
            alpha = db.session.query(Product).filter_by(sku_name="SKU Alpha").one()
            vanilla = db.session.query(Product).filter_by(sku_name="SKU Vanilla").one()
            assert alpha.is_active is True
            assert float(alpha.price) == 120.0
            assert vanilla.is_active is False
            db.session.remove()
            db.engine.dispose()


def test_product_master_can_add_edit_deactivate_and_restore() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()

        response = client.post(
            "/products",
            data={
                "sku_name": "Manual Product",
                "price": "24500",
                "vatable": "yes",
                "uom": "ctn",
                "alt_uom": "pcs",
                "conversion": "12",
            },
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert "was added to the product master" in response.get_data(as_text=True)

        with app.app_context():
            product = db.session.query(Product).filter_by(sku_name="Manual Product").one()
            product_id = product.id

        response = client.post(
            f"/products/{product_id}/edit",
            data={
                "sku_name": "Manual Product Updated",
                "price": "26000",
                "vatable": "no",
                "uom": "ctn",
                "alt_uom": "pcs",
                "conversion": "6",
            },
            follow_redirects=True,
        )
        assert response.status_code == 200
        assert "was updated" in response.get_data(as_text=True)

        response = client.post(f"/products/{product_id}/deactivate", follow_redirects=True)
        assert response.status_code == 200
        assert "removed from the active product master" in response.get_data(as_text=True)

        response = client.post(f"/products/{product_id}/activate", follow_redirects=True)
        assert response.status_code == 200
        assert "is active again" in response.get_data(as_text=True)

        with app.app_context():
            product = db.session.get(Product, product_id)
            assert product is not None
            assert product.sku_name == "Manual Product Updated"
            assert float(product.price) == 26000.0
            assert product.vatable is False
            assert product.is_active is True
            db.session.remove()
            db.engine.dispose()


def test_product_master_global_search_finds_active_inactive_and_alias_matches() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        with app.app_context():
            active = Product(
                sku_name="Golden Palm Oil",
                normalized_name="golden palm oil",
                uom="ctn",
                alt_uom="btt",
                conversion=6,
                price=24500,
                vatable=False,
                is_active=True,
            )
            inactive = Product(
                sku_name="Legacy Palm Oil",
                normalized_name="legacy palm oil",
                uom="ctn",
                alt_uom="btt",
                conversion=6,
                price=19800,
                vatable=False,
                is_active=False,
            )
            other = Product(
                sku_name="Neutral Flour",
                normalized_name="neutral flour",
                uom="pck",
                alt_uom="scht",
                conversion=12,
                price=8700,
                vatable=False,
                is_active=True,
            )
            db.session.add_all([active, inactive, other])
            db.session.flush()
            db.session.add(
                ProductAlias(
                    alias_name="Old Palm Blend",
                    normalized_name="old palm blend",
                    match_method="approved-alias",
                    product_id=active.id,
                )
            )
            db.session.commit()

        client = app.test_client()
        response = client.get("/products?q=palm")
        html = response.get_data(as_text=True)

        assert response.status_code == 200
        assert "Golden Palm Oil" in html
        assert "Legacy Palm Oil" in html
        assert "Old Palm Blend" in html
        assert "Matching aliases" in html
        assert 'value="palm"' in html
        assert "Neutral Flour" not in html

        with app.app_context():
            db.session.remove()
            db.engine.dispose()


def test_product_master_page_exposes_uom_import_and_redirects_back() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()

        page = client.get("/database")
        html = page.get_data(as_text=True)
        assert page.status_code == 200
        assert "Refresh UOM source" in html
        assert "Import UOM file" in html

        response = client.post(
            "/uom/import",
            data={
                "return_to": "product_master",
                "uom_workbook": (BytesIO(build_loading_tracker_uom_workbook().getvalue()), "uom.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        assert response.status_code == 200
        assert "UOM import complete" in html
        assert "Product master" in html

        with app.app_context():
            assert db.session.query(Product).count() == 2
            db.session.remove()
            db.engine.dispose()


def test_stock_category_summary_workbook_imports_into_uom_master() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()

        seed = Workbook()
        seed_sheet = seed.active
        seed_sheet.title = "UOM"
        seed_sheet.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
        seed_sheet.append(["AMB- 100ml Carrot Oil (12x)", "ctn", "unt", 12, "Yes", 30000])
        seed_bytes = BytesIO()
        seed.save(seed_bytes)
        seed_bytes.seek(0)

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(seed_bytes.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        response = client.post(
            "/uom/import",
            data={
                "return_to": "product_master",
                "uom_workbook": (BytesIO(build_stock_category_summary_workbook().getvalue()), "Stock_Items_&_Pricelist.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        assert response.status_code == 200
        assert "UOM import complete" in html or "UOM update complete" in html

        with app.app_context():
            carrot = db.session.query(Product).filter_by(sku_name="AMB- 100ml Carrot Oil (12x)").one()
            palm = db.session.query(Product).filter_by(sku_name="1.5Litre Palm Oil (6X)").one()
            assert carrot.uom == "ctn"
            assert carrot.alt_uom == "unt"
            assert float(carrot.conversion) == 12.0
            assert carrot.vatable is True
            assert float(carrot.price) == 33600.0
            assert palm.uom == "ctn"
            assert palm.alt_uom == "btt"
            assert float(palm.conversion) == 6.0
            assert palm.vatable is False
            assert float(palm.price) == 27759.69
            db.session.remove()
            db.engine.dispose()


def test_uom_layout_on_stock_category_summary_sheet_imports_into_uom_master() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()

        response = client.post(
            "/uom/import",
            data={
                "return_to": "product_master",
                "uom_workbook": (
                    BytesIO(build_stock_category_summary_uom_layout_workbook().getvalue()),
                    "Stock_Items_&_Pricelist (1).xlsx",
                ),
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        assert response.status_code == 200
        assert "UOM import complete" in html or "UOM update complete" in html

        with app.app_context():
            carrot = db.session.query(Product).filter_by(sku_name="AMB- 100ml Carrot Oil (12x)").one()
            palm = db.session.query(Product).filter_by(sku_name="1.5Litre Palm Oil (6X)").one()
            assert carrot.uom == "ctn"
            assert carrot.alt_uom == "unt"
            assert float(carrot.conversion) == 12.0
            assert carrot.vatable is False
            assert float(carrot.price) == 33600.0
            assert palm.uom == "ctn"
            assert palm.alt_uom == "btt"
            assert float(palm.conversion) == 6.0
            assert palm.vatable is True
            assert float(palm.price) == 27759.69
            db.session.remove()
            db.engine.dispose()


def test_uom_replace_sync_matches_minor_name_changes_without_false_removals() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()

        original_rows = [
            ["AF- 1.5Litre Palm Oil (6X)", "ctn", "btt", 6, "No", 5500],
            ["WE Lemongrass Tea (56g)", "pck", "sachet", 12, "No", 2850],
        ]
        refreshed_rows = [
            ["1.5Litre Palm Oil (6X)", "ctn", "btt", 6, "No", 27759.69],
            ["WHE- 56g Lemongrass Tea (12x)_x000D_", "pck", "sachet", 12, "No", 2850],
        ]

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(build_uom_workbook_from_rows(original_rows).getvalue()), "uom-old.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        assert response.status_code == 200

        response = client.post(
            "/uom/import",
            data={
                "return_to": "product_master",
                "uom_workbook": (BytesIO(build_uom_workbook_from_rows(refreshed_rows).getvalue()), "uom-new.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        assert response.status_code == 200

        with app.app_context():
            all_products = list(db.session.query(Product).order_by(Product.sku_name.asc()))
            assert len(all_products) == 2

            palm = db.session.query(Product).filter_by(sku_name="1.5Litre Palm Oil (6X)").one()
            tea = db.session.query(Product).filter_by(sku_name="WHE- 56g Lemongrass Tea (12x)_x000D_").one()
            assert palm.is_active is True
            assert tea.is_active is True
            assert float(palm.price) == 27759.69
            assert db.session.query(Product).filter_by(sku_name="AF- 1.5Litre Palm Oil (6X)").count() == 0
            assert db.session.query(Product).filter_by(sku_name="WE Lemongrass Tea (56g)").count() == 0
            db.session.remove()
            db.engine.dispose()


def test_uom_replace_sync_does_not_auto_merge_ambiguous_name_changes() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()

        original_rows = [
            ["AGS- 200g Jadens Fruitamil (12x)", "ctn", "sachet", 12, "No", 1000],
            ["AGS- 400g Jadens Fruitamil (12x)", "ctn", "sachet", 12, "No", 1800],
        ]
        refreshed_rows = [
            ["AGS- Jadens Fruitamil (12x)", "ctn", "sachet", 12, "No", 1500],
        ]

        client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(build_uom_workbook_from_rows(original_rows).getvalue()), "uom-old.xlsx")},
            content_type="multipart/form-data",
        )
        client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(build_uom_workbook_from_rows(refreshed_rows).getvalue()), "uom-new.xlsx")},
            content_type="multipart/form-data",
        )

        with app.app_context():
            ambiguous = db.session.query(Product).filter_by(sku_name="AGS- Jadens Fruitamil (12x)").one()
            old_200g = db.session.query(Product).filter_by(sku_name="AGS- 200g Jadens Fruitamil (12x)").one()
            old_400g = db.session.query(Product).filter_by(sku_name="AGS- 400g Jadens Fruitamil (12x)").one()
            assert ambiguous.is_active is True
            assert old_200g.is_active is False
            assert old_400g.is_active is False
            assert db.session.query(Product).count() == 3
            db.session.remove()
            db.engine.dispose()


def test_manual_product_stays_active_after_new_uom_import() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()

        workbook = build_test_workbook()
        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(workbook.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        response = client.post(
            "/products",
            data={
                "sku_name": "Manual Product",
                "price": "24500",
                "vatable": "yes",
                "uom": "ctn",
                "alt_uom": "pcs",
                "conversion": "12",
            },
        )
        assert response.status_code == 302

        second = Workbook()
        uom_two = second.active
        uom_two.title = "UOM"
        uom_two.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
        uom_two.append(["SKU Alpha", "ctn", "unt", 12, "Yes", 120])
        second_bytes = BytesIO()
        second.save(second_bytes)
        second_bytes.seek(0)

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(second_bytes.getvalue()), "uom-two.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        with app.app_context():
            alpha = db.session.query(Product).filter_by(sku_name="SKU Alpha").one()
            vanilla = db.session.query(Product).filter_by(sku_name="SKU Vanilla").one()
            manual = db.session.query(Product).filter_by(sku_name="Manual Product").one()
            assert alpha.is_active is True
            assert vanilla.is_active is False
            assert manual.is_active is True
            db.session.remove()
            db.engine.dispose()


def test_seed_bootstrap_populates_database_when_empty() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        with app.app_context():
            seeded = bootstrap_seed_uom_if_empty()
            assert seeded is not None
            assert seeded.product_count > 0
            assert db.session.query(Product).count() == seeded.product_count
            db.session.remove()
            db.engine.dispose()


def test_tracker_import_accepts_mon_test_sheet_name() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()
        workbook = build_test_workbook("Mon_Test")

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(workbook.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        response = client.post(
            "/runs/import",
            data={"tracker_workbook": (BytesIO(workbook.getvalue()), "LT-DN Test 2.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert response.status_code == 302
        assert "/review" in response.headers["Location"]

        with app.app_context():
            db.session.remove()
            db.engine.dispose()


def test_item_list_export_can_merge_into_uom_without_replacing_existing() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()
        workbook = build_test_workbook()

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(workbook.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        assert response.status_code == 200

        response = client.post(
            "/products",
            data={
                "sku_name": "Existing SKU Alpha",
                "price": "1200",
                "vatable": "yes",
                "uom": "ctn",
                "alt_uom": "pcs",
                "conversion": "12",
            },
            follow_redirects=True,
        )
        assert response.status_code == 200

        item_list = build_item_list_export()
        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(item_list.getvalue()), "item_list (5).xls")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        assert response.status_code == 200
        assert "1 new product rows were added and 1 existing items were skipped and 1 items were moved to inactive" in html

        with app.app_context():
            existing = db.session.query(Product).filter_by(sku_name="Existing SKU Alpha").one()
            fresh = db.session.query(Product).filter_by(sku_name="Fresh SKU Beta").one()
            retired = db.session.query(Product).filter_by(sku_name="Retired SKU Gamma").one()
            assert float(existing.price) == 1200.0
            assert fresh.is_active is True
            assert float(fresh.price) == 3450.0
            assert fresh.uom == "ctn"
            assert fresh.alt_uom == "unt"
            assert fresh.vatable is False
            assert retired.is_active is False
            db.session.remove()
            db.engine.dispose()


def test_tracker_ignores_known_inactive_products() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()
        workbook = build_test_workbook()

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(workbook.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        response = client.post(
            "/products",
            data={
                "sku_name": "Mamaologi SKU",
                "price": "1",
                "vatable": "no",
                "uom": "ctn",
                "alt_uom": "pcs",
                "conversion": "10",
            },
        )
        assert response.status_code == 302

        with app.app_context():
            product = db.session.query(Product).filter_by(sku_name="Mamaologi SKU").one()
            product_id = product.id

        response = client.post(f"/products/{product_id}/deactivate", follow_redirects=True)
        assert response.status_code == 200

        tracker = Workbook()
        sheet = tracker.active
        sheet.title = "Sheet1"
        sheet.append(["Sales Order Number", "Stores", "Mamaologi SKU"])
        sheet.append(["SO-200", "Market One", 3])
        tracker_bytes = BytesIO()
        tracker.save(tracker_bytes)
        tracker_bytes.seek(0)

        response = client.post(
            "/runs/import",
            data={"tracker_workbook": (BytesIO(tracker_bytes.getvalue()), "mamaologi.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        assert response.status_code == 200
        assert "Everything active is matched and ready." in html
        assert "Ignored" in html

        with app.app_context():
            db.session.remove()
            db.engine.dispose()


def test_review_can_mark_source_sku_inactive_directly() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()
        workbook = build_test_workbook()

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(workbook.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        response = client.post(
            "/runs/import",
            data={"tracker_workbook": (BytesIO(workbook.getvalue()), "tracker.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert response.status_code == 302
        run_id = response.headers["Location"].split("/runs/")[1].split("/")[0]

        response = client.post(
            f"/runs/{run_id}/review",
            data={"mark_inactive": "SKU Vanila"},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        assert response.status_code == 200
        assert "was moved to inactive" in html
        assert "Inactive items skipped" in html
        assert "SKU Vanila" in html

        ignored_download = client.get(f"/runs/{run_id}/ignored/download")
        assert ignored_download.status_code == 200
        assert ignored_download.mimetype == "application/vnd.ms-excel"
        assert "DALA Ignored Items" in ignored_download.headers["Content-Disposition"]
        assert ignored_download.data

        with app.app_context():
            inactive_product = db.session.query(Product).filter_by(sku_name="SKU Vanila").one()
            run = db.session.get(UploadRun, run_id)
            assert inactive_product.is_active is False
            assert run is not None
            assert run.rows_needing_review == 0
            db.session.remove()
            db.engine.dispose()


def test_general_dashboard_and_delivery_note_module_render() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()

        response = client.get("/")
        assert response.status_code == 200
        html = response.get_data(as_text=True)
        assert "Planning, loading, exports, and future automators in one calm workspace." in html
        assert "Delivery Note" in html
        assert "Loading Tracker" in html
        assert "SKU Automator" in html

        response = client.get("/delivery-note")
        assert response.status_code == 200
        assert "Delivery Note studio" in response.get_data(as_text=True)

        with app.app_context():
            db.session.remove()
            db.engine.dispose()


def test_loading_tracker_import_builds_initial_module_views() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
                "LOADING_TRACKER_IMPORT_SYNC": True,
            }
        )

        client = app.test_client()
        workbook = build_loading_tracker_workbook()

        response = client.post(
            "/loading-tracker/import",
            data={"loading_tracker_workbook": (BytesIO(workbook.getvalue()), "Week 4 Loading Tracker.xlsx")},
            content_type="multipart/form-data",
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        payload = response.get_json()
        assert response.status_code == 202
        assert payload is not None
        assert payload["job"]["status"] == "completed"
        assert payload["job"]["tracker_import_id"]
        import_id = payload["job"]["tracker_import_id"]

        with app.app_context():
            tracker_import = db.session.query(LoadingTrackerImport).one()
            mon = db.session.query(LoadingTrackerDay).filter_by(day_name="Mon").one()
            assert tracker_import.assumptions_sku_count == 2
            assert len(tracker_import.pending_rows_json) == 1
            assert len(tracker_import.planning_rows) == 5
            assert len(tracker_import.inventory_items) == 2
            assert mon.batch_count == 2
            assert mon.active_store_count == 2
            assert float(mon.load_total) == 4.0
            assert tracker_import.id == import_id

        status_response = client.get(payload["status_url"])
        assert status_response.status_code == 200
        status_payload = status_response.get_json()
        assert status_payload["status"] == "completed"

        day_response = client.get(f"/loading-tracker/imports/{import_id}/days/Mon")
        assert day_response.status_code == 200
        day_html = day_response.get_data(as_text=True)
        assert "Mon planning workspace." in day_html
        assert "Generated LL" in day_html
        assert "Store One" in day_html

        pending_response = client.get(f"/loading-tracker/imports/{import_id}/pending")
        assert pending_response.status_code == 200
        pending_html = pending_response.get_data(as_text=True)
        assert "Pending is now a live queue" in pending_html
        assert "Pending Store" in pending_html

        with app.app_context():
            db.session.remove()
            db.engine.dispose()


def test_loading_tracker_live_move_and_inventory_adjustment() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
                "LOADING_TRACKER_IMPORT_SYNC": True,
            }
        )

        client = app.test_client()
        workbook = build_loading_tracker_workbook()

        response = client.post(
            "/loading-tracker/import",
            data={"loading_tracker_workbook": (BytesIO(workbook.getvalue()), "Week 4 Loading Tracker.xlsx")},
            content_type="multipart/form-data",
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        assert response.status_code == 202

        with app.app_context():
            tracker_import = db.session.query(LoadingTrackerImport).one()
            pending_row = db.session.query(LoadingTrackerRow).filter_by(row_state="pending").one()
            import_id = tracker_import.id
            pending_row_id = pending_row.id

        response = client.post(
            f"/loading-tracker/imports/{import_id}/rows/{pending_row_id}/move",
            data={"target_day_name": "Tues"},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        assert response.status_code == 200
        assert "moved into Tues" in html
        assert "Pending Store" in html

        response = client.post(
            f"/loading-tracker/imports/{import_id}/inventory",
            data={"sku_name": "SKU Alpha", "added_qty": "3", "opening_g2g_qty": "12", "opening_remaining_qty": "10"},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        assert response.status_code == 200
        assert "Inventory for" in html and "was updated" in html

        with app.app_context():
            pending_count = db.session.query(LoadingTrackerRow).filter_by(row_state="pending").count()
            tracker_import = db.session.query(LoadingTrackerImport).one()
            sku_alpha = next(item for item in tracker_import.inventory_items if item.sku_name == "SKU Alpha")
            assert pending_count == 0
            assert float(sku_alpha.added_qty) == 3.0
            db.session.remove()
            db.engine.dispose()


def test_loading_tracker_ignores_duplicate_weight_or_value_sections() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
                "LOADING_TRACKER_IMPORT_SYNC": True,
            }
        )

        client = app.test_client()
        workbook = build_loading_tracker_workbook_with_duplicate_sections()

        response = client.post(
            "/loading-tracker/import",
            data={"loading_tracker_workbook": (BytesIO(workbook.getvalue()), "Week 4 Loading Tracker.xlsx")},
            content_type="multipart/form-data",
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        assert response.status_code == 202

        with app.app_context():
            tues = db.session.query(LoadingTrackerDay).filter_by(day_name="Tues").one()
            assert float(tues.loaded_total) == 3.5
            assert tues.active_store_count == 2
            db.session.remove()
            db.engine.dispose()


def test_loading_tracker_reset_clears_live_week_but_keeps_master_data() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
                "LOADING_TRACKER_IMPORT_SYNC": True,
            }
        )

        client = app.test_client()

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(build_loading_tracker_uom_workbook().getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        response = client.post(
            "/loading-tracker/import",
            data={"loading_tracker_workbook": (BytesIO(build_loading_tracker_workbook().getvalue()), "Week 4 Loading Tracker.xlsx")},
            content_type="multipart/form-data",
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        assert response.status_code == 202

        with app.app_context():
            assert db.session.query(LoadingTrackerImport).count() == 1
            assert db.session.query(LoadingTrackerImportJob).count() == 1
            assert db.session.query(LoadingTrackerRow).count() > 0
            assert db.session.query(Product).count() == 2

        response = client.post("/loading-tracker/reset", follow_redirects=True)
        assert response.status_code == 200
        html = response.get_data(as_text=True)
        assert "Loading Tracker was cleared for a clean re-import" in html

        with app.app_context():
            assert db.session.query(LoadingTrackerImport).count() == 0
            assert db.session.query(LoadingTrackerImportJob).count() == 0
            assert db.session.query(LoadingTrackerDay).count() == 0
            assert db.session.query(LoadingTrackerRow).count() == 0
            assert db.session.query(LoadingTrackerEvent).count() == 0
            assert db.session.query(Product).count() == 2
            db.session.remove()
            db.engine.dispose()


def test_loading_tracker_template_can_seed_future_weeks_without_workbook() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
                "LOADING_TRACKER_IMPORT_SYNC": True,
            }
        )

        client = app.test_client()

        response = client.post(
            "/loading-tracker/import",
            data={"loading_tracker_workbook": (BytesIO(build_loading_tracker_workbook().getvalue()), "Week 4 Loading Tracker.xlsx")},
            content_type="multipart/form-data",
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        assert response.status_code == 202

        with app.app_context():
            source_import = db.session.query(LoadingTrackerImport).one()
            source_import_id = source_import.id
            source_pending_count = db.session.query(LoadingTrackerRow).filter_by(tracker_import_id=source_import_id, row_state="pending").count()
            source_planned_count = db.session.query(LoadingTrackerRow).filter_by(tracker_import_id=source_import_id, row_state="planned").count()

        capture = client.post(
            "/loading-tracker/template/capture",
            data={"source_import_id": source_import_id, "template_name": "Main planning template"},
            follow_redirects=True,
        )
        assert capture.status_code == 200
        assert "active backend planning template" in capture.get_data(as_text=True)

        start_from_template = client.post(
            "/loading-tracker/template/start-week",
            data={"source_import_id": source_import_id},
            follow_redirects=True,
        )
        assert start_from_template.status_code == 200
        assert "created from the backend template" in start_from_template.get_data(as_text=True)

        with app.app_context():
            assert db.session.query(LoadingTrackerTemplate).count() == 1
            imports = db.session.query(LoadingTrackerImport).order_by(LoadingTrackerImport.created_at.asc()).all()
            assert len(imports) == 2
            new_import = imports[-1]
            assert len(new_import.days) == 2
            assert db.session.query(LoadingTrackerRow).filter_by(tracker_import_id=new_import.id, row_state="planned").count() == source_planned_count
            assert db.session.query(LoadingTrackerRow).filter_by(tracker_import_id=new_import.id, row_state="pending").count() == source_pending_count

        reset = client.post("/loading-tracker/reset", follow_redirects=True)
        assert reset.status_code == 200

        restart = client.post(
            "/loading-tracker/template/start-week",
            follow_redirects=True,
        )
        assert restart.status_code == 200
        assert "created from the backend template" in restart.get_data(as_text=True)

        with app.app_context():
            assert db.session.query(LoadingTrackerTemplate).count() == 1
            assert db.session.query(LoadingTrackerImport).count() == 1
            restarted_import = db.session.query(LoadingTrackerImport).one()
            assert db.session.query(LoadingTrackerRow).filter_by(tracker_import_id=restarted_import.id, row_state="planned").count() == source_planned_count
            assert db.session.query(LoadingTrackerRow).filter_by(tracker_import_id=restarted_import.id, row_state="pending").count() == 0
            db.session.remove()
            db.engine.dispose()


def test_loading_tracker_counts_handoff_history_and_carry_forward() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
                "LOADING_TRACKER_IMPORT_SYNC": True,
            }
        )

        client = app.test_client()

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(build_loading_tracker_uom_workbook().getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        response = client.post(
            "/loading-tracker/import",
            data={"loading_tracker_workbook": (BytesIO(build_loading_tracker_workbook().getvalue()), "Week 4 Loading Tracker.xlsx")},
            content_type="multipart/form-data",
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        assert response.status_code == 202

        with app.app_context():
            tracker_import = db.session.query(LoadingTrackerImport).one()
            mon = db.session.query(LoadingTrackerDay).filter_by(day_name="Mon").one()
            mon_first_row = (
                db.session.query(LoadingTrackerRow)
                .filter_by(day_id=mon.id, row_state="planned")
                .order_by(LoadingTrackerRow.sort_order.asc())
                .first()
            )
            assert mon_first_row is not None
            import_id = tracker_import.id
            row_id = mon_first_row.id

        counts = client.post(
            f"/loading-tracker/imports/{import_id}/days/Mon/counts",
            data={"count::SKU Alpha": "9", "count::SKU Beta": "3.5"},
            follow_redirects=True,
        )
        counts_html = counts.get_data(as_text=True)
        assert counts.status_code == 200
        assert "Start-of-day physical count saved for Mon" in counts_html
        assert "Planning pulse" in counts_html

        with app.app_context():
            assert db.session.query(LoadingTrackerDailyCount).filter_by(day_id=mon.id).count() == 2
            discrepancy_event = (
                db.session.query(LoadingTrackerEvent)
                .filter_by(day_id=mon.id, event_type="inventory_discrepancy_alert")
                .one()
            )
            assert discrepancy_event.details_json["delivery_status"] == "not_configured"
            assert discrepancy_event.details_json["variance_count"] == 2

        invalid_pending = client.post(
            f"/loading-tracker/imports/{import_id}/rows/{row_id}/move",
            data={"target_day_name": "__pending__"},
            follow_redirects=True,
        )
        invalid_html = invalid_pending.get_data(as_text=True)
        assert invalid_pending.status_code == 200
        assert "Choose a pending reason" in invalid_html

        valid_pending = client.post(
            f"/loading-tracker/imports/{import_id}/rows/{row_id}/move",
            data={
                "target_day_name": "__pending__",
                "reason_code": "stock_shortage",
                "reason_note": "SKU Alpha short",
            },
            follow_redirects=True,
        )
        valid_pending_html = valid_pending.get_data(as_text=True)
        assert valid_pending.status_code == 200
        assert "now waiting in Pending" in valid_pending_html
        assert "Insufficient stock: SKU Alpha short" in valid_pending_html

        tues_day = client.get(f"/loading-tracker/imports/{import_id}/days/Tues")
        tues_html = tues_day.get_data(as_text=True)
        assert tues_day.status_code == 200
        assert "Pack Breaker" in tues_html
        assert "SKU Beta" in tues_html

        handoff = client.post(
            f"/loading-tracker/imports/{import_id}/days/Tues/handoff",
            follow_redirects=True,
        )
        handoff_html = handoff.get_data(as_text=True)
        assert handoff.status_code == 200
        assert "final adjusted day plan has been handed off" in handoff_html
        assert "Download NV" in handoff_html
        assert "Download VT" not in handoff_html
        assert "Download BP" not in handoff_html

        with app.app_context():
            upload_run = db.session.query(UploadRun).one()
            handoff_run_id = upload_run.id

        handoff_download = client.get(f"/runs/{handoff_run_id}/download?category=NV")
        assert handoff_download.status_code == 200
        assert handoff_download.mimetype == "application/vnd.ms-excel"

        history = client.get(f"/loading-tracker/imports/{import_id}/history")
        history_html = history.get_data(as_text=True)
        assert history.status_code == 200
        assert "Track planning changes" in history_html
        assert "Insufficient stock" in history_html

        history_download = client.get(f"/loading-tracker/imports/{import_id}/history/download")
        assert history_download.status_code == 200
        assert history_download.mimetype == "text/csv"
        assert b"moved_to_pending" in history_download.data

        carry = client.post(
            f"/loading-tracker/imports/{import_id}/carry-forward",
            follow_redirects=True,
        )
        carry_html = carry.get_data(as_text=True)
        assert carry.status_code == 200
        assert "fresh week was created" in carry_html

        with app.app_context():
            imports = db.session.query(LoadingTrackerImport).order_by(LoadingTrackerImport.created_at.asc()).all()
            assert len(imports) == 2
            latest_import = imports[-1]
            assert len(latest_import.days) == 6
            assert db.session.query(LoadingTrackerRow).filter_by(tracker_import_id=latest_import.id, row_state="pending").count() == 2
            assert float(latest_import.opening_g2g_total) == 7.0
            assert db.session.query(LoadingTrackerEvent).filter_by(tracker_import_id=latest_import.id).count() >= 1
            assert db.session.query(UploadRun).count() == 1
            db.session.remove()
            db.engine.dispose()


def test_loading_tracker_import_job_can_resume_from_saved_upload() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        with app.app_context():
            job = create_loading_tracker_import_job("Week 4 Loading Tracker.xlsx")
            job_id = job.id
            upload_root = Path(app.instance_path) / "loading_tracker_jobs"
            upload_root.mkdir(parents=True, exist_ok=True)
            saved_path = upload_root / f"{job_id}-Week-4-Loading-Tracker.xlsx"
            saved_path.write_bytes(build_loading_tracker_workbook().getvalue())

            run_loading_tracker_import_job(job_id)

            refreshed_job = db.session.get(LoadingTrackerImportJob, job_id)
            assert refreshed_job is not None
            assert refreshed_job.status == "completed"
            assert refreshed_job.tracker_import_id
            assert db.session.query(LoadingTrackerImport).count() == 1
            assert saved_path.exists() is False
            db.session.remove()
            db.engine.dispose()


def test_loading_tracker_day_suggestions_and_bulk_moves() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
                "LOADING_TRACKER_IMPORT_SYNC": True,
            }
        )

        client = app.test_client()

        response = client.post(
            "/loading-tracker/import",
            data={"loading_tracker_workbook": (BytesIO(build_loading_tracker_workbook().getvalue()), "Week 4 Loading Tracker.xlsx")},
            content_type="multipart/form-data",
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        assert response.status_code == 202

        with app.app_context():
            tracker_import = db.session.query(LoadingTrackerImport).one()
            mon = db.session.query(LoadingTrackerDay).filter_by(tracker_import_id=tracker_import.id, day_name="Mon").one()
            mon_row_ids = [
                row.id
                for row in db.session.query(LoadingTrackerRow)
                .filter_by(day_id=mon.id, row_state="planned")
                .order_by(LoadingTrackerRow.sort_order.asc())
                .all()
            ]
            import_id = tracker_import.id

        counts = client.post(
            f"/loading-tracker/imports/{import_id}/days/Mon/counts",
            data={"count::SKU Alpha": "1", "count::SKU Beta": "4"},
            follow_redirects=True,
        )
        counts_html = counts.get_data(as_text=True)
        assert counts.status_code == 200
        assert "Suggested planner actions" in counts_html
        assert "Apply suggestion" in counts_html

        bulk_to_pending = client.post(
            f"/loading-tracker/imports/{import_id}/days/Mon/bulk-move",
            data={
                "target_day_name": "__pending__",
                "reason_code": "stock_shortage",
                "reason_note": "bulk relief",
                "row_ids": [str(mon_row_ids[0]), str(mon_row_ids[1])],
            },
            follow_redirects=True,
        )
        bulk_pending_html = bulk_to_pending.get_data(as_text=True)
        assert bulk_to_pending.status_code == 200
        assert "2 planner row(s) were moved into Pending." in bulk_pending_html

        with app.app_context():
            pending_rows = list(
                db.session.query(LoadingTrackerRow)
                .filter_by(tracker_import_id=import_id, row_state="pending")
                .order_by(LoadingTrackerRow.id.asc())
                .all()
            )
            moved_pending_ids = [row.id for row in pending_rows if row.store_name in {"Store One", "Second Store"}]
            assert len(moved_pending_ids) == 2
            assert len(pending_rows) == 3

        bulk_to_tues = client.post(
            f"/loading-tracker/imports/{import_id}/pending/bulk-move",
            data={
                "target_day_name": "Tues",
                "row_ids": [str(row_id) for row_id in moved_pending_ids],
            },
            follow_redirects=True,
        )
        bulk_tues_html = bulk_to_tues.get_data(as_text=True)
        assert bulk_to_tues.status_code == 200
        assert "2 pending row(s) were moved into Tues." in bulk_tues_html
        assert "Tues planning workspace." in bulk_tues_html

        with app.app_context():
            tues = db.session.query(LoadingTrackerDay).filter_by(tracker_import_id=import_id, day_name="Tues").one()
            tues_planned = db.session.query(LoadingTrackerRow).filter_by(day_id=tues.id, row_state="planned").count()
            pending_count = db.session.query(LoadingTrackerRow).filter_by(tracker_import_id=import_id, row_state="pending").count()
            assert tues_planned == 4
            assert pending_count == 1
            db.session.remove()
            db.engine.dispose()


def test_sku_automator_run_can_start_loading_tracker_week_directly() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
                "LOADING_TRACKER_IMPORT_SYNC": True,
            }
        )

        client = app.test_client()
        uom = Workbook()
        sheet = uom.active
        sheet.title = "UOM"
        sheet.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
        sheet.append(["SKU Alpha", "ctn", "pcs", 12, "Yes", 100])
        sheet.append(["SKU Vanilla", "ctn", "pcs", 12, "No", 120])
        uom_bytes = BytesIO()
        uom.save(uom_bytes)
        uom_bytes.seek(0)

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(uom_bytes.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        tracker_upload = client.post(
            "/loading-tracker/import",
            data={"loading_tracker_workbook": (BytesIO(build_loading_tracker_workbook().getvalue()), "Week 4 Loading Tracker.xlsx")},
            content_type="multipart/form-data",
            headers={"Accept": "application/json", "X-Requested-With": "XMLHttpRequest"},
        )
        assert tracker_upload.status_code == 202

        with app.app_context():
            source_import = db.session.query(LoadingTrackerImport).one()
            source_import_id = source_import.id

        capture = client.post(
            "/loading-tracker/template/capture",
            data={"source_import_id": source_import_id, "template_name": "Main planning template"},
            follow_redirects=True,
        )
        assert capture.status_code == 200

        upload = client.post(
            "/sku-automator/import",
            data={"sku_automator_workbook": (BytesIO(build_tally_export_workbook().getvalue()), "salesordertest.xls")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert upload.status_code == 302
        run_id = upload.headers["Location"].split("/sku-automator/runs/")[1]

        create_week = client.post(
            f"/sku-automator/runs/{run_id}/loading-tracker",
            data={
                "source_import_id": source_import_id,
                "week_label": "Week 5 from SKU Automator",
                "target_day_name": "Tues",
            },
            follow_redirects=True,
        )
        create_week_html = create_week.get_data(as_text=True)
        assert create_week.status_code == 200
        assert "SKU Automator matrix is now a live Loading Tracker week" in create_week_html

        with app.app_context():
            imports = db.session.query(LoadingTrackerImport).order_by(LoadingTrackerImport.created_at.asc()).all()
            assert len(imports) == 2
            live_week = imports[-1]
            planned_rows = db.session.query(LoadingTrackerRow).filter_by(tracker_import_id=live_week.id, row_state="planned").all()
            pending_count = db.session.query(LoadingTrackerRow).filter_by(tracker_import_id=live_week.id, row_state="pending").count()
            tues = db.session.query(LoadingTrackerDay).filter_by(tracker_import_id=live_week.id, day_name="Tues").one()
            tues_rows = db.session.query(LoadingTrackerRow).filter_by(day_id=tues.id, row_state="planned").all()

            assert live_week.week_label == "Week 5 from SKU Automator"
            assert len(planned_rows) == 2
            assert {row.source_kind for row in planned_rows} == {"sku_automator"}
            assert {row.store_name for row in tues_rows} == {"Store One", "Store Two"}
            assert pending_count == 1
            tues_items = [
                item
                for row in tues_rows
                for item in db.session.query(LoadingTrackerRowItem)
                .filter_by(row_id=row.id)
                .order_by(LoadingTrackerRowItem.id.asc())
                .all()
            ]
            assert {item.invoice_category for item in tues_items} == {"VT", "NV"}
            assert {item.prefixed_reference_no for item in tues_items} == {
                "VT-17551810",
                "NV-17551811",
                "VT-17551812",
            }
            db.session.remove()
            db.engine.dispose()


def test_sales_order_run_generates_tally_ready_workbook() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()
        uom = Workbook()
        sheet = uom.active
        sheet.title = "UOM"
        sheet.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
        sheet.append(["SKU Alpha", "ctn", "pcs", 12, "Yes", 100])
        sheet.append(["SKU Vanilla", "ctn", "pcs", 12, "No", 120])
        uom_bytes = BytesIO()
        uom.save(uom_bytes)
        uom_bytes.seek(0)

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(uom_bytes.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        upload = client.post(
            "/sales-order/import",
            data={"sales_order_workbook": (BytesIO(build_pepup_orders_workbook().getvalue()), "pepup.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert upload.status_code == 302
        assert "/sales-order/runs/" in upload.headers["Location"]
        assert "/review" not in upload.headers["Location"]

        run_id = upload.headers["Location"].split("/sales-order/runs/")[1]
        page = client.get(upload.headers["Location"])
        html = page.get_data(as_text=True)
        assert page.status_code == 200
        assert "Download Sales Order" in html

        download = client.get(f"/sales-order/runs/{run_id}/download")
        assert download.status_code == 200
        assert download.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        workbook = load_workbook(BytesIO(download.data), data_only=True)
        output = workbook["Sales Order"]
        row2 = [output.cell(2, c).value for c in range(1, 13)]
        row3 = [output.cell(3, c).value for c in range(1, 13)]

        assert row2[1] == "VT-17552475"
        assert row2[4] == "SKU Alpha"
        assert row2[5] == "2ctn"
        assert row2[6] == 100
        assert row2[7] == 200
        assert row2[9] == "VAT"

        assert row3[1] == "NV-17552476"
        assert row3[4] == "SKU Vanilla"
        assert row3[5] == "0.5ctn"
        assert row3[6] == 120
        assert row3[7] == 60
        assert row3[9] is None

        with app.app_context():
            run = db.session.get(SalesOrderRun, run_id)
            lines = db.session.query(SalesOrderLine).filter_by(run_id=run_id).order_by(SalesOrderLine.id.asc()).all()
            assert run is not None
            assert run.rows_ready == 2
            assert run.rows_needing_review == 0
            assert [line.invoice_category for line in lines] == ["VT", "NV"]
            db.session.remove()
            db.engine.dispose()


def test_sales_order_run_splits_same_reference_into_bp_vt_and_nv_lines() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()
        uom = Workbook()
        sheet = uom.active
        sheet.title = "UOM"
        sheet.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
        sheet.append(["SKU Alpha", "ctn", "pcs", 12, "Yes", 100])
        sheet.append(["SKU Beta", "ctn", "pcs", 12, "Yes", 120])
        sheet.append(["SKU Vanilla", "ctn", "pcs", 12, "No", 80])
        uom_bytes = BytesIO()
        uom.save(uom_bytes)
        uom_bytes.seek(0)

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(uom_bytes.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        with app.app_context():
            db.session.add(
                BrandPartnerRule(
                    rule_name="Globus Alpha BP",
                    sku_name_pattern="SKU Alpha",
                    normalized_sku_pattern="SKU ALPHA",
                    store_name_pattern="Globus",
                    normalized_store_pattern="GLOBUS",
                    is_active=True,
                )
            )
            db.session.commit()

        upload = client.post(
            "/sales-order/import",
            data={"sales_order_workbook": (BytesIO(build_mixed_category_pepup_orders_workbook().getvalue()), "pepup.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert upload.status_code == 302
        run_id = upload.headers["Location"].split("/sales-order/runs/")[1]

        download = client.get(f"/sales-order/runs/{run_id}/download")
        assert download.status_code == 200

        workbook = load_workbook(BytesIO(download.data), data_only=True)
        output = workbook["Sales Order"]
        row2 = [output.cell(2, c).value for c in range(1, 13)]
        row3 = [output.cell(3, c).value for c in range(1, 13)]
        row4 = [output.cell(4, c).value for c in range(1, 13)]

        assert row2[1] == "BP-17575653"
        assert row2[4] == "SKU Alpha"
        assert row2[9] == "VAT"

        assert row3[1] == "VT-17575653"
        assert row3[4] == "SKU Beta"
        assert row3[9] == "VAT"

        assert row4[1] == "NV-17575653"
        assert row4[4] == "SKU Vanilla"
        assert row4[9] in ("", None)

        with app.app_context():
            lines = db.session.query(SalesOrderLine).filter_by(run_id=run_id).order_by(SalesOrderLine.id.asc()).all()
            assert [line.invoice_category for line in lines] == ["BP", "VT", "NV"]
            assert [line.prefixed_reference_no for line in lines] == [
                "BP-17575653",
                "VT-17575653",
                "NV-17575653",
            ]
            assert lines[0].classification_source == "bp_rule"
            assert lines[0].bp_rule_reason == "Globus Alpha BP"
            db.session.remove()
            db.engine.dispose()


def test_sales_order_run_accepts_tab_delimited_pepup_export_with_xls_extension() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()
        uom = Workbook()
        sheet = uom.active
        sheet.title = "UOM"
        sheet.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
        sheet.append(["SKU Alpha", "ctn", "pcs", 12, "Yes", 100])
        sheet.append(["SKU Vanilla", "ctn", "pcs", 12, "No", 120])
        uom_bytes = BytesIO()
        uom.save(uom_bytes)
        uom_bytes.seek(0)

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(uom_bytes.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        upload = client.post(
            "/sales-order/import",
            data={"sales_order_workbook": (BytesIO(build_pepup_orders_tab_delimited_export().getvalue()), "Order Item List (17).xls")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert upload.status_code == 302
        assert "/sales-order/runs/" in upload.headers["Location"]
        assert "/review" not in upload.headers["Location"]

        run_id = upload.headers["Location"].split("/sales-order/runs/")[1]
        page = client.get(upload.headers["Location"])
        html = page.get_data(as_text=True)
        assert page.status_code == 200
        assert "Download Sales Order" in html

        download = client.get(f"/sales-order/runs/{run_id}/download")
        assert download.status_code == 200
        workbook = load_workbook(BytesIO(download.data))
        output = workbook["Sales Order"]
        row2 = [output.cell(2, c).value for c in range(1, 13)]
        row3 = [output.cell(3, c).value for c in range(1, 13)]

        assert row2[0].date().isoformat() == "2026-03-12"
        assert row2[1] == "VT-17575653"
        assert row2[4] == "SKU Alpha"
        assert row2[5] == "2ctn"
        assert row2[6] == 93.02
        assert row2[7] == 186.04
        assert output["A2"].number_format == "yyyy-mm-dd"
        assert output.column_dimensions["A"].width >= 16

        assert row3[0].date().isoformat() == "2026-03-12"
        assert row3[1] == "NV-17575278"
        assert row3[4] == "SKU Vanilla"
        assert row3[5] == "0.5ctn"
        assert row3[7] == 60

        with app.app_context():
            run = db.session.get(SalesOrderRun, run_id)
            assert run is not None
            assert run.rows_ready == 2
            assert run.rows_needing_review == 0
            db.session.remove()
            db.engine.dispose()


def test_sku_automator_run_generates_register_and_matrix() -> None:
    with TemporaryDirectory() as temp_dir:
        app = create_app(
            {
                "TESTING": True,
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{Path(temp_dir) / 'test.db'}",
                "APP_TIMEZONE": "Africa/Lagos",
            }
        )

        client = app.test_client()
        uom = Workbook()
        sheet = uom.active
        sheet.title = "UOM"
        sheet.append(["ITEM", "UOM", "ALT UOM", "Conversion", "Vatable", "Prices"])
        sheet.append(["SKU Alpha", "ctn", "pcs", 12, "Yes", 100])
        sheet.append(["SKU Vanilla", "ctn", "pcs", 12, "No", 120])
        uom_bytes = BytesIO()
        uom.save(uom_bytes)
        uom_bytes.seek(0)

        response = client.post(
            "/uom/import",
            data={"uom_workbook": (BytesIO(uom_bytes.getvalue()), "uom.xlsx")},
            content_type="multipart/form-data",
        )
        assert response.status_code == 302

        upload = client.post(
            "/sku-automator/import",
            data={"sku_automator_workbook": (BytesIO(build_tally_export_workbook().getvalue()), "salesordertest.xls")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        assert upload.status_code == 302
        assert "/sku-automator/runs/" in upload.headers["Location"]
        assert "/review" not in upload.headers["Location"]

        run_id = upload.headers["Location"].split("/sku-automator/runs/")[1]
        page = client.get(upload.headers["Location"])
        html = page.get_data(as_text=True)
        assert page.status_code == 200
        assert "Download planner output" in html

        download = client.get(f"/sku-automator/runs/{run_id}/download")
        assert download.status_code == 200
        assert download.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        workbook = load_workbook(BytesIO(download.data), data_only=True)
        register = workbook["Sales Order- Dala Register"]
        matrix = workbook["Store SKU Matrix"]

        register_row2 = [register.cell(2, c).value for c in range(1, 9)]
        register_row3 = [register.cell(3, c).value for c in range(1, 9)]
        register_row4 = [register.cell(4, c).value for c in range(1, 9)]

        assert register_row2[1] == "Store One"
        assert register_row2[2] == "SKU Alpha"
        assert register_row2[3] == 2
        assert register_row2[5] == 200
        assert register_row2[6] == "VT-17551810"

        assert register_row3[2] == "SKU Vanilla"
        assert register_row3[3] == 0.5
        assert register_row3[5] == 60
        assert register_row3[6] == "NV-17551811"

        assert register_row4[1] == "Store Two"
        assert register_row4[2] == "SKU Alpha"
        assert register_row4[3] == 1
        assert register_row4[5] == 100
        assert register_row4[6] == "VT-17551812"

        matrix_headers = [matrix.cell(1, c).value for c in range(1, 4)]
        matrix_row2 = [matrix.cell(2, c).value for c in range(1, 4)]
        matrix_row3 = [matrix.cell(3, c).value for c in range(1, 4)]

        assert matrix_headers == ["Stores", "SKU Alpha", "SKU Vanilla"]
        assert matrix_row2 == ["Store One", 2, 0.5]
        assert matrix_row3 == ["Store Two", 1, 0]

        with app.app_context():
            run = db.session.get(SkuAutomatorRun, run_id)
            lines = db.session.query(SkuAutomatorLine).filter_by(run_id=run_id).order_by(SkuAutomatorLine.id.asc()).all()
            assert run is not None
            assert run.rows_ready == 3
            assert run.store_count == 2
            assert [line.invoice_category for line in lines] == ["VT", "NV", "VT"]
            db.session.remove()
            db.engine.dispose()
