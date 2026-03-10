from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import xlwt
from openpyxl import load_workbook

TRACKER_SHEET = "tracker"
UOM_SHEET = "UOM"
TEMPLATE_SHEET = "Delivery Invoice"
VOUCHER_TYPE_NAME = "Delivery Invoice"
SALES_LEDGER_NAME = "Inventory Pool"
VAT_LABEL = "VAT"
VAT_RATE = Decimal("7.5")
DATE_FORMAT = "%Y-%m-%d"


class ConverterError(Exception):
    pass


class MissingSheetError(ConverterError):
    pass


class MissingPriceError(ConverterError):
    def __init__(self, missing_items: list[str]) -> None:
        self.missing_items = missing_items
        super().__init__("Some SKUs are missing prices.")


@dataclass
class WorkbookInspection:
    output_headers: list[str]
    rows_to_export: int
    order_count: int
    supermarket_count: int
    missing_items: list[str]
    invoice_date: str


@dataclass
class ConversionResult:
    output_headers: list[str]
    rows: list[list[Any]]
    rows_to_export: int
    order_count: int
    supermarket_count: int
    invoice_date: str
    corrected_items: list[str]


def inspect_workbook(workbook_path: Path, timezone_name: str) -> WorkbookInspection:
    workbook = load_workbook(workbook_path, data_only=True)
    tracker = _require_sheet(workbook, TRACKER_SHEET)
    uom = _require_sheet(workbook, UOM_SHEET)
    template = _require_sheet(workbook, TEMPLATE_SHEET)

    product_headers = _tracker_product_headers(tracker)
    uom_lookup = _build_uom_lookup(uom)
    missing_items: set[str] = set()
    order_numbers: set[str] = set()
    supermarkets: set[str] = set()
    rows_to_export = 0

    for row_index in range(2, tracker.max_row + 1):
        order_number = _string_value(tracker.cell(row_index, 1).value)
        supermarket = _string_value(tracker.cell(row_index, 2).value)
        if not order_number and not supermarket:
            continue

        if order_number:
            order_numbers.add(order_number)
        if supermarket:
            supermarkets.add(supermarket)

        for column_index, sku in product_headers:
            quantity = _decimal_quantity(tracker.cell(row_index, column_index).value)
            if quantity is None or quantity <= Decimal("0"):
                continue

            rows_to_export += 1
            item = uom_lookup.get(sku)
            if item is None or item.price is None:
                missing_items.add(sku)

    output_headers = _template_headers(template)
    invoice_date = _tomorrow_in_timezone(timezone_name).strftime(DATE_FORMAT)

    return WorkbookInspection(
        output_headers=output_headers,
        rows_to_export=rows_to_export,
        order_count=len(order_numbers),
        supermarket_count=len(supermarkets),
        missing_items=sorted(missing_items),
        invoice_date=invoice_date,
    )


def convert_workbook(
    workbook_path: Path,
    timezone_name: str,
    price_overrides: dict[str, Decimal] | None = None,
) -> ConversionResult:
    workbook = load_workbook(workbook_path, data_only=True)
    tracker = _require_sheet(workbook, TRACKER_SHEET)
    uom = _require_sheet(workbook, UOM_SHEET)
    template = _require_sheet(workbook, TEMPLATE_SHEET)

    product_headers = _tracker_product_headers(tracker)
    uom_lookup = _build_uom_lookup(uom)
    output_headers = _template_headers(template)
    invoice_date = _tomorrow_in_timezone(timezone_name).strftime(DATE_FORMAT)
    overrides = price_overrides or {}

    order_numbers: set[str] = set()
    supermarkets: set[str] = set()
    missing_items: set[str] = set()
    corrected_items: set[str] = set()
    rows: list[list[Any]] = []

    for row_index in range(2, tracker.max_row + 1):
        order_number = _string_value(tracker.cell(row_index, 1).value)
        supermarket = _string_value(tracker.cell(row_index, 2).value)
        if not order_number and not supermarket:
            continue

        if order_number:
            order_numbers.add(order_number)
        if supermarket:
            supermarkets.add(supermarket)

        for column_index, sku in product_headers:
            quantity = _decimal_quantity(tracker.cell(row_index, column_index).value)
            if quantity is None or quantity <= Decimal("0"):
                continue

            item = uom_lookup.get(sku)
            rate = item.price if item else None
            if rate is None and sku in overrides:
                rate = overrides[sku]
                corrected_items.add(sku)

            if rate is None:
                missing_items.add(sku)
                continue

            amount = quantity * rate
            is_vatable = item.vatable if item else False
            vat_amount = (amount * VAT_RATE / Decimal("100")).quantize(Decimal("0.01")) if is_vatable else ""

            rows.append(
                [
                    invoice_date,
                    order_number,
                    VOUCHER_TYPE_NAME,
                    supermarket,
                    sku,
                    float(quantity),
                    float(rate),
                    float(amount),
                    SALES_LEDGER_NAME,
                    VAT_LABEL if is_vatable else "",
                    float(VAT_RATE) if is_vatable else "",
                    float(vat_amount) if vat_amount != "" else "",
                ]
            )

    if missing_items:
        raise MissingPriceError(sorted(missing_items))

    return ConversionResult(
        output_headers=output_headers,
        rows=rows,
        rows_to_export=len(rows),
        order_count=len(order_numbers),
        supermarket_count=len(supermarkets),
        invoice_date=invoice_date,
        corrected_items=sorted(corrected_items),
    )


def write_xls(result: ConversionResult, output_path: Path) -> None:
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet(TEMPLATE_SHEET)

    header_style = xlwt.easyxf(
        "font: bold on, height 240;"
        "pattern: pattern solid, fore_colour ocean_blue;"
        "align: horiz center;"
        "borders: left thin, right thin, top thin, bottom thin;"
    )
    text_style = xlwt.easyxf("borders: left thin, right thin, top thin, bottom thin;")
    date_style = xlwt.easyxf(
        "borders: left thin, right thin, top thin, bottom thin;",
        num_format_str="yyyy-mm-dd",
    )
    decimal_style = xlwt.easyxf(
        "borders: left thin, right thin, top thin, bottom thin;",
        num_format_str="0.00##",
    )

    for col_index, header in enumerate(result.output_headers):
        sheet.write(0, col_index, header, header_style)
        sheet.col(col_index).width = _column_width(header)

    for row_index, row in enumerate(result.rows, start=1):
        for col_index, value in enumerate(row):
            if col_index == 0 and value:
                parsed = datetime.strptime(str(value), DATE_FORMAT)
                sheet.write(row_index, col_index, parsed, date_style)
            elif isinstance(value, (int, float)):
                sheet.write(row_index, col_index, value, decimal_style)
            else:
                sheet.write(row_index, col_index, value, text_style)

    workbook.save(str(output_path))


@dataclass
class UomItem:
    price: Decimal | None
    vatable: bool


def _require_sheet(workbook: Any, sheet_name: str) -> Any:
    if sheet_name not in workbook.sheetnames:
        raise MissingSheetError(f"Sheet '{sheet_name}' is missing.")
    return workbook[sheet_name]


def _tracker_product_headers(tracker_sheet: Any) -> list[tuple[int, str]]:
    headers: list[tuple[int, str]] = []
    for column_index in range(3, tracker_sheet.max_column + 1):
        header = _string_value(tracker_sheet.cell(1, column_index).value)
        if header:
            headers.append((column_index, header))
    return headers


def _template_headers(template_sheet: Any) -> list[str]:
    headers: list[str] = []
    for column_index in range(1, template_sheet.max_column + 1):
        header = _string_value(template_sheet.cell(1, column_index).value)
        if header:
            headers.append(header)
    return headers


def _build_uom_lookup(uom_sheet: Any) -> dict[str, UomItem]:
    lookup: dict[str, UomItem] = {}
    for row_index in range(2, uom_sheet.max_row + 1):
        sku = _string_value(uom_sheet.cell(row_index, 1).value)
        if not sku:
            continue

        lookup[sku] = UomItem(
            price=_decimal_quantity(uom_sheet.cell(row_index, 6).value),
            vatable=_string_value(uom_sheet.cell(row_index, 5).value).lower() == "yes",
        )
    return lookup


def _string_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _decimal_quantity(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None

    try:
        return Decimal(str(value)).normalize()
    except (InvalidOperation, ValueError):
        return None


def _tomorrow_in_timezone(timezone_name: str) -> datetime:
    now = datetime.now(ZoneInfo(timezone_name))
    return now + timedelta(days=1)


def _column_width(header: str) -> int:
    length = max(len(header) + 3, 14)
    return min(length * 256, 35 * 256)
